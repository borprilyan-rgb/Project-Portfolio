#region --- LIBRARY AND SUCH ---
import streamlit as st
import pandas as pd
import altair as alt
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import plotly.graph_objects as go
import num2words as n2w
import ast
import numpy as np
import textwrap
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import json as _json
import os
import tempfile

APP_VERSION = "1.1.0" #app version for future compatibility check
st.set_page_config(page_title="Project Feasibility Study - Agung Sedayu Group",
                    layout="wide", page_icon="Agung-Sedayu.png",)

st.logo("Agung-Sedayu-Group.png")
#endregion

#region 
import copy

# ==================================================
# CENTRAL APP CONFIG DEFAULTS
# ==================================================
DEFAULT_REPORT_CONFIG = {
    "port_meta": {
        "title": "PROJECT PORTFOLIO | PIK2.D2.GINZA.MIDTOWN OPT.2 R(1)",
        "ref": "REF. DATA R(0) | CONCEPT DWG 2026-02-02.DPA",
        "version": "R (1) OPT2",
        "updated": "02-02-2026",
        "created": "02-02-2026"
    },
    "export_settings": {
        "prepared_by": "",
        "checked_by": ""
    },
    "port_assumptions": [
        {"No.": str(i), "Assumption Description": desc}
        for i, desc in enumerate([
            "Include Vacuum Project + Urugan kembali asumsi 1m",
            "Foundation System standard pilecaps.",
            "No Basement and No Parking Podium.",
            "Parking provison limited to ON STREET LEVEL parking; Floor Hardener finish",
            "Floor to Floor Height at 3.5M",
            "Facade Alumunium Window Wall - + Grill Outdoor AC",
            "External Façade Precast, No double skin for parking podium if any.",
            "Ground Lobby Finishes completed with Artificial stone & HT.",
            "Typical Corridor | Floor finishes : HT | Wall Finishes : Cement Sand Plaster c/w Emulsion Paint.",
            "Aircon System | Apartement : AC Split | Hotel : VRF SYSTEM",
            "SBO Rebars @ Rp. 10.000/kg",
            "Excluded Smarthome",
            "Lift : Luxury Apartment : 8 Private Lift + 2 Services Lift | Hotel 3* : 3 Passenger Lift + 1 Services Lift\nTerrace Village : 16 Private Lift + 8 Services Lift | Retail : No Elevator + Escalator 12 units\nApartment 2 : 4 Passenger Lift + 2 Services Lift\nPodium Village : 10 Private Lift + 5 Services Lift",
            "Exclude Wardrobe",
            "FFE : Kitchen cabinet, Hob & Hood, Refrigerator & Washing Machine",
            "Water Heater : Installation only",
            "Based on Resume Calculation DP dated on 2026.02.02"
        ], 1)
    ],
}

def _safe_float(v, default=0.0):
    if v is None:
        return default

    try:
        if pd.isna(v):
            return default
    except Exception:
        pass

    if isinstance(v, str):
        v = (
            v.strip()
            .replace("Rp", "")
            .replace("rp", "")
            .replace(",", "")
            .replace("%", "")
        )

        if v == "" or v.lower() in ["none", "nan", "null", "-"]:
            return default

    try:
        return float(v)
    except Exception:
        return default

def calculate_area_totals_from_table(area_table_data):
    unit_area_col = "Unit Area"
    breakdown_cols = [
        "Parkir", "Roof/Deck", "MEP Outdoor",
        "Koridor/Lobby", "Stair, MEP, Etc",
        unit_area_col, "Office"
    ]

    if not isinstance(area_table_data, list) or len(area_table_data) == 0:
        return {
            "gba": 0.0,
            "gfa": 0.0,
            "sgfa": 0.0,
            "nfa": 0.0
        }

    df = pd.DataFrame(area_table_data)

    if unit_area_col not in df.columns and "Unit" in df.columns:
        df[unit_area_col] = df["Unit"]

    for col in breakdown_cols:
        if col not in df.columns:
            df[col] = 0.0

        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    total = df[breakdown_cols].sum(axis=1)

    gba = _safe_float(total.sum())
    gfa = _safe_float((total - df[["Parkir", "Roof/Deck", "MEP Outdoor"]].sum(axis=1)).sum())
    sgfa = _safe_float(df[[unit_area_col, "Office", "Koridor/Lobby"]].sum(axis=1).sum())
    nfa = _safe_float(df[[unit_area_col, "Office"]].sum(axis=1).sum())

    return {
        "gba": gba,
        "gfa": gfa,
        "sgfa": sgfa,
        "nfa": nfa
    }

def clean_for_json(obj):
    """
    Convert Streamlit / pandas / numpy objects into JSON-safe Python objects.
    This prevents save_data() from crashing when session_state contains
    numpy, pandas, NaN, inf, Timestamp, DataFrame, Series, etc.
    """

    import math
    import datetime
    from decimal import Decimal

    # None
    if obj is None:
        return None

    # Pandas DataFrame / Series
    if isinstance(obj, pd.DataFrame):
        return clean_for_json(obj.to_dict(orient="records"))

    if isinstance(obj, pd.Series):
        return clean_for_json(obj.to_dict())

    # Pandas / datetime objects
    if isinstance(obj, (pd.Timestamp, datetime.datetime, datetime.date)):
        return obj.isoformat()

    # Numpy scalar types
    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        val = float(obj)
        if math.isnan(val) or math.isinf(val):
            return None
        return val

    if isinstance(obj, np.bool_):
        return bool(obj)

    # Normal Python scalar types
    if isinstance(obj, bool):
        return obj

    if isinstance(obj, int):
        return obj

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    if isinstance(obj, Decimal):
        val = float(obj)
        if math.isnan(val) or math.isinf(val):
            return None
        return val

    if isinstance(obj, str):
        return obj

    # Dictionary
    if isinstance(obj, dict):
        return {
            str(k): clean_for_json(v)
            for k, v in obj.items()
        }

    # List / tuple / set
    if isinstance(obj, (list, tuple, set)):
        return [clean_for_json(v) for v in obj]

    # Fallback for unknown objects
    try:
        return str(obj)
    except Exception:
        return None

def init_report_config():
    """
    Creates one global/report-level config bucket.
    Old port_meta / port_assumptions are only migrated ONCE
    when report_config does not exist yet.
    """
    if "report_config" not in st.session_state:
        st.session_state.report_config = copy.deepcopy(DEFAULT_REPORT_CONFIG)

        # One-time migration from old keys
        if "port_meta" in st.session_state:
            st.session_state.report_config["port_meta"] = st.session_state.port_meta

        if "port_assumptions" in st.session_state:
            if isinstance(st.session_state.port_assumptions, pd.DataFrame):
                st.session_state.report_config["port_assumptions"] = (
                    st.session_state.port_assumptions.to_dict("records")
                )
            else:
                st.session_state.report_config["port_assumptions"] = st.session_state.port_assumptions

    st.session_state.report_config.setdefault(
        "port_meta",
        copy.deepcopy(DEFAULT_REPORT_CONFIG["port_meta"])
    )

    st.session_state.report_config.setdefault(
        "export_settings",
        copy.deepcopy(DEFAULT_REPORT_CONFIG["export_settings"])
    )

    st.session_state.report_config.setdefault(
        "port_assumptions",
        copy.deepcopy(DEFAULT_REPORT_CONFIG["port_assumptions"])
    )

def get_report_config():
    init_report_config()
    return st.session_state.report_config


def get_port_meta():
    cfg = get_report_config()
    cfg.setdefault("port_meta", copy.deepcopy(DEFAULT_REPORT_CONFIG["port_meta"]))
    return cfg["port_meta"]


def get_port_assumptions_df():
    cfg = get_report_config()
    cfg.setdefault(
        "port_assumptions",
        copy.deepcopy(DEFAULT_REPORT_CONFIG["port_assumptions"])
    )

    assumptions = cfg["port_assumptions"]

    if isinstance(assumptions, pd.DataFrame):
        return assumptions.copy()

    return pd.DataFrame(assumptions)


def set_port_assumptions_df(df):
    cfg = get_report_config()
    cfg["port_assumptions"] = df.to_dict("records")

def build_app_payload():
    """
    Single source of truth for saving.
    save_data() and save_snapshot() should both use this.
    """
    init_report_config()

    curr_id, _ = repair_projects_state(save=False)

    payload = {
        "app_version": APP_VERSION,
        "projects": st.session_state.get("projects", make_default_projects()),
        "current_proj_id": curr_id,
        "proj_counter": st.session_state.get("proj_counter", 1),

        "loaded_snapshot_id": st.session_state.get("loaded_snapshot_id"),
        "loaded_snapshot_name": st.session_state.get("loaded_snapshot_name"),

        "report_config": st.session_state.get(
            "report_config",
            copy.deepcopy(DEFAULT_REPORT_CONFIG)
        )
    }

    return clean_for_json(payload)

def make_default_projects():
    return {
        "proj_1": {
            "name": "New Project 1",
            "type": "Hotel",
            "data": {}
        }
    }


def repair_projects_state(save=False):
    projects = st.session_state.get("projects", {})

    # Guard 1: projects must be a non-empty dict
    if not isinstance(projects, dict) or len(projects) == 0:
        st.session_state.projects = make_default_projects()
        st.session_state.current_proj_id = "proj_1"
        st.session_state.proj_counter = 1
        if save:
            save_data()
        return "proj_1", st.session_state.projects["proj_1"]

    # Guard 2: remove any corrupt project entries (non-dict values)
    corrupt_keys = [k for k, v in projects.items() if not isinstance(v, dict)]
    for k in corrupt_keys:
        del projects[k]

    # If all entries were corrupt, rebuild from scratch
    if len(projects) == 0:
        st.session_state.projects = make_default_projects()
        st.session_state.current_proj_id = "proj_1"
        st.session_state.proj_counter = 1
        if save:
            save_data()
        return "proj_1", st.session_state.projects["proj_1"]

    # Guard 3: current_proj_id must be a valid string key in projects
    curr_id = st.session_state.get("current_proj_id")

    if not isinstance(curr_id, str) or curr_id not in projects:
        curr_id = list(projects.keys())[0]
        st.session_state.current_proj_id = curr_id
        if save:
            save_data()

    return curr_id, projects[curr_id]

def get_current_project():
    return repair_projects_state(save=True)

def restore_app_payload(data):
    """
    Single source of truth for loading.
    Snapshot load and startup load should both use this.
    """
    if not data:
        data = {}

    projects = data.get("projects", make_default_projects())

    # Critical guard: prevent empty project dictionary
    if not isinstance(projects, dict) or len(projects) == 0:
        projects = make_default_projects()

    st.session_state.projects = projects

    saved_curr_id = data.get("current_proj_id")

    if isinstance(saved_curr_id, str) and saved_curr_id in st.session_state.projects:
        st.session_state.current_proj_id = saved_curr_id
    else:
        st.session_state.current_proj_id = list(st.session_state.projects.keys())[0]

    st.session_state.proj_counter = data.get(
        "proj_counter",
        len(st.session_state.projects)
    )

    # New format
    report_config = copy.deepcopy(DEFAULT_REPORT_CONFIG)
    report_config.update(data.get("report_config", {}))

    # Backward compatibility with old snapshots
    if "port_meta" in data:
        report_config["port_meta"] = data["port_meta"]

    if "port_assumptions" in data:
        report_config["port_assumptions"] = data["port_assumptions"]

    st.session_state.report_config = report_config

    # Active archive reference
    st.session_state.loaded_snapshot_id = data.get("loaded_snapshot_id")
    st.session_state.loaded_snapshot_name = data.get("loaded_snapshot_name")

    # Optional backward compatibility aliases
    st.session_state.port_meta = st.session_state.report_config["port_meta"]
    st.session_state.port_assumptions = pd.DataFrame(
        st.session_state.report_config["port_assumptions"]
    )
    #endregion 

#region --- DO NOT CHANGE (OR I WILL KICK YOUR BUTT)---
from supabase import create_client, Client
url: str = st.secrets["SUPABASE_URL"]
key: str = st.secrets["SUPABASE_KEY"]
supabase: Client = create_client(url, key)


def calculate_project_totals(pdata, curr_type):
    """Calculates all totals dynamically for a given project."""
    d = pdata.get("data", {})
    # Default to empty dict if project type isn't in database yet
    pt_data = PROJECT_DATABASE.get(curr_type, {})

    def get_val(key, default=0.0):
        val = d.get(key, default)
        if isinstance(val, list): return val
        try: return _safe_float(val)
        except: return val

    # --- Area Calculations ---
    area_table = get_val("area_table_data", [])
    if isinstance(area_table, list) and len(area_table) > 0:
        calc_gba = 0.0
        calc_gfa = 0.0
        calc_sgfa = 0.0
        unit_area_col = "Unit Area"
        breakdown_cols = ["Parkir", "Roof/Deck", "MEP Outdoor", "Koridor/Lobby", "Stair, MEP, Etc", unit_area_col, "Office"]
        
        for row in area_table:
            row_total = sum(_safe_float(row.get(c, 0)) for c in breakdown_cols)
            if unit_area_col not in row and "Unit" in row:
                row_total += _safe_float(row.get("Unit", 0))
            calc_gba += row_total
            calc_gfa += row_total - sum(_safe_float(row.get(c, 0)) for c in ["Parkir", "Roof/Deck", "MEP Outdoor"])
            unit_area = _safe_float(row.get(unit_area_col, row.get("Unit", 0)))
            calc_sgfa += unit_area + sum(_safe_float(row.get(c, 0)) for c in ["Office", "Koridor/Lobby"])
    else:
        calc_gba = get_val("m_gba", 0.0)
        calc_gfa = get_val("m_gfa", 0.0)
        calc_sgfa = get_val("m_sgfa", 0.0)

    # --- Cost Calculations ---
    struc_earth = get_val("u_earth", pt_data.get("struc_earth", 0))
    struc_found = get_val("u_found", pt_data.get("struc_found", 0))
    struc_work = get_val("u_struc", pt_data.get("struc_work", 0))
    arch_base = get_val("u_arch", pt_data.get("arch_base", 0))
    
    facade = get_val("m_facade", 0.0)
    fac_precast_pct = get_val("r_fac_pre", pt_data.get("facade_precast_pct", 0))
    fac_precast_rate = get_val("u_f_pre", pt_data.get("facade_precast_rate", 0))
    
    rooms = get_val("m_rooms", 0.0)
    ffe_rate = get_val("u_ffe", pt_data.get("ffe", 0))
    mep_rate = get_val("u_mep", pt_data.get("mep", 0))
    utility_rate = get_val("u_util", pt_data.get("utility", 0))
    
    consultancy_rate = get_val("sc_cons", pt_data.get("cons", 0))
    insurance_pct = get_val("sc_ins", 0.12)
    
    smart_custom_costs = sum(_safe_float(item.get("Rate (Rp)", 0)) * _safe_float(item.get("Quantity", 1)) for item in get_val("smart_custom_costs", []))

    t_earth = calc_gba * struc_earth
    t_found = calc_gba * struc_found
    t_struc = calc_gba * struc_work
    t_arch_base = calc_gfa * arch_base
    t_precast = facade * (fac_precast_pct / 100) * fac_precast_rate
    t_ffe = rooms * ffe_rate
    t_mep = calc_gba * mep_rate
    t_utility = calc_gba * utility_rate
    
    construction_subtotal = sum([t_earth, t_found, t_struc, t_arch_base, t_precast, t_ffe, t_mep, t_utility, smart_custom_costs])

    t_preliminary = construction_subtotal * 0.05
    t_contingency = (construction_subtotal + t_preliminary) * 0.03
    grand_total_hc = construction_subtotal + t_preliminary + t_contingency

    total_soft_cost = (calc_gfa * consultancy_rate) + (grand_total_hc * (insurance_pct / 100.0)) 
    
    calc_budget = grand_total_hc + total_soft_cost

    return calc_gba, calc_gfa, calc_sgfa, calc_budget, rooms

from datetime import datetime, timedelta

def _get_authed_snapshot_client(show_error=True):
    token = st.session_state.get("access_token")
    user = st.session_state.get("user")

    user_id = getattr(user, "id", None)
    if user_id is None and isinstance(user, dict):
        user_id = user.get("id")

    if not token or not user_id:
        if show_error:
            st.error("Not authenticated.")
        return None, None

    authed_client = create_client(url, key)
    authed_client.postgrest.auth(token)

    return authed_client, user_id

def format_snapshot_time(created_at):
    if not created_at:
        return ""

    try:
        created_utc = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        created_local = created_utc + timedelta(hours=7)
        return created_local.strftime("%d %b %Y, %H:%M WIB")
    except Exception:
        return ""

def save_snapshot(snapshot_name):
    authed_client, user_id = _get_authed_snapshot_client()

    if not authed_client or not user_id:
        return False

    payload = build_app_payload()

    try:
        response = authed_client.table("project_snapshots").insert({
            "user_id": user_id,
            "snapshot_name": snapshot_name,
            "data": payload
        }).execute()

        if response.data:
            st.session_state.loaded_snapshot_id = response.data[0].get("id")
            st.session_state.loaded_snapshot_name = response.data[0].get("snapshot_name", snapshot_name)

        save_data()
        return True

    except Exception as e:
        st.error(f"Project Save Error: {e}")
        return False

def overwrite_current_snapshot():
    snapshot_id = st.session_state.get("loaded_snapshot_id")
    snapshot_name = st.session_state.get("loaded_snapshot_name")

    if not snapshot_id:
        st.error("No archive is currently linked. Please save this project first from Archive.")
        return False

    token = st.session_state.get("access_token")
    user = st.session_state.get("user")
    user_id = getattr(user, "id", None)

    if not token or not user_id:
        st.error("Not authenticated.")
        return False

    authed_client = create_client(url, key)
    authed_client.postgrest.auth(token)

    payload = build_app_payload()

    try:
        authed_client.table("project_snapshots") \
            .update({
                "snapshot_name": snapshot_name,
                "data": payload
            }) \
            .eq("id", snapshot_id) \
            .eq("user_id", user_id) \
            .execute()

        return True

    except Exception as e:
        st.error(f"Project Overwrite Error: {e}")
        return False

def rename_snapshot(snapshot_id, new_name):
    token = st.session_state.get("access_token")
    user = st.session_state.get("user")
    user_id = getattr(user, "id", None)

    if not token or not user_id:
        st.error("Not authenticated.")
        return False

    if not snapshot_id or not str(new_name).strip():
        st.error("Invalid project name.")
        return False

    authed_client = create_client(url, key)
    authed_client.postgrest.auth(token)

    clean_name = str(new_name).strip()

    try:
        authed_client.table("project_snapshots") \
            .update({
                "snapshot_name": clean_name
            }) \
            .eq("id", snapshot_id) \
            .eq("user_id", user_id) \
            .execute()

        if st.session_state.get("loaded_snapshot_id") == snapshot_id:
            st.session_state.loaded_snapshot_name = clean_name
            save_data()

        return True

    except Exception as e:
        st.error(f"Project Rename Error: {e}")
        return False

def overwrite_snapshot(snapshot_id, snapshot_name=None):
    authed_client, user_id = _get_authed_snapshot_client()

    if not authed_client or not user_id:
        return False

    if not snapshot_id:
        st.error("No saved project selected.")
        return False

    payload = build_app_payload()

    update_payload = {
        "data": payload
    }

    # Optional rename while overwriting
    if snapshot_name is not None and str(snapshot_name).strip() != "":
        update_payload["snapshot_name"] = str(snapshot_name).strip()

    try:
        authed_client.table("project_snapshots") \
            .update(update_payload) \
            .eq("id", snapshot_id) \
            .eq("user_id", user_id) \
            .execute()

        st.session_state.loaded_snapshot_id = snapshot_id

        if snapshot_name is not None and str(snapshot_name).strip() != "":
            st.session_state.loaded_snapshot_name = str(snapshot_name).strip()

        return True

    except Exception as e:
        st.error(f"Project Overwrite Error: {e}")
        return False

def load_snapshots():
    authed_client, user_id = _get_authed_snapshot_client(show_error=False)

    if not authed_client or not user_id:
        return []

    try:
        response = authed_client.table("project_snapshots") \
            .select("id, snapshot_name, created_at") \
            .eq("user_id", user_id) \
            .order("created_at", desc=True) \
            .execute()

        return response.data if response.data else []

    except Exception as e:
        st.error(f"Project Load Error: {e}")
        return []

def load_snapshot_data(snapshot_id):
    authed_client, user_id = _get_authed_snapshot_client()

    if not authed_client or not user_id:
        return None

    try:
        response = authed_client.table("project_snapshots") \
            .select("data, snapshot_name") \
            .eq("id", snapshot_id) \
            .eq("user_id", user_id) \
            .execute()

        if response.data:
            st.session_state.loaded_snapshot_id = snapshot_id
            st.session_state.loaded_snapshot_name = response.data[0].get("snapshot_name", "")
            return response.data[0]["data"]

    except Exception as e:
        st.error(f"Saved Project Fetch Error: {e}")

    return None

def delete_snapshot(snapshot_id):
    authed_client, user_id = _get_authed_snapshot_client()

    if not authed_client or not user_id:
        return False

    try:
        authed_client.table("project_snapshots") \
            .delete() \
            .eq("id", snapshot_id) \
            .eq("user_id", user_id) \
            .execute()

        if st.session_state.get("loaded_snapshot_id") == snapshot_id:
            st.session_state.loaded_snapshot_id = None
            st.session_state.loaded_snapshot_name = None

        return True

    except Exception as e:
        st.error(f"Project Delete Error: {e}")
        return False

def save_data():
    if not st.session_state.get("storage_loaded", False):
        return
    if not st.session_state.get("logged_in", False):
        return

    token = st.session_state.get("access_token")
    user = st.session_state.get("user")
    user_id = getattr(user, "id", None)

    if not token or not user_id:
        st.error("Not authenticated.")
        return False

    authed_client = create_client(url, key)
    authed_client.postgrest.auth(token)

    payload = build_app_payload()

    try:
        authed_client.table("project_storage").upsert({
            "id": f"storage_{user_id}",  # unique row per user
            "user_id": user_id,
            "data": payload
        }).execute()
    except Exception as e:
        st.error(f"Cloud Save Error: {e}")

def load_data():
    token = st.session_state.get("access_token")
    user = st.session_state.get("user")

    if not token or not user:
        return None

    try:
        authed_client = create_client(url, key)
        authed_client.postgrest.auth(token)

        response = authed_client.table("project_storage") \
            .select("data") \
            .eq("id", f"storage_{user.id}") \
            .execute()

        if response.data:
            return response.data[0]["data"]
    except Exception as e:
        st.error(f"Cloud Load Error: {e}")
    return None

def ensure_app_state_loaded():
    """
    Load cloud state only after login.
    This prevents Streamlit from creating default projects before user authentication.
    """
    if st.session_state.get("storage_loaded", False):
        repair_projects_state(save=False)
        return

    stored_data = load_data()

    if stored_data:
        restore_app_payload(stored_data)
    else:
        restore_app_payload({
            "app_version": APP_VERSION,
            "projects": make_default_projects(),
            "current_proj_id": "proj_1",
            "proj_counter": 1,
            "report_config": copy.deepcopy(DEFAULT_REPORT_CONFIG)
        })

    # Repair after load in case old cloud data contains broken/empty projects
    repair_projects_state(save=False)

    st.session_state.storage_loaded = True

    # Save repaired state back to cloud
    save_data()

def n2w(amount):
    try:
        amount = _safe_float(amount)
        if amount >= 1_000_000_000_000: # Trillion
            return f"{amount / 1_000_000_000_000:,.2f} Triliun"
        elif amount >= 1_000_000_000: # Billion (Miliar)
            return f"{amount / 1_000_000_000:,.2f} Miliar"
        elif amount >= 1_000_000: # Million (Juta)
            return f"{amount / 1_000_000:,.2f} Juta"
        else:
            return f"{amount:,.0f}"
    except:
        return "0"
#endregion

PROJECT_DATABASE = { #Change only when asked
    "Apartment": {
        #Foundation & Structure
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        #Architecture
        "arch_base": 1058000.0, "lobby": 1500000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 2500000.0,
        #Pintu & Hardware
        "door_wood": 3500000.0, "door_steel": 7000000.0,
        "hw_wood": 750000.0, "hw_steel": 1850000.0, "door_glass": 1000000.0, 
        #Sanitari
        "san_room_rate": 26875000.0, "san_pub_f": 98075000.0, "san_pub_m": 77050000.0,
        "san_dis": 30275000.0, "san_mushola": 36500000.0,
        #Lantai, Finishing & Interior
        "fl_waste" : 10.0,
        "fl_ht_rate": {"Type1": 150000.0, "Type2": 350000.0},
        "fl_vinyl_rate": {"Type1": 500000.0, "Type2": 750000.0},
        "fl_marmer_rate": {"Type1": 750000.0, "Type2": 1500000.0},

        "gondola": 600000000.0, "carpet": 1200000.0, "glass": 700000.0,
        
        "ffe": 32000000.0, "misc": 32000000.0, "kitchen": 0.0,

        "facade_precast_pct": 10.0, "facade_window_pct": 80.0, "facade_double_pct": 10.0,
        "fl_ht_pct": 90.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 10.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "mep": 4000000.0, "utility": 150000.0,
        "railing_rate": 2200000.0, "skylight_rate": 4500000.0,
         "ext_land": 1563000.0,
        "fac_pub": 31000000.0, "fac_res": 10000000.0, "fac_proj": 2000000000.0,
        "cons": 174000.0
    },
    "Hotel": {
        "arch_base": 1079000.0, "door_wood": 4750000.0, "door_steel": 8000000.0,
        "hw_wood": 8250000.0, "hw_steel": 2850000.0, "lobby": 2000000.0,
        "gondola": 2000000000.0, "carpet": 1200000.0, "glass": 800000.0,
        "san_room_rate": 62050000.0, "san_pub_f": 107825000.0, "san_pub_m": 86050000.0,
        "ffe": 59650000.0, "misc": 52500000.0, "kitchen": 0.0,
        "fl_ht_rate": {"Type1": 150000.0, "Type2": 350000.0},
        "fl_vinyl_rate": {"Type1": 500000.0, "Type2": 750000.0},
        "fl_marmer_rate": {"Type1": 750000.0, "Type2": 1500000.0},
        "facade_precast_pct": 60.0, "facade_window_pct": 30.0, "facade_double_pct": 10.0,
        "fl_ht_pct": 90.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 10.0,
        "san_room_qty": 3.0, "railing_qty": 5.0,
        "mep": 2810941.24, "utility": 92098.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 2500000.0,
        "door_glass": 1000000.0, "railing_rate": 2200000.0, "skylight_rate": 4500000.0,
        "san_dis": 30275000.0, "san_mushola": 36500000.0, "ext_land": 1563000.0,
        "fac_pub": 31000000.0, "fac_res": 10000000.0, "fac_proj": 2000000000.0,
        "cons": 199000.0
    },
    "Retail": {
        "arch_base": 1084000.0, "door_wood": 6000000.0, "door_steel": 8000000.0,
        "hw_wood": 6500000.0, "hw_steel": 2850000.0, "lobby": 2500000.0,
        "gondola": 2500000000.0, "carpet": 1500000.0, "glass": 800000.0,
        "san_room_rate": 0.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "ffe": 0.0, "misc": 0.0, "kitchen": 0.0,
        "fl_ht_rate": {"Type1": 150000.0, "Type2": 350000.0},
        "fl_vinyl_rate": {"Type1": 500000.0, "Type2": 750000.0},
        "fl_marmer_rate": {"Type1": 750000.0, "Type2": 1500000.0},
        "facade_precast_pct": 10.0, "facade_window_pct": 80.0, "facade_double_pct": 10.0,
        "fl_ht_pct": 90.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 10.0,
        "san_room_qty": 0.0, "railing_qty": 0.0,
        "mep": 4000000.0, "utility": 150000.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 2500000.0,
        "door_glass": 1000000.0, "railing_rate": 2200000.0, "skylight_rate": 4500000.0,
        "san_dis": 30275000.0, "san_mushola": 36500000.0, "ext_land": 1563000.0,
        "fac_pub": 31000000.0, "fac_res": 10000000.0, "fac_proj": 2000000000.0,
        "cons": 174000.0
    },
    "Parking": {
        "arch_base": 668000.0,
        "mep": 4000000.0,
        "utility": 150000.0,
        "door_wood": 0.0, "door_steel": 0.0, "door_glass": 0.0,
        "hw_wood": 0.0, "hw_steel": 0.0, "lobby": 0.0, "gondola": 0.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 0.0, "misc": 0.0, "kitchen": 0.0,
        "san_room_rate": 0.0, "san_pub_f": 0.0, "san_pub_m": 0.0,
        "fl_ht_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 10.0, "facade_window_pct": 80.0, "facade_double_pct": 10.0,
        "fl_ht_pct": 90.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 10.0,
        "san_room_qty": 0.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 2500000.0,
        "railing_rate": 2200000.0, "skylight_rate": 4500000.0,
        "san_dis": 30275000.0, "san_mushola": 36500000.0, "ext_land": 1563000.0,
        "fac_pub": 31000000.0, "fac_res": 10000000.0, "fac_proj": 2000000000.0,
        "cons": 53000.0
    },
    "Luxury Apartment": {
        "arch_base": 1517450.0,
        "mep": 3295000.0,
        "utility": 53362.0,
        "door_wood": 4500000.0, "door_steel": 6750000.0, "door_glass": 1000000.0,
        "hw_wood": 1500000.0, "hw_steel": 1400000.0, "lobby": 2500000.0, "gondola": 2000000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 29500000.0, "misc": 5250000000.0, "kitchen": 0.0,
        "san_room_rate": 24450000.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 1000000.0,
        "railing_rate": 1100000.0, "skylight_rate": 4500000.0,
        "san_dis": 24481500.0, "san_mushola": 36500000.0, "ext_land": 632077.0,
        "fac_pub": 0.0, "fac_res": 221432.0, "fac_proj": 2000000000.0,
        "cons": 172813.093
    },
    "Apartment2": {
        "arch_base": 1614800.0,
        "mep": 2247000.0,
        "utility": 84274.0,
        "door_wood": 3500000.0, "door_steel": 6750000.0, "door_glass": 1000000.0,
        "hw_wood": 1500000.0, "hw_steel": 1400000.0, "lobby": 2500000.0, "gondola": 1500000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 29500000.0, "misc": 5250000000.0, "kitchen": 0.0,
        "san_room_rate": 24450000.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 1000000.0,
        "railing_rate": 1100000.0, "skylight_rate": 4500000.0,
        "san_dis": 24481500.0, "san_mushola": 36500000.0, "ext_land": 459538.0,
        "fac_pub": 0.0, "fac_res": 72243.0, "fac_proj": 2000000000.0,
        "cons": 171378.327
    },
    "Hotel 3 Star": {
        "arch_base": 1517450.0,
        "mep": 4223000.0,
        "utility": 131787.0,
        "door_wood": 4500000.0, "door_steel": 6750000.0, "door_glass": 1000000.0,
        "hw_wood": 8000000.0, "hw_steel": 1900000.0, "lobby": 5000000.0, "gondola": 1500000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 64567647.0, "misc": 5250000000.0, "kitchen": 0.0,
        "san_room_rate": 62050000.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 2400000.0, "facade_double_rate": 1000000.0,
        "railing_rate": 1100000.0, "skylight_rate": 4500000.0,
        "san_dis": 62050000.0, "san_mushola": 0.0, "ext_land": 213198.0,
        "fac_pub": 0.0, "fac_res": 108200.0, "fac_proj": 2000000000.0,
        "cons": 173906.6059
    },
    "Retail2": {
        "arch_base": 984500.0,
        "mep": 2651000.0,
        "utility": 23724.0,
        "door_wood": 3500000.0, "door_steel": 5000000.0, "door_glass": 1000000.0,
        "hw_wood": 750000.0, "hw_steel": 1900000.0, "lobby": 2500000.0, "gondola": 2000000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 0.0, "misc": 0.0, "kitchen": 0.0,
        "san_room_rate": 0.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 2500000.0,
        "railing_rate": 2000000.0, "skylight_rate": 4500000.0,
        "san_dis": 62050000.0, "san_mushola": 36500000.0, "ext_land": 559277.0,
        "fac_pub": 0.0, "fac_res": 0.0, "fac_proj": 2000000000.0,
        "cons": 181035.685
    },
    "Terrace Villa": {
        "arch_base": 1517450.0,
        "mep": 4337000.0,
        "utility": 61168.0,
        "door_wood": 3500000.0, "door_steel": 5000000.0, "door_glass": 1000000.0,
        "hw_wood": 750000.0, "hw_steel": 1400000.0, "lobby": 2500000.0, "gondola": 2000000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 0.0, "misc": 5250000000.0, "kitchen": 0.0,
        "san_room_rate": 24450000.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 1250000.0, "facade_double_rate": 1000000.0,
        "railing_rate": 2000000.0, "skylight_rate": 4500000.0,
        "san_dis": 24481500.0, "san_mushola": 36500000.0, "ext_land": 1024897.0,
        "fac_pub": 0.0, "fac_res": 1715946.0, "fac_proj": 2000000000.0,
        "cons": 168148.6486
    },
    "Podium Villa": {
        "arch_base": 1517450.0,
        "mep": 3142000.0,
        "utility": 45249.0,
        "door_wood": 4500000.0, "door_steel": 6750000.0, "door_glass": 1000000.0,
        "hw_wood": 8000000.0, "hw_steel": 1900000.0, "lobby": 5000000.0, "gondola": 1500000000.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 0.0, "misc": 5250000000.0, "kitchen": 0.0,
        "san_room_rate": 62050000.0, "san_pub_f": 154175000.0, "san_pub_m": 126225000.0,
        "fl_ht_rate": {"Type1": 350000.0, "Type2": 200000.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 800000.0, "facade_window_rate": 2400000.0, "facade_double_rate": 1000000.0,
        "railing_rate": 1100000.0, "skylight_rate": 4500000.0,
        "san_dis": 62050000.0, "san_mushola": 0.0, "ext_land": 416735.0,
        "fac_pub": 0.0, "fac_res": 100897.0, "fac_proj": 2000000000.0,
        "cons": 173505.7655
    },
    "Parking2": {
        "arch_base": 1040000.0,
        "mep": 900000.0,
        "utility": 0.0,
        "door_wood": 0.0, "door_steel": 0.0, "door_glass": 0.0,
        "hw_wood": 0.0, "hw_steel": 0.0, "lobby": 0.0, "gondola": 0.0,
        "carpet": 0.0, "glass": 0.0, "ffe": 0.0, "misc": 0.0, "kitchen": 0.0,
        "san_room_rate": 0.0, "san_pub_f": 0.0, "san_pub_m": 0.0,
        "fl_ht_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_vinyl_rate": {"Type1": 0.0, "Type2": 0.0},
        "fl_marmer_rate": {"Type1": 0.0, "Type2": 0.0},
        "facade_precast_pct": 0.0, "facade_window_pct": 0.0, "facade_double_pct": 0.0,
        "fl_ht_pct": 0.0, "fl_vinyl_pct": 0.0, "fl_marmer_pct": 0.0,
        "san_room_qty": 1.0, "railing_qty": 0.0,
        "struc_earth": 25000.0, "struc_found": 400000.0, "struc_work": 1933000.0,
        "facade_precast_rate": 0.0, "facade_window_rate": 0.0, "facade_double_rate": 0.0,
        "railing_rate": 0.0, "skylight_rate": 0.0,
        "san_dis": 0.0, "san_mushola": 0.0, "ext_land": 0.0,
        "fac_pub": 0.0, "fac_res": 0.0, "fac_proj": 0.0,
        "cons": 52500.0
    }
}

#region --- DO NOT CHANGE#2 (OR I WILL KICK YOUR FACE)---
def mi(name):
    return f":material/{name}:"

def icon_safe(name):
    return mi(name) if "mi" in globals() else None    

def cb_add_project():
    st.session_state.proj_counter += 1
    new_id = f"proj_{st.session_state.proj_counter}"
    st.session_state.projects[new_id] = {
        "name": f"New Project {st.session_state.proj_counter}",
        "type": "Hotel",
        "data": {}
    }
    st.session_state.current_proj_id = new_id
    save_data()

def cb_delete_project():
    projects = st.session_state.get("projects", {})
    curr_id = st.session_state.get("current_proj_id")

    if not isinstance(projects, dict) or len(projects) <= 1:
        st.warning("At least one project must remain.")
        repair_projects_state(save=True)
        return

    if curr_id in projects:
        del projects[curr_id]

    repair_projects_state(save=True)

def cb_switch_project():
    # Use .get() to avoid the AttributeError if the key is missing
    selected_label = st.session_state.get("project_selector")
    
    if not selected_label:
        return

    proj_ids = list(st.session_state.projects.keys())
    proj_labels = [f"{st.session_state.projects[pid]['name']} ({st.session_state.projects[pid]['type']})" for pid in proj_ids]
    
    if selected_label in proj_labels:
        selected_idx = proj_labels.index(selected_label)
        st.session_state.current_proj_id = proj_ids[selected_idx]
        save_data()     

def create_new_feasibility_study(study_name, project_type="Hotel"):
    clean_name = str(study_name).strip()

    if clean_name == "":
        st.error("Please enter feasibility study name.")
        return False

    st.session_state.projects = {
        "proj_1": {
            "name": clean_name,
            "type": project_type,
            "data": {}
        }
    }

    st.session_state.current_proj_id = "proj_1"
    st.session_state.proj_counter = 1

    st.session_state.loaded_snapshot_id = None
    st.session_state.loaded_snapshot_name = None

    st.session_state.report_config = copy.deepcopy(DEFAULT_REPORT_CONFIG)

    keys_to_clear = [
        k for k in st.session_state.keys()
        if "base_table_" in k
        or "area_editor_" in k
        or "rename_input_" in k
        or "renaming_" in k
        or "deleting_" in k
        or "fs_load_page" in k
    ]

    for k in keys_to_clear:
        del st.session_state[k]

    return save_snapshot(clean_name)

#endregion

def render_feasibility_study_landing(): #start page
    st.title("Feasibility Study")

    st.divider()

    active_file_id = st.session_state.get("loaded_snapshot_id")
    active_file_name = st.session_state.get("loaded_snapshot_name")

    # ==================================================
    # LANDING MODE STATE
    # ==================================================
    if "fs_landing_mode" not in st.session_state:
        st.session_state.fs_landing_mode = None

    # ==================================================
    # AFTER FILE IS CREATED / LOADED
    # ==================================================

    if active_file_id and st.session_state.fs_landing_mode is None:
        
        c1, c2 = st.columns([1, 1])
        
        c1.success(f"**{active_file_name}** is currently loaded (You can start calculating your project)", icon=":material/check:")
        c2.info("Use **Quick Save** button on the sidebar to save calculation", icon=":material/help:")

        col_msg, col_back = st.columns([1, 5])
    
        with col_msg:
            if st.button("Previous Page", icon=":material/arrow_back:", key="go_back_to_load_list", use_container_width=True):
                st.session_state.fs_landing_mode = "home"
                st.rerun()

        return

    # ==================================================
    # FIRST-SIGHT QUESTION
    # ==================================================
    if st.session_state.fs_landing_mode is None or st.session_state.fs_landing_mode == "home":

        st.info("""

        **Welcome to Project Feasibility Study - Agung Sedayu Group**

        To start calculating, first create a new project by clicking the button below.""", icon=":material/waving_hand:")

        st.space(size="small")

        col_create_btn, col_load_btn, col_empty = st.columns([1, 1, 2], gap="small", vertical_alignment="center")

        with col_create_btn:
            if st.button(
                "Create New Feasibility Study",
                key="landing_choose_create_study",
                type="primary",
                use_container_width=True,
                icon=":material/create_new_folder:"
            ):
                st.session_state.fs_landing_mode = "create"
                st.rerun()

        with col_load_btn:
            st.info("**or load saved FS below:**")


    # ==================================================
    # CREATE NEW STUDY MODE
    # ==================================================
    elif st.session_state.fs_landing_mode == "create":
        col_title, col_back = st.columns([5, 1])

        with col_title:
            st.subheader("Create New Feasibility Study")
            st.caption("Enter a study name. A new default project will be created automatically.")

        with col_back:
            if st.button("Back", key="landing_create_back", use_container_width=True):
                st.session_state.fs_landing_mode = None
                st.rerun()


        col_title, col_back = st.columns([5, 1])

        with col_title.form("create_new_feasibility_study_form", clear_on_submit=False):
            study_name = st.text_input(
                "Feasibility Study Name",
                placeholder="e.g. Project X - Option 1 - Rev 0"
            )

            create_clicked = st.form_submit_button(
                "Create New Feasibility Study",
                type="primary",
                use_container_width=True
            )

            if create_clicked:
                if study_name.strip() == "":
                    st.warning("Please enter feasibility study name.")
                else:
                    # Uses default project type from create_new_feasibility_study()
                    # Usually default = Hotel
                    if create_new_feasibility_study(study_name):
                        st.session_state.fs_landing_mode = None
                        st.success(f"Created **{study_name.strip()}**.")
                        st.rerun()
    
    snapshots = load_snapshots()

    if not snapshots:
        st.info("No saved feasibility studies yet.")

    else:
        # ==================================================
        # PAGINATION SETUP
        # ==================================================
        PAGE_SIZE = 10

        if "fs_load_page" not in st.session_state:
            st.session_state.fs_load_page = 0

        total_items = len(snapshots)
        total_pages = max(1, (total_items + PAGE_SIZE - 1) // PAGE_SIZE)

        # Safety clamp if files were deleted / changed
        if st.session_state.fs_load_page >= total_pages:
            st.session_state.fs_load_page = total_pages - 1

        if st.session_state.fs_load_page < 0:
            st.session_state.fs_load_page = 0

        start_idx = st.session_state.fs_load_page * PAGE_SIZE
        end_idx = start_idx + PAGE_SIZE
        page_snapshots = snapshots[start_idx:end_idx]

        st.divider()

        # ==================================================
        # SAVED FILE LIST
        # ==================================================
        for snap in page_snapshots:
            snap_id = snap["id"]
            snap_name = snap.get("snapshot_name", "Untitled File")

            is_active = st.session_state.get("loaded_snapshot_id") == snap_id
            active_label = " — ACTIVE" if is_active else ""

            saved_time = ""
            if "format_snapshot_time" in globals():
                saved_time = format_snapshot_time(snap.get("created_at"))

            col_file, col_date, col_action= st.columns([4, 2, 1])

            with col_file:
                st.markdown(f"**{snap_name}**")
                if is_active:
                    st.badge("Currently loaded file", icon=":material/check:", color="green")

                if saved_time:
                    st.caption(f"Saved: {saved_time}{active_label}")
                else:
                    st.caption(f"Saved file{active_label}")

            with col_action:
                if st.button(
                    "Load",
                    key=f"landing_load_file_{snap_id}",
                    type="primary",
                    use_container_width=True
                ):
                    data = load_snapshot_data(snap_id)

                    if data:
                        restore_app_payload(data)

                        st.session_state.loaded_snapshot_id = snap_id
                        st.session_state.loaded_snapshot_name = snap_name
                        st.session_state.fs_landing_mode = None

                        save_data()
                        st.success(f"Loaded **{snap_name}**.")
                        st.rerun()

            st.divider()

        # ==================================================
        # PAGINATION CONTROLS
        # ==================================================
        c1, col_prev, col_page, col_next, c2 = st.columns([5, 1, 2, 1, 5])

        with col_prev:
            if st.button(
                "Previous",
                key="fs_load_prev_page",
                use_container_width=True,
                disabled=st.session_state.fs_load_page <= 0
            ):
                st.session_state.fs_load_page -= 1
                st.rerun()

        with col_page:
            st.markdown(
                f"<div style='text-align:center; padding-top: 0.45rem;'>"
                f"Page {st.session_state.fs_load_page + 1} of {total_pages}"
                f"</div>",
                unsafe_allow_html=True
            )

        with col_next:
            if st.button(
                "Next",
                key="fs_load_next_page",
                use_container_width=True,
                disabled=st.session_state.fs_load_page >= total_pages - 1
            ):
                st.session_state.fs_load_page += 1
                st.rerun()

    return                
    st.divider()
    st.caption("For rename, delete, import, export, or full archive management, use the Feasibility Study Archive page.")

def show_project_database():  # database page
    st.title("Project Database")

    # ==================================================
    # CONTEXT MAP
    # Converts internal database keys into readable QS labels
    # ==================================================
    FIELD_CONTEXT = {
        # Structure
        "struc_earth": {
            "Group": "Structure",
            "Item": "Earthwork",
            "Basis": "Rp / m² GBA",
            "Type": "currency",
            "Note": "Applied to gross building area."
        },
        "struc_found": {
            "Group": "Structure",
            "Item": "Foundation Work",
            "Basis": "Rp / m² GBA",
            "Type": "currency",
            "Note": "Applied to gross building area."
        },
        "struc_work": {
            "Group": "Structure",
            "Item": "Main Structure Work",
            "Basis": "Rp / m² GBA",
            "Type": "currency",
            "Note": "Applied to gross building area."
        },

        # Architecture Base
        "arch_base": {
            "Group": "Architecture",
            "Item": "Base Architectural Work",
            "Basis": "Rp / m² GFA",
            "Type": "currency",
            "Note": "General architectural finishing rate."
        },
        "lobby": {
            "Group": "Architecture",
            "Item": "Lobby Finishing Premium",
            "Basis": "Rp / m²",
            "Type": "currency",
            "Note": "Additional lobby finishing allowance."
        },

        # Facade
        "facade_precast_rate": {
            "Group": "Facade",
            "Item": "Precast Facade Rate",
            "Basis": "Rp / m² facade",
            "Type": "currency",
            "Note": "Applied according to precast facade ratio."
        },
        "facade_window_rate": {
            "Group": "Facade",
            "Item": "Window / Glass Facade Rate",
            "Basis": "Rp / m² facade",
            "Type": "currency",
            "Note": "Applied according to window facade ratio."
        },
        "facade_double_rate": {
            "Group": "Facade",
            "Item": "Double Facade Rate",
            "Basis": "Rp / m² facade",
            "Type": "currency",
            "Note": "Applied according to double facade ratio."
        },
        "facade_precast_pct": {
            "Group": "Facade",
            "Item": "Precast Facade Ratio",
            "Basis": "% of facade area",
            "Type": "percent",
            "Note": "Facade composition assumption."
        },
        "facade_window_pct": {
            "Group": "Facade",
            "Item": "Window Facade Ratio",
            "Basis": "% of facade area",
            "Type": "percent",
            "Note": "Facade composition assumption."
        },
        "facade_double_pct": {
            "Group": "Facade",
            "Item": "Double Facade Ratio",
            "Basis": "% of facade area",
            "Type": "percent",
            "Note": "Facade composition assumption."
        },

        # Doors & Hardware
        "door_wood": {
            "Group": "Doors & Hardware",
            "Item": "Wooden Door",
            "Basis": "Rp / unit",
            "Type": "currency",
            "Note": "Door supply and installation allowance."
        },
        "door_steel": {
            "Group": "Doors & Hardware",
            "Item": "Steel Door",
            "Basis": "Rp / unit",
            "Type": "currency",
            "Note": "Door supply and installation allowance."
        },
        "door_glass": {
            "Group": "Doors & Hardware",
            "Item": "Glass Door",
            "Basis": "Rp / unit",
            "Type": "currency",
            "Note": "Door supply and installation allowance."
        },
        "hw_wood": {
            "Group": "Doors & Hardware",
            "Item": "Wooden Door Hardware",
            "Basis": "Rp / set",
            "Type": "currency",
            "Note": "Hardware set allowance."
        },
        "hw_steel": {
            "Group": "Doors & Hardware",
            "Item": "Steel Door Hardware",
            "Basis": "Rp / set",
            "Type": "currency",
            "Note": "Hardware set allowance."
        },

        # Flooring
        "fl_waste": {
            "Group": "Flooring",
            "Item": "Flooring Wastage",
            "Basis": "%",
            "Type": "percent",
            "Note": "Material waste allowance."
        },
        "fl_ht_pct": {
            "Group": "Flooring",
            "Item": "Homogeneous Tile Ratio",
            "Basis": "% of floor area",
            "Type": "percent",
            "Note": "Floor finish composition."
        },
        "fl_vinyl_pct": {
            "Group": "Flooring",
            "Item": "Vinyl Floor Ratio",
            "Basis": "% of floor area",
            "Type": "percent",
            "Note": "Floor finish composition."
        },
        "fl_marmer_pct": {
            "Group": "Flooring",
            "Item": "Marble Floor Ratio",
            "Basis": "% of floor area",
            "Type": "percent",
            "Note": "Floor finish composition."
        },

        # Interior / Specialist
        "gondola": {
            "Group": "Specialist Works",
            "Item": "Gondola System",
            "Basis": "Rp / project",
            "Type": "currency",
            "Note": "Facade maintenance equipment allowance."
        },
        "carpet": {
            "Group": "Interior",
            "Item": "Carpet Finish",
            "Basis": "Rp / m²",
            "Type": "currency",
            "Note": "Carpet finishing rate."
        },
        "glass": {
            "Group": "Interior",
            "Item": "Interior Glass / Mirror",
            "Basis": "Rp / m²",
            "Type": "currency",
            "Note": "Interior glass allowance."
        },
        "ffe": {
            "Group": "Interior",
            "Item": "FF&E",
            "Basis": "Rp / room or unit",
            "Type": "currency",
            "Note": "Furniture, fixtures, and equipment allowance."
        },
        "misc": {
            "Group": "Interior",
            "Item": "Miscellaneous Interior Allowance",
            "Basis": "Rp allowance",
            "Type": "currency",
            "Note": "Project-specific miscellaneous allowance."
        },
        "kitchen": {
            "Group": "Interior",
            "Item": "Kitchen Equipment",
            "Basis": "Rp allowance",
            "Type": "currency",
            "Note": "Kitchen equipment allowance."
        },

        # Sanitary
        "san_room_rate": {
            "Group": "Sanitary",
            "Item": "Typical Room Sanitary",
            "Basis": "Rp / room",
            "Type": "currency",
            "Note": "Sanitary allowance for typical room or unit."
        },
        "san_pub_f": {
            "Group": "Sanitary",
            "Item": "Public Female Toilet",
            "Basis": "Rp / toilet set",
            "Type": "currency",
            "Note": "Public toilet sanitary allowance."
        },
        "san_pub_m": {
            "Group": "Sanitary",
            "Item": "Public Male Toilet",
            "Basis": "Rp / toilet set",
            "Type": "currency",
            "Note": "Public toilet sanitary allowance."
        },
        "san_dis": {
            "Group": "Sanitary",
            "Item": "Accessible Toilet",
            "Basis": "Rp / toilet set",
            "Type": "currency",
            "Note": "Disabled toilet sanitary allowance."
        },
        "san_mushola": {
            "Group": "Sanitary",
            "Item": "Mushola Ablution Area",
            "Basis": "Rp / area",
            "Type": "currency",
            "Note": "Wudhu / mushola sanitary allowance."
        },
        "san_room_qty": {
            "Group": "Sanitary",
            "Item": "Sanitary Quantity per Room",
            "Basis": "Qty / room",
            "Type": "number",
            "Note": "Typical sanitary quantity assumption."
        },

        # MEP & Utility
        "mep": {
            "Group": "MEP",
            "Item": "MEP Works",
            "Basis": "Rp / m² GBA",
            "Type": "currency",
            "Note": "Mechanical, electrical, and plumbing rate."
        },
        "utility": {
            "Group": "Utility",
            "Item": "Infrastructure / Utility Works",
            "Basis": "Rp / m² GBA",
            "Type": "currency",
            "Note": "External or supporting utility allowance."
        },

        # External Works
        "ext_land": {
            "Group": "External Works",
            "Item": "Landscape / External Works",
            "Basis": "Rp / m²",
            "Type": "currency",
            "Note": "External area and landscape allowance."
        },
        "railing_rate": {
            "Group": "External Works",
            "Item": "Railing",
            "Basis": "Rp / m",
            "Type": "currency",
            "Note": "Railing work allowance."
        },
        "railing_qty": {
            "Group": "External Works",
            "Item": "Railing Quantity Ratio",
            "Basis": "Qty",
            "Type": "number",
            "Note": "Default railing quantity assumption."
        },
        "skylight_rate": {
            "Group": "External Works",
            "Item": "Skylight",
            "Basis": "Rp / m²",
            "Type": "currency",
            "Note": "Skylight work allowance."
        },

        # Facilities
        "fac_pub": {
            "Group": "Facilities",
            "Item": "Public Facility Allowance",
            "Basis": "Rp / room or unit",
            "Type": "currency",
            "Note": "Public facility cost allowance."
        },
        "fac_res": {
            "Group": "Facilities",
            "Item": "Residential Facility Allowance",
            "Basis": "Rp / room or unit",
            "Type": "currency",
            "Note": "Residential facility cost allowance."
        },
        "fac_proj": {
            "Group": "Facilities",
            "Item": "Project Facility Allowance",
            "Basis": "Rp / project",
            "Type": "currency",
            "Note": "Lump-sum project facility allowance."
        },

        # Soft Cost
        "cons": {
            "Group": "Soft Cost",
            "Item": "Consultancy Cost",
            "Basis": "Rp / m² GFA",
            "Type": "currency",
            "Note": "Consultant / professional fee allowance."
        },
    }

    GROUP_ORDER = [
        "Structure",
        "Architecture",
        "Facade",
        "Doors & Hardware",
        "Flooring",
        "Interior",
        "Sanitary",
        "MEP",
        "Utility",
        "External Works",
        "Facilities",
        "Specialist Works",
        "Soft Cost",
        "Other"
    ]

    # ==================================================
    # HELPER FUNCTIONS
    # ==================================================
    def prettify_key(raw_key):
        text = raw_key.replace("_", " ").replace(".", " - ")
        return text.title()

    def get_field_context(raw_key):
        # Exact match
        if raw_key in FIELD_CONTEXT:
            return FIELD_CONTEXT[raw_key]

        # Nested flooring rates, e.g. fl_ht_rate.Type1
        if raw_key.startswith("fl_ht_rate."):
            subtype = raw_key.split(".", 1)[1]
            return {
                "Group": "Flooring",
                "Item": f"Homogeneous Tile Rate - {subtype}",
                "Basis": "Rp / m²",
                "Type": "currency",
                "Note": "Floor finish unit rate."
            }

        if raw_key.startswith("fl_vinyl_rate."):
            subtype = raw_key.split(".", 1)[1]
            return {
                "Group": "Flooring",
                "Item": f"Vinyl Floor Rate - {subtype}",
                "Basis": "Rp / m²",
                "Type": "currency",
                "Note": "Floor finish unit rate."
            }

        if raw_key.startswith("fl_marmer_rate."):
            subtype = raw_key.split(".", 1)[1]
            return {
                "Group": "Flooring",
                "Item": f"Marble Floor Rate - {subtype}",
                "Basis": "Rp / m²",
                "Type": "currency",
                "Note": "Floor finish unit rate."
            }

        return {
            "Group": "Other",
            "Item": prettify_key(raw_key),
            "Basis": "-",
            "Type": "number",
            "Note": "Unmapped database item."
        }

    def format_value(value, value_type):
        try:
            num = _safe_float(value)
        except Exception:
            return str(value)

        if value_type == "currency":
            return f"Rp {num:,.0f}"
        elif value_type == "percent":
            return f"{num:,.1f}%"
        elif value_type == "number":
            return f"{num:,.2f}".rstrip("0").rstrip(".")
        else:
            return str(value)

    def flatten_project_database():
        rows = []

        for project_type, metrics in PROJECT_DATABASE.items():
            for key_name, value in metrics.items():

                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        raw_key = f"{key_name}.{sub_key}"
                        ctx = get_field_context(raw_key)

                        rows.append({
                            "Project Type": project_type,
                            "Group": ctx["Group"],
                            "Cost Item": ctx["Item"],
                            "Basis": ctx["Basis"],
                            "Value": sub_value,
                            "Formatted Value": format_value(sub_value, ctx["Type"]),
                            "Note": ctx["Note"],
                            "Internal Key": raw_key
                        })

                else:
                    raw_key = key_name
                    ctx = get_field_context(raw_key)

                    rows.append({
                        "Project Type": project_type,
                        "Group": ctx["Group"],
                        "Cost Item": ctx["Item"],
                        "Basis": ctx["Basis"],
                        "Value": value,
                        "Formatted Value": format_value(value, ctx["Type"]),
                        "Note": ctx["Note"],
                        "Internal Key": raw_key
                    })

        df = pd.DataFrame(rows)

        df["Group Sort"] = df["Group"].apply(
            lambda x: GROUP_ORDER.index(x) if x in GROUP_ORDER else 999
        )

        df = df.sort_values(
            by=["Group Sort", "Cost Item", "Project Type"]
        ).drop(columns=["Group Sort"])

        return df

    df_long = flatten_project_database()

    project_types = list(PROJECT_DATABASE.keys())

    # ==================================================
    # PAGE STYLE
    # ==================================================
    st.markdown("""
    <style>
        .db-card {
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            padding: 14px 16px;
            box-shadow: 0 1px 3px rgba(16,24,40,0.04);
        }

        .db-label {
            font-size: 11px;
            color: #6B7280;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            font-weight: 700;
            margin-bottom: 5px;
        }

        .db-value {
            font-size: 20px;
            color: #111827;
            font-weight: 750;
            line-height: 1.2;
        }

        .db-sub {
            font-size: 12px;
            color: #6B7280;
            margin-top: 4px;
        }

        .db-note {
            background: #F9FAFB;
            border: 1px solid #E5E7EB;
            border-left: 4px solid #3E4095;
            border-radius: 12px;
            padding: 12px 14px;
            margin-bottom: 14px;
            font-size: 13px;
            color: #4B5563;
            line-height: 1.55;
        }
    </style>
    """, unsafe_allow_html=True)

    # ==================================================
    # SUMMARY CARDS
    # ==================================================
    total_project_types = len(project_types)
    total_items = df_long["Cost Item"].nunique()
    total_groups = df_long["Group"].nunique()

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown(
            f"""
            <div class="db-card">
                <div class="db-label">Project Types</div>
                <div class="db-value">{total_project_types}</div>
                <div class="db-sub">Available benchmark templates</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c2:
        st.markdown(
            f"""
            <div class="db-card">
                <div class="db-label">Cost Items</div>
                <div class="db-value">{total_items}</div>
                <div class="db-sub">Contextual database fields</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c3:
        st.markdown(
            f"""
            <div class="db-card">
                <div class="db-label">Cost Groups</div>
                <div class="db-value">{total_groups}</div>
                <div class="db-sub">Grouped by QS discipline</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    st.divider()
    # ==================================================
    # MAIN TABS
    # ==================================================
    tab_search, tab_project, tab_matrix, tab_raw = st.tabs([
        "Search Database",
        "Project Type View",
        "Comparison Matrix",
        "Raw Audit View"
    ])

    # ==================================================
    # TAB 1: PROJECT TYPE VIEW
    # ==================================================
    with tab_project:
        selected_project_type = st.selectbox(
            "Select Project Type",
            options=project_types,
            index=0
        )

        selected_groups = st.multiselect(
            "Filter Cost Group",
            options=GROUP_ORDER,
            default=[
                "Structure",
                "Architecture",
                "Facade",
                "Flooring",
                "MEP",
                "Utility",
                "Soft Cost"
            ]
        )

        df_project = df_long[
            (df_long["Project Type"] == selected_project_type)
            & (df_long["Group"].isin(selected_groups))
        ].copy()

        display_df = df_project[[
            "Group",
            "Cost Item",
            "Basis",
            "Formatted Value",
            "Note"
        ]].rename(columns={
            "Formatted Value": "Rate / Assumption"
        })

        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Group": st.column_config.TextColumn("Cost Group", width="medium"),
                "Cost Item": st.column_config.TextColumn("Cost Item", width="large"),
                "Basis": st.column_config.TextColumn("Basis", width="medium"),
                "Rate / Assumption": st.column_config.TextColumn("Rate / Assumption", width="medium"),
                "Note": st.column_config.TextColumn("Context", width="large"),
            }
        )

    # ==================================================
    # TAB 2: COMPARISON MATRIX
    # ==================================================
    with tab_matrix:
        st.caption("Compare contextual cost items across project types.")

        matrix_groups = st.multiselect(
            "Select Groups to Compare",
            options=GROUP_ORDER,
            default=["Structure", "Architecture", "Facade", "MEP", "Utility", "Soft Cost"],
            key="db_matrix_groups"
        )

        compare_project_types = st.multiselect(
            "Select Project Types",
            options=project_types,
            default=project_types[:5],
            key="db_compare_project_types"
        )

        df_matrix_source = df_long[
            (df_long["Group"].isin(matrix_groups))
            & (df_long["Project Type"].isin(compare_project_types))
        ].copy()

        matrix_df = df_matrix_source.pivot_table(
            index=["Group", "Cost Item", "Basis"],
            columns="Project Type",
            values="Formatted Value",
            aggfunc="first"
        ).reset_index()

        st.dataframe(
            matrix_df,
            use_container_width=True,
            hide_index=True
        )

    # ==================================================
    # TAB 3: RAW AUDIT VIEW
    # ==================================================
    with tab_raw:
        st.caption("Audit view showing internal keys beside contextual labels.")

        raw_groups = st.multiselect(
            "Filter Raw View by Group",
            options=GROUP_ORDER,
            default=GROUP_ORDER,
            key="db_raw_groups"
        )

        raw_df = df_long[df_long["Group"].isin(raw_groups)].copy()

        st.dataframe(
            raw_df[[
                "Project Type",
                "Group",
                "Cost Item",
                "Basis",
                "Formatted Value",
                "Internal Key",
                "Note"
            ]].rename(columns={
                "Formatted Value": "Rate / Assumption"
            }),
            use_container_width=True,
            hide_index=True
        )

    # ==================================================
    # TAB 4: SEARCH DATABASE
    # ==================================================
    with tab_search:

        c1, c2, c3 = st.columns(3)

        with c1:
            search_item = st.text_input(
                "Item Name",
                placeholder="e.g. Foundation, Utility, MEP",
                key="db_search_item"
            )

        with c2:
            search_group = st.text_input(
                "Cost Group",
                placeholder="e.g. Structure, Architecture, Flooring",
                key="db_search_group"
            )

        with c3:
            search_project_type = st.text_input(
                "Project Type",
                placeholder="e.g. Hotel, Apartment, Retail",
                key="db_search_project_type"
            )

        has_search_input = (
            search_item.strip() != ""
            or search_group.strip() != ""
            or search_project_type.strip() != ""
        )

        if not has_search_input:
            st.info("Enter item name, cost group, or project type to search the database.")
        else:
            df_search = df_long.copy()

            if search_item.strip():
                item_query = search_item.strip().lower()
                df_search = df_search[
                    df_search["Cost Item"].str.lower().str.contains(item_query, na=False)
                    | df_search["Internal Key"].str.lower().str.contains(item_query, na=False)
                    | df_search["Note"].str.lower().str.contains(item_query, na=False)
                ]

            if search_group.strip():
                group_query = search_group.strip().lower()
                df_search = df_search[
                    df_search["Group"].str.lower().str.contains(group_query, na=False)
                ]

            if search_project_type.strip():
                project_query = search_project_type.strip().lower()
                df_search = df_search[
                    df_search["Project Type"].str.lower().str.contains(project_query, na=False)
                ]

            st.divider()

            st.caption(f"Search result: {len(df_search):,} rows")

            if df_search.empty:
                st.info("No matching database item found.")
            else:
                st.dataframe(
                    df_search[[
                        "Group",
                        "Cost Item",
                        "Project Type",
                        "Basis",
                        "Formatted Value",
                        "Internal Key",
                        "Note"
                    ]].rename(columns={
                        "Group": "Cost Group",
                        "Formatted Value": "Rate / Assumption"
                    }),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Cost Group": st.column_config.TextColumn("Cost Group", width="medium"),
                        "Cost Item": st.column_config.TextColumn("Cost Item", width="large"),
                        "Project Type": st.column_config.TextColumn("Project Type", width="medium"),
                        "Basis": st.column_config.TextColumn("Basis", width="medium"),
                        "Rate / Assumption": st.column_config.TextColumn("Rate / Assumption", width="medium"),
                        "Internal Key": st.column_config.TextColumn("Internal Key", width="medium"),
                        "Note": st.column_config.TextColumn("Context", width="large"),
                    }
                )

def show_area_calculator(): #area calculator page
    st.title("Area Analysis")
    
    # 1. Identify the Active Project
    curr_id, curr_proj = get_current_project()
    
    def get_area_val(key, default=0.0):
        return curr_proj["data"].get(key, default)

    tab1, tab2 = st.tabs(["GBA", "Details"])
    
    with tab1:
        st.subheader("1. Bottom-Up Feasibility Study")
        st.caption("Input your breakdown per floor. GBA, GFA, SGFA, and NFA will be calculated automatically.")

        # ==========================================
        # THE FIX: THE STATE ANCHOR
        # ==========================================
        # We isolate the base data so the data_editor doesn't reset on every keystroke.
        base_key = f"base_table_{curr_id}"
        
        if base_key not in st.session_state:
            if "area_table_data" in curr_proj["data"]:
                st.session_state[base_key] = curr_proj["data"]["area_table_data"]
            else:
                default_stack = [
                    {"FL": "Roof Machine", "Space Type": "Roof", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "Roof", "Space Type": "Roof", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "5F", "Space Type": "Unit", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "4F", "Space Type": "Unit", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "3F", "Space Type": "Unit", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "2F", "Space Type": "Unit", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0},
                    {"FL": "1F", "Space Type": "Lobby", "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0}
                ]
                st.session_state[base_key] = default_stack
                curr_proj["data"]["area_table_data"] = default_stack

        # --- TOP SETUP CONTROLS ---
        c_name, c_h, c_b, c_u = st.columns(4)

        # Using get_area_val ensures the data survives page reloads
        tower_name = curr_proj.get("name", "Unnamed Component")
        curr_proj["data"]["tname"] = tower_name

        c_name.text_input(
            "Project Name",
            value=tower_name,
            key=f"wid_tname_{curr_id}",
            disabled=True,
            help="This follows the active component name from the sidebar."
        )

        basements = c_h.number_input(
            "Basements / LG",
            min_value=0,
            value=int(get_area_val("base_in", 1)),
            step=1,
            key=f"wid_base_{curr_id}",
            help="1 = LG, 2 = B1, 3 = B2, etc"
        )
        curr_proj["data"]["base_in"] = basements

        upper_floors = c_b.number_input(
            "Floors",
            min_value=1,
            value=int(get_area_val("up_in", 5)),
            step=1,
            key=f"wid_up_{curr_id}"
        )
        curr_proj["data"]["up_in"] = upper_floors

        land_area = c_u.number_input(
            "Luas Tanah (m²)",
            min_value=0.0,
            value=_safe_float(get_area_val("m_land", 0.0)),
            step=100.0,
            key=f"wid_m_land_{curr_id}"
        )
        curr_proj["data"]["m_land"] = land_area
        
        if st.button("Generate Table", type="primary"):
            new_data = [
                {"FL": "Roof Machine", "Space Type": "Roof", "Floor to Floor Height (m)": 3.0, "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 50.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 150.0, "Unit Area": 0.0, "Office": 0.0},
                {"FL": "Roof", "Space Type": "Roof", "Floor to Floor Height (m)": 3.0, "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 200.0, "MEP Outdoor": 80.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 790.0, "Unit Area": 0.0, "Office": 0.0}
            ]
            for i in range(upper_floors, 1, -1):
                new_data.append({"FL": f"{i}F", "Space Type": "Unit", "Floor to Floor Height (m)": 3.2, "Typical Unit": 8, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 75.0, "Stair, MEP, Etc": 150.0, "Unit Area": 500.0, "Office": 0.0})
            new_data.append({"FL": "1F", "Space Type": "Lobby", "Floor to Floor Height (m)": 4.5, "Typical Unit": 0, "Parkir": 0.0, "Roof/Deck": 0.0, "MEP Outdoor": 20.0, "Koridor/Lobby": 75.0, "Stair, MEP, Etc": 0.0, "Unit Area": 0.0, "Office": 0.0})
            for i in range(1, basements + 1):
                fl_name = "LG" if i == 1 else f"B{i-1}"
                new_data.append({"FL": fl_name, "Space Type": "Carpark", "Floor to Floor Height (m)": 3.2, "Typical Unit": 0, "Parkir": 800.0, "Roof/Deck": 0.0, "MEP Outdoor": 0.0, "Koridor/Lobby": 0.0, "Stair, MEP, Etc": 150.0, "Unit Area": 0.0, "Office": 0.0})
            
            # Update Anchor and clear editor memory
            st.session_state[base_key] = new_data
            curr_proj["data"]["area_table_data"] = new_data
            save_data()
            if f"area_editor_{curr_id}" in st.session_state:
                del st.session_state[f"area_editor_{curr_id}"]
            st.rerun()

        st.markdown(f"##### Tabel GBA - {tower_name}")

        # --- BUILD DATAFRAME FROM ANCHOR ---
        
        base_data = st.session_state.get(base_key, [])

        # ✅ Normalize area table data
        if not isinstance(base_data, list) or not base_data or not isinstance(base_data[0], dict):
            base_data = [
                {
                    "FL": "1F",
                    "Space Type": "Lobby",
                    "Floor to Floor Height (m)": 4.5,
                    "Parkir": 0.0,
                    "Roof/Deck": 0.0,
                    "MEP Outdoor": 0.0,
                    "Koridor/Lobby": 0.0,
                    "Stair, MEP, Etc": 0.0,
                    "Typical Unit": 0,
                    "Unit Area": 0.0,
                    "Office": 0.0
                }
            ]
            st.session_state[base_key] = base_data
            curr_proj["data"]["area_table_data"] = base_data

        F2F_COL = "Floor to Floor Height (m)"
        UNIT_AREA_COL = "Unit Area"
        TYPICAL_UNIT_COL = "Typical Unit"

        def guess_f2f_height(row):
            fl = str(row.get("FL", "")).strip()
            space_type = str(row.get("Space Type", "")).strip()

            if fl == "1F" or space_type == "Lobby":
                return 4.5
            if fl == "LG" or fl.startswith("B") or space_type == "Carpark":
                return 3.2
            if fl in ["Roof", "Roof Machine"] or space_type == "Roof":
                return 3.0
            if space_type == "Office":
                return 3.6
            if space_type == "Facility":
                return 4.0

            return 3.2  # default typical apartment/unit floor height


        # Add / migrate columns for old and new saved tables
        for row in base_data:
            # Old saved data used "Unit" as an area column.
            # New standard: "Unit Area" = area, "Typical Unit" = count/unit number per floor.
            if UNIT_AREA_COL not in row:
                row[UNIT_AREA_COL] = _safe_float(row.get("Unit", 0.0))

            # Optional: remove old confusing key from saved records
            if "Unit" in row:
                row.pop("Unit", None)

            if TYPICAL_UNIT_COL not in row:
                row[TYPICAL_UNIT_COL] = 0

            if F2F_COL not in row:
                row[F2F_COL] = guess_f2f_height(row)

        st.session_state[base_key] = base_data
        curr_proj["data"]["area_table_data"] = base_data

        df_area = pd.DataFrame(base_data)

        breakdown_cols = [
            "Parkir",
            "Roof/Deck",
            "MEP Outdoor",
            "Koridor/Lobby",
            "Stair, MEP, Etc",
            UNIT_AREA_COL,
            "Office"
        ]

        detail_input_cols = [TYPICAL_UNIT_COL]
        
        # FULL WIDTH EDITOR
        edited_df = st.data_editor(
            df_area,
            num_rows="dynamic",
            use_container_width=True,
            key=f"area_editor_{curr_id}",
            hide_index=True,
            column_order=["FL", "Space Type", F2F_COL, TYPICAL_UNIT_COL] + breakdown_cols,
            column_config={
                "Space Type": st.column_config.SelectboxColumn(
                    "Space Type",
                    options=["Roof", "Unit", "Lobby", "Ramp", "Carpark", "Facility"],
                    required=True
                ),
                F2F_COL: st.column_config.NumberColumn(
                    "Floor to Floor Height (m)",
                    min_value=0.0,
                    step=0.1,
                    format="%.2f m",
                    help="Vertical height from one finished floor level to the next."
                ),
                TYPICAL_UNIT_COL: st.column_config.NumberColumn(
                    "Typical Unit",
                    min_value=0,
                    step=1,
                    format="%d",
                    help="Number of typical units on this floor. This is a count, not an area."
                ),
                UNIT_AREA_COL: st.column_config.NumberColumn(
                    "Unit Area",
                    min_value=0.0,
                    step=10.0,
                    format="%.2f m²",
                    help="Total unit area on this floor."
                )
            }
        )

        # --- CLEAN EDITED DATA FIRST ---
        for col in breakdown_cols:
            if col not in edited_df.columns:
                edited_df[col] = 0.0
            edited_df[col] = pd.to_numeric(edited_df[col], errors="coerce").fillna(0.0)

        if TYPICAL_UNIT_COL not in edited_df.columns:
            edited_df[TYPICAL_UNIT_COL] = 0

        edited_df[TYPICAL_UNIT_COL] = (
            pd.to_numeric(edited_df[TYPICAL_UNIT_COL], errors="coerce")
            .fillna(0)
            .astype(int)
        )

        edited_df[F2F_COL] = pd.to_numeric(
            edited_df[F2F_COL],
            errors="coerce"
        ).fillna(0.0)

        # Keep only editable source columns
        editable_cols = ["FL", "Space Type", F2F_COL, TYPICAL_UNIT_COL] + breakdown_cols

        saved_area_records = edited_df[editable_cols].to_dict("records")

        # ✅ Save to session anchor (so it persists on next rerun)
        st.session_state[base_key] = saved_area_records

        # ✅ Also save to project data
        curr_proj["data"]["area_table_data"] = saved_area_records

        # Optional but recommended: persist to file
        save_data()

        # --- CALCULATIONS ---
        edited_df["TOTAL"] = edited_df[breakdown_cols].sum(axis=1)
        edited_df["GBA"] = edited_df["TOTAL"]
        edited_df["GFA"] = edited_df["TOTAL"] - edited_df[["Parkir", "Roof/Deck", "MEP Outdoor"]].sum(axis=1)
        edited_df["SGFA"] = edited_df[[UNIT_AREA_COL, "Office", "Koridor/Lobby"]].sum(axis=1)
        edited_df["NFA"] = edited_df[[UNIT_AREA_COL, "Office"]].sum(axis=1)

        # --- COST ESTIMATOR SYNC REFERENCES ---
        # Ruang (unit) = total Typical Unit from full Tab 1 table

        total_typical_units = (
            int(pd.to_numeric(edited_df[TYPICAL_UNIT_COL], errors="coerce").fillna(0).sum())
            if TYPICAL_UNIT_COL in edited_df.columns
            else 0
        )

        curr_proj["data"]["area_rooms_calc"] = total_typical_units
        curr_proj["data"]["area_typical_units_total_calc"] = total_typical_units

        # --- DETAIL AREA REFERENCES FOR COST ESTIMATOR SYNC ---
        # Lobby Interior for cost estimator = total Koridor/Lobby area from the full Tab 1 table.
        # Do not filter by Space Type, because corridor/lobby area usually appears on Unit floors too.

        total_lobby_interior = (
            _safe_float(edited_df["Koridor/Lobby"].sum())
            if "Koridor/Lobby" in edited_df.columns
            else 0.0
        )

        # Optional supporting references
        total_koridor_lobby_gba_source = total_lobby_interior

        # Store calculated references for future sync into cost estimator
        curr_proj["data"]["area_lobby_interior_calc"] = total_lobby_interior
        curr_proj["data"]["area_koridor_lobby_total_calc"] = total_lobby_interior

        st.divider()

        # ==========================================
        # VISUALIZATION DASHBOARD SECTION — PROFESSIONAL SETTINGS DASHBOARD STYLE
        # ==========================================

        col_viz_left, col_viz_right = st.columns([1.55, 1], gap="large")

        # --------------------------------------------------
        # PROFESSIONAL LIGHT CORPORATE PALETTE
        # Light enough for black text, but not too pale
        # --------------------------------------------------
        PRO_COLORS = {
            "Unit Area": "#9FBBD6",        # Main residential blue
            "Office": "#B5CEE5",           # Office blue
            "Koridor/Lobby": "#B6D0AA",    # Circulation sage
            "Stair, MEP, Etc": "#CCC7BE",  # Service warm gray
            "Parkir": "#AEB3BA",           # Parking gray
            "Roof/Deck": "#D1D8E2",        # Roof slate
            "MEP Outdoor": "#BDC5CF",      # Outdoor MEP gray
            "Lobby_Override": "#C8B4D2"    # Lobby lavender
        }

        PRO_TEXT = "#111827"
        PRO_MUTED = "#6B7280"
        PRO_BORDER = "#111827"
        PRO_CARD_BORDER = "#E5E7EB"
        PRO_BG = "#FFFFFF"
        PRO_PANEL_BG = "#F9FAFB"
        PRO_GRID = "#E5E7EB"

        area_cols = [
            "Office",
            UNIT_AREA_COL,
            "Koridor/Lobby",
            "Stair, MEP, Etc",
            "Parkir",
            "MEP Outdoor",
            "Roof/Deck"
        ]

        # --------------------------------------------------
        # DEFENSIVE NUMERIC CLEANUP
        # --------------------------------------------------
        for c in ["GBA", "GFA", "SGFA", "NFA"] + area_cols:
            if c not in edited_df.columns:
                edited_df[c] = 0
            edited_df[c] = pd.to_numeric(edited_df[c], errors="coerce").fillna(0)


        def safe_sum(df, col):
            return _safe_float(df[col].sum()) if col in df.columns else 0.0


        # ==================================================
        # LEFT: BUILDING AREA SECTION — STYLIZED BOX + FLOOR LINES
        # ==================================================
        with col_viz_left:
            # Stylized Streamlit container
            with st.container(border=True):
                st.markdown(
                    """
<div style="
padding: 2px 2px 4px 2px;
">
<div style="
font-size: 13px;
font-weight: 700;
letter-spacing: 0.04em;
text-transform: uppercase;
color: #111827;
margin-bottom: 2px;
">
Building Area Section
</div>
<div style="
font-size: 12px;
color: #6B7280;
margin-bottom: 8px;
">
Stacked floor composition by area category
</div>
</div>
                    """,
                    unsafe_allow_html=True
                )

                draw_df = edited_df.iloc[::-1].reset_index(drop=True).copy()

                floor_labels = draw_df["FL"].astype(str).tolist()
                y_positions = list(range(len(draw_df)))

                bases = {col: [] for col in area_cols}
                widths = {col: [] for col in area_cols}
                hover_texts = {col: [] for col in area_cols}
                text_labels = {col: [] for col in area_cols}
                unit_colors = []

                for _, row in draw_df.iterrows():
                    gba = _safe_float(row.get("GBA", 0))
                    sp_type = str(row.get("Space Type", ""))

                    curr_x = -gba / 2 if gba > 0 else 0

                    is_lobby = "Lobby" in sp_type
                    unit_colors.append(
                        PRO_COLORS["Lobby_Override"] if is_lobby else PRO_COLORS[UNIT_AREA_COL]
                    )

                    for col in area_cols:
                        val = _safe_float(row.get(col, 0))

                        if val > 0 and gba > 0:
                            bases[col].append(curr_x)
                            widths[col].append(val)

                            display_name = "Lobby" if col == UNIT_AREA_COL and is_lobby else col

                            hover_texts[col].append(
                                f"<b>{display_name}</b>"
                                f"<br>Area: {val:,.0f} m²"
                                f"<br>Floor GBA: {gba:,.0f} m²"
                                f"<br>Share: {(val / gba * 100):.1f}%"
                            )

                            if val / gba >= 0.115:
                                if col == "Koridor/Lobby":
                                    short_name = "Corridor"
                                elif col == "Stair, MEP, Etc":
                                    short_name = "Service"
                                elif col == "MEP Outdoor":
                                    short_name = "MEP Out."
                                elif col == "Roof/Deck":
                                    short_name = "Roof"
                                else:
                                    short_name = display_name

                                text_labels[col].append(f"{short_name}<br>{val:,.0f}")
                            else:
                                text_labels[col].append("")

                            curr_x += val
                        else:
                            bases[col].append(0)
                            widths[col].append(0)
                            hover_texts[col].append("")
                            text_labels[col].append("")

                fig_mass = go.Figure()

                floor_bar_height = 0.84

                for col in area_cols:
                    marker_color = unit_colors if col == UNIT_AREA_COL else PRO_COLORS.get(col, "#DDDDDD")

                    fig_mass.add_trace(
                        go.Bar(
                            y=y_positions,
                            x=widths[col],
                            base=bases[col],
                            width=floor_bar_height,
                            name=col,
                            orientation="h",
                            marker=dict(
                                color=marker_color,
                                line=dict(
                                    color="#111827",
                                    width=0.75
                                )
                            ),
                            text=text_labels[col],
                            textposition="inside",
                            insidetextanchor="middle",
                            hoverinfo="text",
                            hovertext=hover_texts[col],
                            textfont=dict(
                                color="#111827",
                                size=10.5,
                                family="Arial"
                            ),
                            cliponaxis=False
                        )
                    )

                max_gba = _safe_float(draw_df["GBA"].max()) if len(draw_df) else 0

                # --------------------------------------------------
                # FULL OUTER BUILDING OUTLINE PER FLOOR
                # Keeps every floor clearly boxed regardless of area composition
                # --------------------------------------------------
                for i, row in draw_df.iterrows():
                    gba = _safe_float(row.get("GBA", 0))

                    if gba > 0:
                        fig_mass.add_shape(
                            type="rect",
                            x0=-gba / 2,
                            x1=gba / 2,
                            y0=i - floor_bar_height / 2,
                            y1=i + floor_bar_height / 2,
                            line=dict(
                                color="#111827",
                                width=1.8
                            ),
                            fillcolor="rgba(0,0,0,0)",
                            layer="above"
                        )

                # --------------------------------------------------
                # HORIZONTAL FLOOR SEPARATION LINES
                # These read like floor slabs / level lines
                # --------------------------------------------------
                for i, row in draw_df.iterrows():
                    gba = _safe_float(row.get("GBA", 0))

                    if gba > 0:
                        # Top slab line
                        fig_mass.add_shape(
                            type="line",
                            x0=-gba / 2,
                            x1=gba / 2,
                            y0=i + floor_bar_height / 2,
                            y1=i + floor_bar_height / 2,
                            line=dict(
                                color="#111827",
                                width=2.2
                            ),
                            layer="above"
                        )

                        # Bottom slab line
                        fig_mass.add_shape(
                            type="line",
                            x0=-gba / 2,
                            x1=gba / 2,
                            y0=i - floor_bar_height / 2,
                            y1=i - floor_bar_height / 2,
                            line=dict(
                                color="#111827",
                                width=1.2
                            ),
                            layer="above"
                        )

                # --------------------------------------------------
                # Optional: subtle background frame inside the chart
                # --------------------------------------------------
                if max_gba > 0 and len(draw_df) > 0:
                    fig_mass.add_shape(
                        type="rect",
                        x0=-max_gba * 0.585,
                        x1=max_gba * 0.585,
                        y0=-0.75,
                        y1=len(draw_df) - 0.25,
                        line=dict(
                            color="#D1D5DB",
                            width=1
                        ),
                        fillcolor="rgba(249,250,251,0.45)",
                        layer="below"
                    )

                fig_mass.update_layout(
                    barmode="stack",
                    height=max(500, len(draw_df) * 40),
                    margin=dict(l=8, r=8, t=30, b=12),
                    plot_bgcolor="#FFFFFF",
                    paper_bgcolor="#FFFFFF",
                    hovermode="closest",
                    showlegend=True,
                    bargap=0.02,
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=1.015,
                        xanchor="left",
                        x=0,
                        font=dict(
                            size=10,
                            color="#6B7280"
                        ),
                        itemclick=False,
                        itemdoubleclick=False
                    ),
                    xaxis=dict(
                        range=[-max_gba * 0.60, max_gba * 0.60] if max_gba > 0 else None,
                        showgrid=False,
                        zeroline=False,
                        showticklabels=False,
                        title=None,
                        fixedrange=True
                    ),
                    yaxis=dict(
                        tickmode="array",
                        tickvals=y_positions,
                        ticktext=floor_labels,
                        showgrid=False,
                        tickfont=dict(
                            size=12,
                            color="#111827"
                        ),
                        title=None,
                        fixedrange=True
                    ),
                    uniformtext=dict(
                        minsize=8,
                        mode="hide"
                    )
                )

                # Ground datum line
                fig_mass.add_hline(
                    y=-0.5,
                    line_width=3,
                    line_color="#111827"
                )

                st.plotly_chart(
                    fig_mass,
                    use_container_width=True,
                    config={"displayModeBar": False}
                )


        # ==================================================
        # RIGHT: PROFESSIONAL DASHBOARD / SETTINGS PANEL
        # ==================================================
        with col_viz_right:
            total_gba = safe_sum(edited_df, "GBA")
            total_gfa = safe_sum(edited_df, "GFA")
            total_sgfa = safe_sum(edited_df, "SGFA")
            total_nfa = safe_sum(edited_df, "NFA")

            total_circ = safe_sum(edited_df, "Koridor/Lobby")
            total_serv = safe_sum(edited_df, "Stair, MEP, Etc")
            total_non_gfa = (
                safe_sum(edited_df, "Parkir")
                + safe_sum(edited_df, "Roof/Deck")
                + safe_sum(edited_df, "MEP Outdoor")
            )

            efficiency = (total_nfa / total_gfa * 100) if total_gfa > 0 else 0
            sgfa_ratio = (total_sgfa / total_gba * 100) if total_gba > 0 else 0
            gfa_ratio = (total_gfa / total_gba * 100) if total_gba > 0 else 0

            floor_count = len(edited_df)
            total_f2f_height = safe_sum(edited_df, F2F_COL)
            total_typical_units = int(edited_df[TYPICAL_UNIT_COL].sum()) if TYPICAL_UNIT_COL in edited_df.columns else 0

            composition = [
                ("NFA", total_nfa, PRO_COLORS[UNIT_AREA_COL]),
                ("Circulation", total_circ, PRO_COLORS["Koridor/Lobby"]),
                ("Services", total_serv, PRO_COLORS["Stair, MEP, Etc"]),
                ("Non-GFA", total_non_gfa, PRO_COLORS["Parkir"])
            ]

            st.markdown("##### PROJECT CONTROL PANEL")

            # --------------------------------------------------
            # SETTINGS DASHBOARD STYLE CSS
            # --------------------------------------------------
            st.markdown(
                f"""
<style>
.pro-panel {{
    background: {PRO_BG};
    border: 1px solid {PRO_CARD_BORDER};
    border-radius: 14px;
    padding: 14px;
    margin-bottom: 12px;
    box-shadow: 0 1px 2px rgba(16,24,40,0.04);
}}

.pro-panel-title {{
    font-size: 12px;
    font-weight: 700;
    color: {PRO_TEXT};
    letter-spacing: 0.04em;
    text-transform: uppercase;
    margin-bottom: 10px;
}}

.pro-metric-main {{
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    border-bottom: 1px solid {PRO_CARD_BORDER};
    padding-bottom: 10px;
    margin-bottom: 10px;
}}

.pro-metric-label {{
    font-size: 12px;
    color: {PRO_MUTED};
    text-transform: uppercase;
    letter-spacing: 0.035em;
    font-weight: 600;
}}

.pro-metric-value {{
    font-size: 24px;
    color: {PRO_TEXT};
    font-weight: 750;
    line-height: 1.1;
}}

.pro-row {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 8px 0;
    border-bottom: 1px solid #F0F2F5;
}}

.pro-row:last-child {{
    border-bottom: none;
}}

.pro-row-label {{
    font-size: 13px;
    color: {PRO_MUTED};
}}

.pro-row-value {{
    font-size: 13px;
    color: {PRO_TEXT};
    font-weight: 700;
}}

.pro-chip-row {{
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    margin-top: 8px;
}}

.pro-chip {{
    background: {PRO_PANEL_BG};
    border: 1px solid {PRO_CARD_BORDER};
    border-radius: 999px;
    padding: 5px 10px;
    font-size: 12px;
    color: {PRO_TEXT};
    font-weight: 600;
}}

.pro-bar-item {{
    margin-bottom: 12px;
}}

.pro-bar-top {{
    display: flex;
    justify-content: space-between;
    font-size: 12.5px;
    margin-bottom: 5px;
}}

.pro-bar-label {{
    color: {PRO_TEXT};
    font-weight: 650;
}}

.pro-bar-value {{
    color: {PRO_MUTED};
    font-weight: 650;
}}

.pro-bar-track {{
    height: 10px;
    background: #EEF1F5;
    border-radius: 999px;
    overflow: hidden;
    border: 1px solid #E5E7EB;
}}

.pro-bar-fill {{
    height: 100%;
    border-right: 1px solid rgba(17,24,39,0.35);
}}
</style>
                """,
                unsafe_allow_html=True
            )

            # --------------------------------------------------
            # MAIN KPI PANEL
            # --------------------------------------------------
            st.markdown(
                f"""
<div class="pro-panel">
<div class="pro-panel-title">Primary Area Metrics</div>

<div class="pro-metric-main">
<div>
<div class="pro-metric-label">Efficiency (NFA/GFA)</div>
<div class="pro-metric-value">{efficiency:.1f}%</div>
</div>
<div style="text-align:right;">
<div class="pro-metric-label">Gross Building Area (GBA)</div>
<div class="pro-metric-value">{total_gba:,.0f} m²</div>
</div>
</div>

<div class="pro-row">
<div class="pro-row-label">Gross Floor Area (GFA)</div>
<div class="pro-row-value">{total_gfa:,.0f} m²</div>
</div>

<div class="pro-row">
<div class="pro-row-label">Semi-Gross Floor Area (SGFA)</div>
<div class="pro-row-value">{total_sgfa:,.0f} m²</div>
</div>

<div class="pro-row">
<div class="pro-row-label">Net Floor Area (NFA)</div>
<div class="pro-row-value">{total_nfa:,.0f} m²</div>
</div>

<div class="pro-chip-row">
<div class="pro-chip">{floor_count} Floors</div>
<div class="pro-chip">{total_typical_units:,} Typical Units</div>
<div class="pro-chip">Total Height {total_f2f_height:.1f} m</div>
<div class="pro-chip">GFA/GBA {gfa_ratio:.1f}%</div>
<div class="pro-chip">SGFA/GBA {sgfa_ratio:.1f}%</div>
</div>
</div>
                """,
                unsafe_allow_html=True
            )

            # --------------------------------------------------
            # AREA COMPOSITION PANEL
            # --------------------------------------------------
            comp_html = """
            <div class="pro-panel">
                <div class="pro-panel-title">Area Composition</div>
            """

            for label, value, color in composition:
                pct = (value / total_gba * 100) if total_gba > 0 else 0
                bar_width = min(max(pct, 0), 100)

                comp_html += f"""
<div class="pro-bar-item">
    <div class="pro-bar-top">
        <div class="pro-bar-label">{label}</div>
        <div class="pro-bar-value">{value:,.0f} m² · {pct:.1f}%</div>
    </div>
    <div class="pro-bar-track">
        <div class="pro-bar-fill" style="width:{bar_width:.1f}%; background:{color};"></div>
    </div>
</div>
                """

            comp_html += "</div>"

            st.markdown(comp_html, unsafe_allow_html=True)

            # --------------------------------------------------
            # COMPACT REFERENCE CHART
            # --------------------------------------------------
            st.markdown("##### DISTRIBUTION REFERENCE")

            categories = [x[0] for x in composition]
            percentages = [
                (x[1] / total_gba * 100) if total_gba > 0 else 0
                for x in composition
            ]
            chart_colors = [x[2] for x in composition]

            fig_dist = go.Figure()

            fig_dist.add_trace(
                go.Bar(
                    x=percentages,
                    y=categories,
                    orientation="h",
                    marker=dict(
                        color=chart_colors,
                        line=dict(
                            color=PRO_BORDER,
                            width=1
                        )
                    ),
                    text=[f"{p:.1f}%" for p in percentages],
                    textposition="outside",
                    textfont=dict(
                        color=PRO_TEXT,
                        size=11
                    ),
                    hovertemplate="<b>%{y}</b><br>%{x:.1f}% of GBA<extra></extra>"
                )
            )

            fig_dist.update_layout(
                height=260,
                margin=dict(l=8, r=42, t=8, b=8),
                plot_bgcolor=PRO_BG,
                paper_bgcolor=PRO_BG,
                showlegend=False,
                xaxis=dict(
                    showgrid=True,
                    gridcolor=PRO_GRID,
                    zeroline=False,
                    ticksuffix="%",
                    tickfont=dict(
                        size=10,
                        color=PRO_MUTED
                    ),
                    range=[0, max(percentages) * 1.22 if max(percentages) > 0 else 100],
                    fixedrange=True
                ),
                yaxis=dict(
                    tickfont=dict(
                        size=11,
                        color=PRO_TEXT
                    ),
                    autorange="reversed",
                    fixedrange=True
                )
            )

            st.plotly_chart(
                fig_dist,
                use_container_width=True,
                config={"displayModeBar": False}
            )

    with tab2:
        st.subheader("2. Area Details")
        st.caption("Calculated detail areas from Tab 1. These values are prepared for later sync into the cost estimator.")

        st.markdown("##### Total Area Summary From Tab 1")

        summary_cols = [
            TYPICAL_UNIT_COL,
            "Parkir",
            "Roof/Deck",
            "MEP Outdoor",
            "Koridor/Lobby",
            "Stair, MEP, Etc",
            UNIT_AREA_COL,
            "Office",
            "GBA",
            "GFA",
            "SGFA",
            "NFA",
        ]

        summary_cols = [c for c in summary_cols if c in edited_df.columns]

        total_summary = pd.DataFrame([
            {
                "Summary": "TOTAL",
                **{
                    col: (
                        int(edited_df[col].sum())
                        if col == TYPICAL_UNIT_COL
                        else _safe_float(edited_df[col].sum())
                    )
                    for col in summary_cols
                }
            }
        ])

        st.dataframe(
            total_summary,
            use_container_width=True,
            hide_index=True
        )

        st.markdown("##### Area Summary by Space Type")

        group_cols = [
            TYPICAL_UNIT_COL,
            "Parkir",
            "Roof/Deck",
            "MEP Outdoor",
            "Koridor/Lobby",
            "Stair, MEP, Etc",
            UNIT_AREA_COL,
            "Office",
            "GBA",
            "GFA",
            "SGFA",
            "NFA",
        ]

        group_cols = [c for c in group_cols if c in edited_df.columns]

        area_by_space_type = (
            edited_df
            .groupby("Space Type", dropna=False)[group_cols]
            .sum()
            .reset_index()
        )

        if TYPICAL_UNIT_COL in area_by_space_type.columns:
            area_by_space_type[TYPICAL_UNIT_COL] = area_by_space_type[TYPICAL_UNIT_COL].astype(int)

        st.dataframe(
            area_by_space_type,
            use_container_width=True,
            hide_index=True
        )

        st.markdown("##### Cost Analysis Sync Sources")

        c1, c2, c3 = st.columns(3)

        c1.metric(
            "Lobby Interior Sync Source",
            f"{total_lobby_interior:,.0f} m²",
            help="Calculated from total Koridor/Lobby in Tab 1."
        )

        c2.metric(
            "Ruang / Units Sync Source",
            f"{total_typical_units:,} unit",
            help="Calculated from total Typical Unit in Tab 1."
        )

        c3.metric(
            "Future Targets",
            "m_lobby / m_rooms",
            help="These values will be copied into Cost Analysis inputs by Use Area Analysis."
        )

        st.info(
            f"Sync sources prepared: "
            f"`area_lobby_interior_calc = {total_lobby_interior:,.0f} m²`, "
            f"`area_rooms_calc = {total_typical_units:,} unit`."
        )

def update_price(metric_key, db_key): #this function pulls~
    """Update flooring price based on spec radio selection."""
    c_id = st.session_state.current_proj_id
    
    # 1. Safety check: Ensure the project actually exists
    if c_id not in st.session_state.projects:
        return 

    p_type = st.session_state.projects[c_id]["type"]
    c_type_key = f"{c_id}_{p_type}"
    widget_key = f"temp_spec_{metric_key}_{c_type_key}"
    
    # 2. Use .get() to securely fetch the value without throwing a KeyError
    selected_spec = st.session_state.get(widget_key)
    
    # 3. If the widget was destroyed or doesn't exist, quietly exit
    if selected_spec is None:
        return

    # 4. Proceed with normal update
    st.session_state.projects[c_id]["data"][f"{metric_key}_spec_type"] = selected_spec
    db_val = PROJECT_DATABASE.get(p_type, {}).get(db_key, {})
    
    if isinstance(db_val, dict):
        new_val = db_val.get(selected_spec, 0.0)
        st.session_state[f"u_fl_{metric_key}_{c_type_key}"] = _safe_float(new_val)

def show_cost_estimator(): #cost calculator page
    st.title("Cost Analysis")

    st.markdown("""
        <style>
            .metric-container {
                position: relative;
                display: inline-block;
                width: 100%;
            }
            .custom-tooltip {
                visibility: hidden;
                width: 240px;
                background-color: #FFFFFF;
                color: #262730;
                text-align: left;
                border-radius: 8px;
                padding: 12px;
                position: absolute;
                z-index: 1000;
                bottom: 110%; 
                left: 50%;
                transform: translateX(-50%);
                opacity: 0;
                transition: opacity 0.3s;
                font-size: 12px;
                border: 1px solid #CDDC39;
                box-shadow: 0px 4px 15px rgba(0,0,0,0.4);
                line-height: 1.5;
            }
            .metric-container:hover .custom-tooltip {
                visibility: visible;
                opacity: 1;
            }
            /* Triangle arrow for tooltip */
            .custom-tooltip::after {
                content: "";
                position: absolute;
                top: 100%;
                left: 50%;
                margin-left: -5px;
                border-width: 5px;
                border-style: solid;
                border-color: #262730 transparent transparent transparent;
            }
        </style>
        """, unsafe_allow_html=True)

    curr_id, curr_proj = get_current_project()

    if "data" not in curr_proj:
        st.session_state.projects[curr_id]["data"] = {}

    def get_val(key, default=0.0):
        data_dict = st.session_state.projects[curr_id]["data"]
        val = data_dict.get(key, default)
        
        # If the value is a list (for Custom Items), return it immediately
        if isinstance(val, list):
            return val
            
        # For everything else (numbers/strings), try to force to _safe_float
        try:
            return _safe_float(val)
        except (ValueError, TypeError):
            # If it's not a number (like "Type1"), return it as is
            return val

    # --- PROJECT SETUP ---
    curr_type = curr_proj["type"]
    pt_data = PROJECT_DATABASE[curr_type]
    curr_type_key = f"{curr_id}_{curr_type}"

    # --- TABS ---
    tab2, tab3, tab5, tab4, tab6, tab7, tab8 = st.tabs([
        "1. Ukuran", "2. Rasio",
        "3. Soft Costs", "4. Harga",
        "5. Item Tambahan", "6. Hasil",
        "7. Pembuktian",
    ])

    with tab2:

        # ==================================================
        # IMPORT AREA ANALYSIS TOTALS INTO COST ANALYSIS
        # ==================================================
        area_table_data = curr_proj.get("data", {}).get("area_table_data", [])
        area_totals = calculate_area_totals_from_table(area_table_data)
        area_land = _safe_float(curr_proj.get("data", {}).get("m_land", 0.0))
        area_lobby_interior = _safe_float(curr_proj.get("data", {}).get("area_lobby_interior_calc", 0.0))
        area_rooms = int(_safe_float(curr_proj.get("data", {}).get("area_rooms_calc", 0)))

        if isinstance(area_table_data, list) and len(area_table_data) > 0:
            area_sync_df = pd.DataFrame(area_table_data)
            if "Koridor/Lobby" in area_sync_df.columns:
                area_lobby_interior = _safe_float(
                    pd.to_numeric(
                        area_sync_df["Koridor/Lobby"],
                        errors="coerce"
                    ).fillna(0.0).sum()
                )
            if "Typical Unit" in area_sync_df.columns:
                area_rooms = int(
                    pd.to_numeric(
                        area_sync_df["Typical Unit"],
                        errors="coerce"
                    ).fillna(0).sum()
                )

        sync_col1, sync_col2 = st.columns([1, 3])

        with sync_col1:
            if st.button(
                "Use Area Analysis",
                key=f"use_area_analysis_{curr_id}",
                type="primary",
                use_container_width=True,
                icon=mi("sync") if "mi" in globals() else None,
                help="Replace GBA, GFA, SGFA, NFA, Lobby Interior, and Ruang using the current component's Area Analysis table."
            ):

                st.session_state.projects[curr_id]["data"]["m_land"] = area_land
                st.session_state.projects[curr_id]["data"]["m_gba"] = area_totals["gba"]
                st.session_state.projects[curr_id]["data"]["m_gfa"] = area_totals["gfa"]
                st.session_state.projects[curr_id]["data"]["m_sgfa"] = area_totals["sgfa"]
                st.session_state.projects[curr_id]["data"]["m_nfa"] = area_totals["nfa"]
                st.session_state.projects[curr_id]["data"]["m_lobby"] = area_lobby_interior
                st.session_state.projects[curr_id]["data"]["m_rooms"] = area_rooms

                st.session_state[f"m_land_{curr_id}"] = area_land
                st.session_state[f"m_gba_{curr_id}"] = area_totals["gba"]
                st.session_state[f"m_gfa_{curr_id}"] = area_totals["gfa"]
                st.session_state[f"m_sgfa_{curr_id}"] = area_totals["sgfa"]
                st.session_state[f"m_lobby_{curr_id}"] = area_lobby_interior
                st.session_state[f"m_rooms_{curr_id}"] = area_rooms

                save_data()
                st.success("Area Analysis totals, lobby area, and rooms applied to Cost Analysis.")
                st.rerun()

        with sync_col2:
            st.caption(
                f"Area Analysis totals: "
                f"Land {area_land:,.0f} m² | "
                f"GBA {area_totals['gba']:,.0f} m² | "
                f"GFA {area_totals['gfa']:,.0f} m² | "
                f"SGFA {area_totals['sgfa']:,.0f} m² | "
                f"NFA {area_totals['nfa']:,.0f} m² | "
                f"Lobby {area_lobby_interior:,.0f} m² | "
                f"Ruang {area_rooms:,} unit"
            )

        st.divider()
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)

        with col_m1:
            with st.expander("Ukuran Proyek", expanded=True):
                st.subheader("Ukuran")
                land_area = st.number_input("Luas Tanah (m2)", value=get_val("m_land", 0.0), step=100.0, key=f"m_land_{curr_id}")
                gba = st.number_input("GBA (m2)", value=get_val("m_gba", 0.0), step=100.0, key=f"m_gba_{curr_id}")
                gfa = st.number_input("GFA (m2)", value=get_val("m_gfa", 0.0), step=100.0, key=f"m_gfa_{curr_id}")
                sgfa = st.number_input("SGFA (m2)", value=get_val("m_sgfa", 0.0), step=100.0, key=f"m_sgfa_{curr_id}")

        with col_m2:
            with st.expander("Arsitektur", expanded=True):
                st.subheader("Interior")
                rooms = st.number_input(
                    "Ruang (unit)",
                    help="cth. 500 unit untuk 1 proyek Apartement A",
                    value=int(_safe_float(get_val("m_rooms", 0))),
                    step=1,
                    key=f"m_rooms_{curr_id}"
                )

                lobby_interior = st.number_input(
                    "Lobby Interior (m2)",
                    help="cth. 500 m2 lobby untuk 1 proyek Apartement A",
                    value=_safe_float(get_val("m_lobby", 0.0)),
                    step=10.0,
                    key=f"m_lobby_{curr_id}"
                )

                carpet_m2 = st.number_input("Karpet (m2)", value=get_val("m_carpet", 0.0), step=10.0, key=f"m_carpet_{curr_id}")
                glass_m2 = st.number_input("Kaca (m2)", value=get_val("m_glass", 0.0), step=10.0, key=f"m_glass_{curr_id}")
                st.subheader("Eksterior")
                facade = st.number_input("Facade (m2)", value=get_val("m_facade", 0.0), step=100.0, key=f"m_facade_{curr_id}")
                gondola_unit = st.number_input("Gondola (unit)", value=get_val("m_gondola", 0.0), step=1.0, key=f"m_gondola_{curr_id}")
                skylight_area = st.number_input("Skylight (m2)", value=get_val("m_skylight", 0.0), step=10.0, key=f"m_skylight_{curr_id}")    
                st.subheader("Pintu")
                glass_door = st.number_input("Glass Door (unit)", value=get_val("m_door_g", 0.0), step=1.0, key=f"m_door_g_{curr_id}")
                wooden_door = st.number_input("Wooden Door (unit)", value=get_val("m_door_w", 0.0), step=10.0, key=f"m_door_w_{curr_id}")
                steel_door = st.number_input("Steel Door (unit)", value=get_val("m_door_s", 0.0), step=10.0, key=f"m_door_s_{curr_id}")

        with col_m3:
            with st.expander("Sanitari", expanded=True):
                st.subheader("Toilet Unit")
                san_qty_room = st.number_input("Toilet Private (unit/ruang)", help="Cth. 3 Toilet/1 Kamar (Apt)", value=get_val("r_san_qty", pt_data["san_room_qty"]), key=f"r_san_qty_{curr_type_key}")
                st.subheader("Toilet Umum")
                toilet_male = st.number_input("Toilet Umum - Pria (units)", value=get_val("m_toil_m", 0.0), step=1.0, key=f"m_toil_m_{curr_id}")
                toilet_female = st.number_input("Toilet Umum - Wanita (units)", value=get_val("m_toil_f", 0.0), step=1.0, key=f"m_toil_f_{curr_id}")
                disabled_toil = st.number_input("Toilet Difabel (units)", value=get_val("m_toil_d", 0.0), step=1.0, key=f"m_toil_d_{curr_id}")
                st.subheader("Mushola")
                mushola_unit = st.number_input("Mushola (units)", value=get_val("m_mushola", 0.0), step=1.0, key=f"m_mushola_{curr_id}")

        with col_m4:
            with st.expander("Fasilitas", expanded=True):
                st.subheader("Fasilitas")
                res_fac_m2 = st.number_input("Fasilitas Penghuni (m2)", value=get_val("m_fac_res", 0.0), step=10.0, key=f"m_fac_res_{curr_id}")
                pub_fac_m2 = st.number_input("Fasilitas Publik (m2)", value=get_val("m_fac_pub", 0.0), step=10.0, key=f"m_fac_pub_{curr_id}")
                proj_fac_u = st.number_input("Fasilitas Proyek (unit)", value=get_val("m_fac_proj", 0.0), step=1.0, key=f"m_fac_proj_{curr_id}")
                land_m2 = st.number_input("Area Lanskap (m2)", value=get_val("m_land_m2", 0.0), step=100.0, key=f"m_land_m2_{curr_id}")
                st.subheader("Miscellaneous")
                m_status = st.radio("Ada Gym/Linen?", ["Tidak", "Ada"],
                                    index=1 if get_val("misc_switch", 0) == 1 else 0,
                                    key=f"misc_sw_{curr_id}", horizontal=True)
                misc_switch = 1 if m_status == "Ada" else 0
                st.session_state.projects[curr_id]["data"]["misc_switch"] = misc_switch

    # --- TAB 2: RATIOS & MULTIPLIERS ---
    with tab3:
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            st.subheader("Facade Ratio (%)")
            facade_precast_pct = st.number_input("Precast (%)", value=get_val("r_fac_pre", pt_data["facade_precast_pct"]), step=5.0, key=f"r_fac_pre_{curr_type_key}")
            facade_window_pct = st.number_input("Window Wall (%)", value=get_val("r_fac_win", pt_data["facade_window_pct"]), step=5.0, key=f"r_fac_win_{curr_type_key}")
            facade_double_pct = st.number_input("Double Skin (%)", value=get_val("r_fac_doub", pt_data["facade_double_pct"]), step=5.0, key=f"r_fac_doub_{curr_type_key}")
            t_fac_pct = facade_precast_pct + facade_window_pct + facade_double_pct

            if t_fac_pct != 100:
                st.warning(f"⚠️ Total is **{t_fac_pct}%** (bukan 100%)")
        with col_r3:
            st.subheader("Waste & Skirting (%)")
            fl_skirt = st.number_input(
                "Skirting (%)", 
                value=get_val("s_floor", pt_data.get("fl_skirt", 20)), 
                key=f"s_floor{curr_type_key}"
            )
            fl_waste = st.number_input(
                "Floor Waste (%)", 
                value=get_val("w_floor", pt_data.get("fl_waste", 10)), # Use .get() with 1.1 as default
                key=f"w_floor{curr_type_key}"
            )        

            st.caption(f"Luas Lantai + Skirting + Waste: GFA: {gfa:.2f} x {1 + (fl_skirt/100):.2f} x {1 + (fl_waste/100):.2f} = {gfa*(1 + (fl_waste/100))*(1 + (fl_skirt/100)):.2f} m2")

            f_mult = (1 + (fl_waste/100)) * (1 + (fl_skirt/100))
        
        with col_r2:
            st.subheader("Flooring Ratio (%)")
            fl_ht_pct = st.number_input("HT/Ceramic Tile (%)", value=get_val("r_fl_ht", pt_data["fl_ht_pct"]), step=5.0, key=f"r_fl_ht_{curr_type_key}")
            fl_vinyl_pct = st.number_input("Vinyl (%)", value=get_val("r_fl_vin", pt_data["fl_vinyl_pct"]), step=5.0, key=f"r_fl_vin_{curr_type_key}")
            fl_marmer_pct = st.number_input("Marmer (%)", value=get_val("r_fl_mar", pt_data["fl_marmer_pct"]), step=5.0, key=f"r_fl_mar_{curr_type_key}")
            t_fl_pct = fl_ht_pct + fl_vinyl_pct + fl_marmer_pct

            if t_fl_pct != 100:
                st.warning(f"⚠️ Total is **{t_fl_pct}%** (bukan 100%)")

        with col_r4:
            st.subheader("Railing (m')")
            railing_qty = st.number_input("Railing Length (m'/room)", value=get_val("r_rail_qty", pt_data["railing_qty"]), step=1.0, key=f"r_rail_qty_{curr_type_key}")
            st.caption(f"Total: {railing_qty:.0f} m x {rooms:.0f} unit = {railing_qty * rooms:.0f} m'")

    # --- TAB 3: UNIT RATES ---
    with tab4:
        with st.expander("Harga Fondasi & Struktur", expanded=True):
            c1, c2, c3 = st.columns(3)
            struc_earth = c1.number_input("Earthwork Rate (Rp)", value=get_val("u_earth", pt_data["struc_earth"]), key=f"u_earth_{curr_type_key}")
            struc_found = c2.number_input("Foundation Rate (Rp)", value=get_val("u_found", pt_data["struc_found"]), key=f"u_found_{curr_type_key}")
            struc_work = c3.number_input("Structural Work Rate (Rp)", value=get_val("u_struc", pt_data["struc_work"]), key=f"u_struc_{curr_type_key}")
            #caption
            c1.caption(f"""Hitungan: Rp {struc_earth:,.0f} x GBA: {gba:,.0f} m2  \n  Total Earthwork: Rp {struc_earth * gba:,.0f}  \n  Terbilang: {n2w(struc_earth * gba)}""")
            c2.caption(f"""Hitungan: Rp {struc_found:,.0f} x GBA: {gba:,.0f} m2  \n  Total Foundation: Rp {struc_found * gba:,.0f}  \n  Terbilang: {n2w(struc_found * gba)}""")
            c3.caption(f"""Hitungan: Rp {struc_work:,.0f} x GBA: {gba:,.0f} m2  \n  Total Structural Work: Rp {struc_work * gba:,.0f}  \n  Terbilang: {n2w(struc_work * gba)}""")

        with st.expander("Arsitektur & Fasad"):
            c1, c2 = st.columns(2)
            arch_base = c1.number_input("Architecture Base (Rp)", value=get_val("u_arch", pt_data["arch_base"]), key=f"u_arch_{curr_type_key}")
            lobby_rate = c2.number_input("Lobby Interior Rate (Rp)", value=get_val("u_lobby", pt_data["lobby"]), key=f"u_lobby_{curr_type_key}")
            c3, c4, c5 = st.columns(3)
            fac_precast_rate = c3.number_input("Precast Rate (Rp)", value=get_val("u_f_pre", pt_data["facade_precast_rate"]), key=f"u_f_pre_{curr_type_key}")
            fac_window_rate = c4.number_input("Window Wall Rate (Rp)", value=get_val("u_f_win", pt_data["facade_window_rate"]), key=f"u_f_win_{curr_type_key}")
            fac_double_rate = c5.number_input("Double Skin Rate (Rp)", value=get_val("u_f_doub", pt_data["facade_double_rate"]), key=f"u_f_doub_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {arch_base:,.0f} x GFA: {gfa:,.0f} m2  \n  Total Architecture Base: Rp {arch_base * gfa:,.0f}  \n  Terbilang: {n2w(arch_base * gfa)}""")
            c2.caption(f"""Hitungan: Rp {lobby_rate:,.0f} x Lobby Interior: {lobby_interior:,.0f} m2  \n  Total Lobby Interior: Rp {lobby_rate * lobby_interior:,.0f}  \n  Terbilang: {n2w(lobby_rate * lobby_interior)}""")
            c3.caption(f"""Hitungan: Rp {fac_precast_rate:,.0f} x Facade: {facade:,.0f} m2 x {facade_precast_pct}%  \n  Total Precast: Rp {fac_precast_rate * facade * (facade_precast_pct/100):,.0f}  \n  Terbilang: {n2w(fac_precast_rate * facade * (facade_precast_pct/100))}""")
            c4.caption(f"""Hitungan: Rp {fac_window_rate:,.0f} x Facade: {facade:,.0f} m2 x {facade_window_pct}%  \n  Total Window Wall: Rp {fac_window_rate * facade * (facade_window_pct/100):,.0f}  \n  Terbilang: {n2w(fac_window_rate * facade * (facade_window_pct/100))}""")
            c5.caption(f"""Hitungan: Rp {fac_double_rate:,.0f} x Facade: {facade:,.0f} m2 x {facade_double_pct}%  \n  Total Double Skin: Rp {fac_double_rate * facade * (facade_double_pct/100):,.0f}  \n  Terbilang: {n2w(fac_double_rate * facade * (facade_double_pct/100))}""")
        with st.expander("Pintu dan Hardware"):
            c1, c2, c3 = st.columns(3)
            door_wood = c1.number_input("Wooden Door Rate (Rp)", value=get_val("u_d_wood", pt_data["door_wood"]), key=f"u_d_wood_{curr_type_key}")
            door_glass = c3.number_input("Glass Door Rate (Rp)", value=get_val("u_d_glass", pt_data["door_glass"]), key=f"u_d_glass_{curr_type_key}")
            door_steel = c2.number_input("Steel Door Rate (Rp)", value=get_val("u_d_steel", pt_data["door_steel"]), key=f"u_d_steel_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {door_wood:,.0f} x  {wooden_door:,.0f} unit  \n  Total Pintu Kayu: Rp {door_wood * wooden_door:,.0f}  \n  Terbilang: {n2w(door_wood * wooden_door)}""")
            c3.caption(f"""Hitungan: Rp {door_glass:,.0f} x  {glass_door:,.0f} unit  \n  Total Pintu Kaca: Rp {door_glass * glass_door:,.0f}  \n  Terbilang: {n2w(door_glass * glass_door)}""")
            c2.caption(f"""Hitungan: Rp {door_steel:,.0f} x  {steel_door:,.0f} unit  \n  Total Pintu Baja: Rp {door_steel * steel_door:,.0f}  \n  Terbilang: {n2w(door_steel * steel_door)}""")
            hw_wood = c1.number_input("Hardware Wooden Door (Rp)", value=get_val("u_hw_wood", pt_data["hw_wood"]), key=f"u_hw_wood_{curr_type_key}")
            hw_steel = c2.number_input("Hardware Steel Door (Rp)", value=get_val("u_hw_steel", pt_data["hw_steel"]), key=f"u_hw_steel_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {hw_wood:,.0f} x  {wooden_door:,.0f} unit  \n  Total Hardware Pintu Kayu: Rp {hw_wood * wooden_door:,.0f}  \n  Terbilang: {n2w(hw_wood * wooden_door)}""")
            c2.caption(f"""Hitungan: Rp {hw_steel:,.0f} x  {steel_door:,.0f} unit  \n  Total Hardware Pintu Baja: Rp {hw_steel * steel_door:,.0f}  \n  Terbilang: {n2w(hw_steel * steel_door)}""")

        with st.expander("Sanitari"):
            c1, c2 = st.columns(2)
            san_room_rate = c1.number_input("Typical Unit Sanitary Rate (Rp)", value=get_val("u_s_room", pt_data["san_room_rate"]), key=f"u_s_room_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {san_room_rate:,.0f} x {rooms:,.0f} Rooms x {san_qty_room} Units / Room  \n  Total Private Bathroom: Rp {rooms * san_qty_room * san_room_rate:,.0f}  \n  Terbilang: {n2w(rooms * san_qty_room * san_room_rate)}""")
            san_pub_m = c2.number_input("Public Toilet Male Rate (Rp)", value=get_val("u_s_pub_m", pt_data["san_pub_m"]), key=f"u_s_pub_m_{curr_type_key}")
            c2.caption(f"""Hitungan: Rp {san_pub_m:,.0f} x {toilet_male:,.0f} Units  \n  Total Public Toilet Male: Rp {toilet_male * san_pub_m:,.0f}  \n  Terbilang: {n2w(toilet_male * san_pub_m)}""")
            san_pub_f = c2.number_input("Public Toilet Female Rate (Rp)", value=get_val("u_s_pub_f", pt_data["san_pub_f"]), key=f"u_s_pub_f_{curr_type_key}")
            c2.caption(f"""Hitungan: Rp {san_pub_f:,.0f} x {toilet_female:,.0f} Units  \n  Total Public Toilet Female: Rp {toilet_female * san_pub_f:,.0f}  \n  Terbilang: {n2w(toilet_female * san_pub_f)}""")
            san_dis = c1.number_input("Disabled Toilet Rate (Rp)", value=get_val("u_s_dis", pt_data["san_dis"]), key=f"u_s_dis_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {san_dis:,.0f} x {disabled_toil:,.0f} Units  \n  Total Toilet Difabel: Rp {disabled_toil * san_dis:,.0f}  \n  Terbilang: {n2w(disabled_toil * san_dis)}""")
            san_mushola = c1.number_input("Mushola Rate (Rp)", value=get_val("u_s_mushola", pt_data["san_mushola"]), key=f"u_s_mushola_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {san_mushola:,.0f} x {mushola_unit:,.0f} Units  \n  Total Mushola: Rp {mushola_unit * san_mushola:,.0f}  \n  Terbilang: {n2w(mushola_unit * san_mushola)}""")
        
        with st.expander("Lantai, Finishing, dan Interior"):
            st.subheader("Harga ")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.radio("Spek HT", ["Type1", "Type2"],
                    key=f"temp_spec_ht_{curr_type_key}",
                    horizontal=True,
                    on_change=update_price, args=("ht", "fl_ht_rate"))
                fl_ht_rate = st.number_input("HT Rate (Rp)",
                                            value=get_val("u_fl_ht", pt_data["fl_ht_rate"]["Type1"]),
                                            key=f"u_fl_ht_{curr_type_key}")
                c1.caption(f"""Hitungan: {fl_ht_pct}% of GFA x Rp {fl_ht_rate:,.0f} x {gfa:,.0f} m2 x {f_mult}  \n  Total HT: Rp {gfa * (fl_ht_pct / 100) * fl_ht_rate * f_mult:,.0f}  \n  Terbilang: {n2w(gfa * (fl_ht_pct / 100) * fl_ht_rate * f_mult)}""")

            with c2:
                st.radio("Spek Vinyl", ["Type1", "Type2"],
                    key=f"temp_spec_vin_{curr_type_key}",
                    horizontal=True,
                    on_change=update_price, args=("vin", "fl_vinyl_rate"))
                fl_vinyl_rate = st.number_input("Vinyl Rate (Rp)",
                                                value=get_val("u_fl_vin", pt_data["fl_vinyl_rate"]["Type1"]),
                                                key=f"u_fl_vin_{curr_type_key}")
                c2.caption(f"""Hitungan: {fl_vinyl_pct}% of GFA x Rp {fl_vinyl_rate:,.0f} x {gfa:,.0f} m2 x {f_mult}  \n  Total Vinyl: Rp {gfa * (fl_vinyl_pct / 100) * fl_vinyl_rate * f_mult:,.0f}  \n  Terbilang: {n2w(gfa * (fl_vinyl_pct / 100) * fl_vinyl_rate * f_mult)}""")
            
            with c3:
                st.radio("Spek Marmer", ["Type1", "Type2"],
                    key=f"temp_spec_mar_{curr_type_key}",
                    horizontal=True,
                    on_change=update_price, args=("mar", "fl_marmer_rate"))
                fl_marmer_rate = st.number_input("Marmer Rate (Rp)",
                                                value=get_val("u_fl_mar", pt_data["fl_marmer_rate"]["Type1"]),
                                                key=f"u_fl_mar_{curr_type_key}")
                c3.caption(f"""Hitungan: {fl_marmer_pct}% of GFA x Rp {fl_marmer_rate:,.0f} x {gfa:,.0f} m2 x {f_mult}  \n  Total Marmer: Rp {gfa * (fl_marmer_pct / 100) * fl_marmer_rate * f_mult:,.0f}  \n  Terbilang: {n2w(gfa * (fl_marmer_pct / 100) * fl_marmer_rate * f_mult)}""")
                
            carpet_rate = c1.number_input("Carpet Rate (Rp)", value=get_val("u_carpet", pt_data["carpet"]), key=f"u_carpet_{curr_type_key}")
            c1.caption(f"""Hitungan: {carpet_m2:,.0f} m2 x Rp {carpet_rate:,.0f}  \n  Total Carpet Work: Rp {carpet_m2 * carpet_rate:,.0f}  \n  Terbilang: {n2w(carpet_m2 * carpet_rate)}""")
            glass_rate = c2.number_input("Glass Work Rate (Rp)", value=get_val("u_glass", pt_data["glass"]), key=f"u_glass_{curr_type_key}")
            c2.caption(f"""Hitungan: {glass_m2:,.0f} m2 x Rp {glass_rate:,.0f}  \n  Total Glass Work: Rp {glass_m2 * glass_rate:,.0f}  \n  Terbilang: {n2w(glass_m2 * glass_rate)}""")            
            skylight_rate = c3.number_input("Skylight Rate (Rp)", value=get_val("u_sky", pt_data["skylight_rate"]), key=f"u_sky_{curr_type_key}")
            c3.caption(f"""Hitungan: {skylight_area:,.0f} m2 Total x Rp {skylight_rate:,.0f}  \n  Total Skylight Work: Rp {skylight_area * skylight_rate:,.0f}  \n  Terbilang: {n2w(skylight_area * skylight_rate)}""")
            gondola_rate = c1.number_input("Gondola Rate (Rp)", value=get_val("u_gondola", pt_data["gondola"]), key=f"u_gondola_{curr_type_key}")
            c1.caption(f"""Hitungan: {gondola_unit:,.0f} Units x Rp {gondola_rate:,.0f}  \n  Total Gondola: Rp {gondola_unit * gondola_rate:,.0f}  \n  Terbilang: {n2w(gondola_unit * gondola_rate)}""")
            railing_rate = c2.number_input("Railing Rate (Rp)", value=get_val("u_rail", pt_data["railing_rate"]), key=f"u_rail_{curr_type_key}")
            c2.caption(f"""Hitungan: {rooms * railing_qty:,.0f} m' Total x Rp {railing_rate:,.0f}  \n  Total Railing Work: Rp {rooms * railing_qty * railing_rate:,.0f}  \n  Terbilang: {n2w(rooms * railing_qty * railing_rate)}""")

        with st.expander("MEP, Dapur dan FF&E"):
            c1, c2 = st.columns(2)
            mep_rate = c1.number_input("MEP Works (Rp)", value=get_val("u_mep", pt_data["mep"]), key=f"u_mep_{curr_type_key}")
            c1.caption(f"""Hitungan: {gba:,.0f} m2 x Rp {mep_rate:,.0f}  \n  Total MEP Works: Rp {gba * mep_rate:,.0f}  \n  Terbilang: {n2w(gba * mep_rate)}""")
            utility_rate = c2.number_input("Utility Connection (Rp)", value=get_val("u_util", pt_data["utility"]), key=f"u_util_{curr_type_key}")
            c2.caption(f"""Hitungan: {gba:,.0f} m2 x Rp {utility_rate:,.0f}  \n  Total Utility Connection: Rp {gba * utility_rate:,.0f}  \n  Terbilang: {n2w(gba * utility_rate)}""")

            ffe_rate = c1.number_input("FF&E (Rp)", value=get_val("u_ffe", pt_data["ffe"]), key=f"u_ffe_{curr_type_key}")
            c1.caption(f"""Hitungan: {rooms:,.0f} Rooms x Rp {ffe_rate:,.0f}  \n  Total FF&E: Rp {rooms * ffe_rate:,.0f}  \n  Terbilang: {n2w(rooms * ffe_rate)}""")

            kitchen_rate = c2.number_input("Kitchen Equipment (Rp)", value=get_val("u_kit", pt_data["kitchen"]), key=f"u_kit_{curr_type_key}")
            c2.caption(f"""Hitungan: {rooms:,.0f} Rooms x Rp {kitchen_rate:,.0f}  \n  Total Kitchen Equipment: Rp {rooms * kitchen_rate:,.0f}  \n  Terbilang: {n2w(rooms * kitchen_rate)}""")

            misc_rate = c1.number_input("Misc (Linen/Gym Equipment) (Rp)", value=get_val("u_misc", pt_data["misc"]), key=f"u_misc_{curr_type_key}")
            c1.caption(f"""Hitungan: Rp {misc_rate if misc_switch else 0:,.0f}  \n  Total Misc. Costs: Rp {misc_rate * misc_switch:,.0f}  \n  Terbilang: {n2w(misc_rate * misc_switch)}""")
        
        with st.expander("External & Facility Rates"):
            c1, c2 = st.columns(2)
            ext_land_rate = c1.number_input("External Works (Landscape) (Rp)", value=get_val("u_ext", pt_data["ext_land"]), key=f"u_ext_{curr_type_key}")
            fac_pub_rate = c2.number_input("Public Facilities (Rp)", value=get_val("u_fac_p", pt_data["fac_pub"]), key=f"u_fac_p_{curr_type_key}")
            c1.caption(f"""Hitungan: {land_m2:,.0f} m2 x Rp {ext_land_rate:,.0f}  \n  Total External Works: Rp {land_m2 * ext_land_rate:,.0f}  \n  Terbilang: {n2w(land_m2 * ext_land_rate)}""")
            c2.caption(f"""Hitungan: {pub_fac_m2:,.0f} m2 x Rp {fac_pub_rate:,.0f}  \n  Total Public Facilities: Rp {pub_fac_m2 * fac_pub_rate:,.0f}  \n  Terbilang: {n2w(pub_fac_m2 * fac_pub_rate)}""")
            fac_res_rate = c1.number_input("Resident Facilities (Rp)", value=get_val("u_fac_r", pt_data["fac_res"]), key=f"u_fac_r_{curr_type_key}")
            fac_proj_rate = c2.number_input("Project Facilities (Rp)", value=get_val("u_fac_pr", pt_data["fac_proj"]), key=f"u_fac_pr_{curr_type_key}")
            c1.caption(f"""Hitungan: {res_fac_m2:,.0f} m2 x Rp {fac_res_rate:,.0f}  \n  Total Resident Facilities: Rp {res_fac_m2 * fac_res_rate:,.0f}  \n  Terbilang: {n2w(res_fac_m2 * fac_res_rate)}""")
            c2.caption(f"""Hitungan: {proj_fac_u:,.0f} Units x Rp {fac_proj_rate:,.0f}  \n  Total Project Facilities: Rp {proj_fac_u * fac_proj_rate:,.0f}  \n  Terbilang: {n2w(proj_fac_u * fac_proj_rate)}""")

    # --- TAB 4: SOFT COSTS ---
    with tab5:
        sc_col1, sc_col2, sc_col3 = st.columns(3)
        with sc_col1:
            with st.expander("QS", expanded=True):
                qs_months = st.number_input("Durasi QS (Bulan)", value=get_val("sc_qs_m", 0.0), step=1.0, key=f"sc_qs_m_{curr_id}")
                qs_rate = st.number_input("Harga QS (per Bulan) (Rp)", value=get_val("sc_qs_r", 0.0), step=1000000.0, key=f"sc_qs_r_{curr_id}")
                st.caption(f"""Hitungan: {qs_months} Months x Rp {qs_rate:,.0f}/Mo  \n  Total QS Services: Rp {qs_months * qs_rate:,.0f}  \n  Terbilang: {n2w(qs_months * qs_rate)}""")
        with sc_col2:
            with st.expander("PM", expanded=True):
                pm_months = st.number_input("Durasi PM (Bulan)", value=get_val("sc_pm_m", 0.0), step=1.0, key=f"sc_pm_m_{curr_id}")
                pm_rate = st.number_input("Harga PM (per Bulan) (Rp)", value=get_val("sc_pm_r", 0.0), step=1000000.0, key=f"sc_pm_r_{curr_id}")
                st.caption(f"""Hitungan: {pm_months} Months x Rp {pm_rate:,.0f}/Mo  \n  Total PM Services: Rp {pm_months * pm_rate:,.0f}  \n  Terbilang: {n2w(pm_months * pm_rate)}""")                
        with sc_col3:
            with st.expander("Lainnya", expanded=True):
                consultancy_rate = st.number_input("Biaya Konsultan (Rp) per m2 GFA", help="Biaya konsultan per m2 GFA", value=get_val("sc_cons", pt_data["cons"]), key=f"sc_cons_{curr_type_key}")
                st.caption(f"""Hitungan: {gfa:,.0f} m2 x Rp {consultancy_rate:,.0f}  \n  Total Consultancy Fee: Rp {gfa * consultancy_rate:,.0f}  \n  Terbilang: {n2w(gfa * consultancy_rate)}""")
                insurance_pct = st.number_input("Insurance (%)", help="Persentase premi asuransi", value=get_val("sc_ins", 0.12), step=0.01, key=f"sc_ins_{curr_id}")
    
    # --- TAB 5: CUSTOM ITEMS ---
    with tab6:
        st.subheader("Item Tambahan")

        default_smart_cc = [{"Item Description": "", "Rate (Rp)": 0.0, "Quantity": 1.0}]
        current_smart_cc = get_val("smart_custom_costs", default_smart_cc)

        edited_smart_cc = st.data_editor(
            pd.DataFrame(current_smart_cc),
            num_rows="dynamic",
            key=f"edit_smart_cc_{curr_id}",
            # ADD THIS LINE BELOW
            column_order=["Item Description", "Quantity", "Rate (Rp)"], 
            column_config={
                "Item Description": st.column_config.TextColumn("Item Description", width="large"),
                "Quantity": st.column_config.NumberColumn(
                    "Quantity",
                    min_value=0,
                    default=1.0,
                    format="%.2f"
                )
            },
            use_container_width=True
        )

        smart_custom_costs = 0.0
        
        # New: List to store breakdown strings
        smart_custom_inputs = {}
        smart_custom_costs = 0.0
        breakdown_details = []

        for index, row in edited_smart_cc.iterrows():
            # Create a 1-based suffix for your keys (1, 2, 3...)
            suffix = index + 1
            
            desc = row.get("Item Description", "")
            rate = _safe_float(row.get("Rate (Rp)", 0.0))
            qty = _safe_float(row.get("Quantity", 1.0))
            
            # Store individual keys as requested
            smart_custom_inputs[f"input_name{suffix}"] = desc
            smart_custom_inputs[f"input_rate{suffix}"] = rate
            smart_custom_inputs[f"input_qty{suffix}"] = qty
            
            # Keep the total for the UI/Summary
            item_total = rate * qty
            smart_custom_costs += item_total

            if rate > 0:
                calc_str = f"**{desc}**: Rp {rate:,.2f} × {qty} qty = **Rp {item_total:,.2f}**"
                breakdown_details.append(calc_str)

# ... (keep your loop the same) ...
        
        # 1. Update the 'smart_custom_data' for export logic (what you already have)
        st.session_state[f"smart_custom_data_{curr_id}"] = smart_custom_inputs
        
        # 2. ADD THIS: Update the raw list so the table editor actually remembers the rows!
        st.session_state.projects[curr_id]["data"]["smart_custom_costs"] = edited_smart_cc.to_dict('records')
        
        # 3. Now save to JSON
        save_data()
        
        st.markdown("---")
        st.markdown(f"### Total Harga Item Custom: Rp {smart_custom_costs:,.2f}")

    # --- LIVE AUTO-CALCULATIONS ---
    t_earth = gba * struc_earth
    t_found = gba * struc_found
    t_struc = gba * struc_work
    t_arch_base = gfa * arch_base
    t_precast = facade * (facade_precast_pct / 100) * fac_precast_rate
    t_window  = facade * (facade_window_pct / 100) * fac_window_rate
    t_double  = facade * (facade_double_pct / 100) * fac_double_rate
    t_w_door = wooden_door * door_wood
    t_g_door = glass_door * door_glass
    t_s_door = steel_door * door_steel
    t_lobby  = lobby_interior * lobby_rate
    t_gondola = gondola_unit * gondola_rate
    t_unit_san = rooms * san_qty_room * san_room_rate
    t_t_male   = toilet_male * san_pub_m
    t_t_female = toilet_female * san_pub_f
    t_t_dis    = disabled_toil * san_dis
    t_mushola  = mushola_unit * san_mushola
    t_kitchen = rooms * kitchen_rate
    t_hw_w    = wooden_door * hw_wood
    t_hw_s    = steel_door * hw_steel
    f_mult = (1 + (fl_waste/100)) * (1 + (fl_skirt/100))
    t_ht      = gfa * (fl_ht_pct / 100) * fl_ht_rate * f_mult
    t_vinyl   = gfa * (fl_vinyl_pct / 100) * fl_vinyl_rate * f_mult
    t_marmer  = gfa * (fl_marmer_pct / 100) * fl_marmer_rate * f_mult
    t_carpet     = carpet_m2 * carpet_rate
    t_glass_work = glass_m2 * glass_rate
    t_ffe        = rooms * ffe_rate
    t_misc       = misc_rate * misc_switch
    t_mep        = gba * mep_rate
    t_utility    = gba * utility_rate
    t_railing    = (rooms * railing_qty) * railing_rate
    t_skylight   = skylight_area * skylight_rate
    t_external = land_m2 * ext_land_rate
    t_pub_fac  = pub_fac_m2 * fac_pub_rate
    t_res_fac  = res_fac_m2 * fac_res_rate
    t_proj_fac = proj_fac_u * fac_proj_rate

    construction_subtotal = sum([
        t_earth, t_found, t_struc, t_arch_base, t_precast, t_window, t_double,
        t_w_door, t_g_door, t_s_door, t_lobby, t_gondola, t_unit_san, t_t_male,
        t_t_female, t_t_dis, t_mushola, t_kitchen, t_hw_w, t_hw_s, t_ht, t_vinyl,
        t_marmer, t_carpet, t_glass_work, t_ffe, t_misc, t_mep, t_utility,
        t_railing, t_skylight, t_external, t_pub_fac, t_res_fac, t_proj_fac,
        smart_custom_costs
    ])

    t_preliminary = (construction_subtotal) * 0.05
    t_contingency = (construction_subtotal + t_preliminary) * 0.03
    grand_total_hc = construction_subtotal + t_preliminary + t_contingency

    t_consultancy = gfa * consultancy_rate
    t_qs = qs_months * qs_rate
    t_pm = pm_months * pm_rate
    t_insurance = (grand_total_hc) * (insurance_pct / 100.0)

    total_soft_cost = t_consultancy + t_qs + t_pm + t_insurance
    grand_total_project = grand_total_hc + total_soft_cost

    group_earth = t_earth 
    group_found = t_found 
    group_struc = t_struc
    
    group_facade = t_precast + t_window + t_double
    group_sanitary = t_unit_san + t_t_male + t_t_female + t_t_dis + t_mushola
    group_floor =  t_ht + t_vinyl + t_marmer 
    group_door = t_w_door + t_g_door + t_s_door + t_hw_w + t_hw_s
    group_arch = (t_arch_base + + t_lobby + t_carpet + t_gondola 
                  + t_glass_work + t_kitchen  + t_railing + t_skylight 
                  + group_facade + group_sanitary + group_floor + group_door
                  + smart_custom_costs)
    
    group_ffe = t_ffe + t_misc 
    group_mep = t_mep 
    group_utility = t_utility
    group_ext = t_external
    group_misc = t_pub_fac + t_res_fac + t_proj_fac
    group_prelim = t_preliminary
    group_conting = t_contingency
    group_soft_cost = total_soft_cost
    group_hard_cost = grand_total_hc
    group_total = total_soft_cost + grand_total_hc

    with tab7:
        tab1, tab2, tab3 = st.tabs([
        "Hasil",
        "Tabel", "Chart"
        ])
        
        with tab1:
            box_base = "margin-bottom: 12px; padding: 8px; border-radius: 5px; background-color: #FFFFFF; border: 1px solid #E0E0E0;"
            label_style = "font-size: 12px; color: #666666; font-weight: bold;"
            val_style = "font-size: 14px; font-weight: bold; color: #000000; margin-top: 4px;"
            with st.expander("Detail Hard Cost", expanded=False):
                cs1, cs2, cs3, cs4, cs5, cs6 = st.columns(6)

                # 1. Preliminary (Army Green - Start)
                cs1.markdown(f"""<div style="{box_base} border-left: 5px solid #CDDC39;">
                    <div style="{label_style}">Preliminary</div>
                    <div style="{val_style}">{n2w(group_prelim)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_prelim:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 2. Earthwork (Lime)
                cs2.markdown(f"""<div style="{box_base} border-left: 5px solid #8BC34A;">
                    <div style="{label_style}">Earthwork</div>
                    <div style="{val_style}">{n2w(group_earth)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_earth:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 3. Utility (Sage Green)
                cs3.markdown(f"""<div style="{box_base} border-left: 5px solid #4CAF50;">
                    <div style="{label_style}">Utility</div>
                    <div style="{val_style}">{n2w(group_utility)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_utility:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 4. Foundation (Soft Green)
                cs4.markdown(f"""<div style="{box_base} border-left: 5px solid #689F38;">
                    <div style="{label_style}">Foundation</div>
                    <div style="{val_style}">{n2w(group_found)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_found:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 5. Structural (Success Green)
                cs5.markdown(f"""<div style="{box_base} border-left: 5px solid #388E3C;">
                    <div style="{label_style}">Structural</div>
                    <div style="{val_style}">{n2w(group_struc)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_struc:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 6. External Works (Olive)
                cs6.markdown(f"""<div style="{box_base} border-left: 5px solid #254E18;">
                    <div style="{label_style}">External Works</div>
                    <div style="{val_style}">{n2w(group_ext)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_ext:,.0f}</div>
                </div>""", unsafe_allow_html=True)


                # --- ROW 2: ct1 to ct6 ---
                ct1, ct2, ct3, ct4, ct5, ct6 = st.columns(6)

                # 7. Architecture (Medium Green)
                ct2.markdown(f"""<div style="{box_base} border-left: 5px solid #9CCC65;">
                    <div style="{label_style}">Architecture</div>
                    <div style="{val_style}">{n2w(group_arch)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_arch:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 8. Miscellaneous (Dark Olive)
                ct3.markdown(f"""<div style="{box_base} border-left: 5px solid #558B2F;">
                    <div style="{label_style}">Miscellaneous</div>
                    <div style="{val_style}">{n2w(group_misc)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_misc:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 9. FF & E (Dark Green)
                ct4.markdown(f"""<div style="{box_base} border-left: 5px solid #2E7D32;">
                    <div style="{label_style}">FF & E</div>
                    <div style="{val_style}">{n2w(group_ffe)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_ffe:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 10. Contingency (Deep Army)
                ct6.markdown(f"""<div style="{box_base} border-left: 5px solid #1B5E20;">
                    <div style="{label_style}">Contingency</div>
                    <div style="{val_style}">{n2w(group_conting)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_conting:,.0f}</div>
                </div>""", unsafe_allow_html=True)

                # 11. MEP Works (Forest Green)
                ct5.markdown(f"""<div style="{box_base} border-left: 5px solid #33691E;">
                    <div style="{label_style}">MEP Works</div>
                    <div style="{val_style}">{n2w(group_mep)}</div>
                    <div style="font-size: 10px; color: #888888; margin-top: 2px;">Rp {group_mep:,.0f}</div>
                </div>""", unsafe_allow_html=True)
            
    # --- TOTALS SUMMARY ---
            c1, c2 = st.columns(2)

            # Common styles
            summary_base = "margin-bottom: 20px; padding: 15px; border-radius: 8px; background-color: #FFFFFF; border: 1px solid #E0E0E0;"
            summary_label = "font-size: 16px; color: #666666; font-weight: bold;"
            summary_val = "font-size: 24px; font-weight: bold; color: #000000; line-height: 1.2; margin-top: 5px;"
            summary_n2w = "font-size: 14px; color: #888888; font-weight: normal; margin-top: 5px;"

            # Hard Cost - Starting with Lime Accent
            c1.markdown(f"""
                <div style="{summary_base} border-left: 8px solid #CDDC39; background-color: #F9F9F0">
                    <div style="{summary_label}">Total Hard Cost</div>
                    <div style="{summary_val}">Rp {grand_total_hc:,.2f}</div>
                    <div style="{summary_n2w}">Terbilang: {n2w(grand_total_hc)} Rupiah</div>
                </div>
            """, unsafe_allow_html=True)


            # Soft Cost - Moving to Success Green Accent
            c2.markdown(f"""
                <div style="{summary_base} border-left: 8px solid #4CAF50; background-color: #F1F8F1">
                    <div style="{summary_label}">Total Soft Cost</div>
                    <div style="{summary_val}">Rp {total_soft_cost:,.2f}</div>
                    <div style="{summary_n2w}">Terbilang: {n2w(total_soft_cost)} Rupiah</div>
                </div>
            """, unsafe_allow_html=True)

            # Grand Total - Final Army Green Accent (Thicker border and larger font)
            st.markdown(f"""
                <div style="{summary_base} border-left: 12px solid #254E18; padding: 20px; background-color: #F5F7F5">
                    <div style="font-size: 18px; color: #666666; font-weight: bold;">Project Cost</div>
                    <div style="font-size: 12px; color: #888888; margin-top: 8px;">
                        Rp {grand_total_hc:,.2f} + Rp {total_soft_cost:,.2f} =
                    </div>
                    <div style="font-size: clamp(24px, 4vw, 30px); font-weight: bold; color: #000000; line-height: 1.2; margin-top: 10px;">
                        Rp {grand_total_project:,.2f}
                    </div>
                    <div style="font-size: 16px; color: #888888; margin-top: 8px;">
                        Terbilang: {n2w(grand_total_project)} Rupiah
                    </div>
                </div>
            """, unsafe_allow_html=True)

        with tab2:
            if 'show_details' not in st.session_state:
                st.session_state.show_details = True

            # 2. Add the Toggle Button
            label = "Hide Details" if st.session_state.show_details else "Show Details"
            if st.button(label):
                st.session_state.show_details = not st.session_state.show_details
                st.rerun()

            # 3. Define the full dataset as a Dictionary
            data_dict = {
                # --- MAIN CONSTRUCTION ---
                "1. Preliminary Works": t_preliminary,
                "2. Earthwork": t_earth,
                "3. Foundation": t_found,
                "4. Structural Work": t_struc,
                
                # --- ARCHITECTURE ---
                "5. Total Architecture": group_arch,
                "5.1 Basic Architecture": t_arch_base,
                "5.2 Facade - Precast": t_precast,
                "5.3 Facade - Window Wall": t_window,
                "5.4 Facade - Double Skin": t_double,
                "5.5 Wooden Doors": t_w_door,
                "5.6 Glass Doors": t_g_door,
                "5.7 Steel Doors": t_s_door,
                "5.8 Lobby Interior": t_lobby,
                "5.9 Gondola": t_gondola,
                "5.10 Typical Unit Sanitary": t_unit_san,
                "5.11 Public Toilet Male": t_t_male,
                "5.12 Public Toilet Female": t_t_female,
                "5.13 Disabled Toilet": t_t_dis,
                "5.14 Mushola": t_mushola,
                "5.15 Kitchen Equipment": t_kitchen,
                "5.16 Hardware Pintu Kayu": t_hw_w,
                "5.17 Hardware Pintu Besi": t_hw_s,
                "5.18 HT/Ceramic Tile": t_ht,
                "5.19 Vinyl Flooring": t_vinyl,
                "5.20 Marmer Flooring": t_marmer,
                "5.21 Carpet Work": t_carpet,
                "5.22 Railing Work": t_railing,
                "5.23 Skylight Work": t_skylight,
                "5.24 Glass Work": t_glass_work,
                "5.25 Custom Item (Architecture)": smart_custom_costs,

                # --- FF&E & SERVICES ---
                "6. Total FF&E": group_ffe,
                "6.1 FF&E": t_ffe,
                "6.2 Misc. (Linen/Gym)": t_misc,
                "7. MEP Works": t_mep,
                "8. Utility Connection": t_utility,

                # --- EXTERNAL & FACILITIES ---
                "9. External Works": t_external,
                "10. Miscellanous/Facility": group_misc,
                "10.1 Public Facilities": t_pub_fac,
                "10.2 Resident Facilities": t_res_fac,
                "10.3 Project Facilities": t_proj_fac,

                # --- CONTINGENCY ---
                "11. Contingencies": t_contingency,

                # --- SOFT COSTS / CONSULTANTS ---
                "12. Consultancy Fee": t_consultancy,
                "13. Quantity Surveyor": t_qs,
                "14. Project Management": t_pm,
                "15. Insurance Coverage": t_insurance
            }

            # 4. Extract into lists for your DataFrame automatically
            original_descriptions = list(data_dict.keys())
            raw_amounts = list(data_dict.values())

            # 4. Filter and Indent Logic
            filtered_data = []
            major_numbers = [f"{i}. " for i in range(1, 16)]

            for desc, amt in zip(original_descriptions, raw_amounts):
                is_major = any(desc.startswith(num) for num in major_numbers)
                
                if st.session_state.show_details:
                    # If showing details, indent sub-items
                    display_desc = desc if is_major else f"    {desc}"
                    filtered_data.append({"Description": display_desc, "Amount": amt})
                else:
                    # If hiding details, only append major items
                    if is_major:
                        filtered_data.append({"Description": desc, "Amount": amt})

            # 5. Create DataFrame
            df = pd.DataFrame(filtered_data)
            df["Amount"] = df["Amount"].apply(lambda x: f"Rp {x:,.2f}")

            # 6. Styling Logic (Bold Major Items)
            def style_major_rows(row):
                clean_desc = row["Description"].strip()
                is_major = any(clean_desc.startswith(num) for num in major_numbers)
                return ['font-weight: bold' if is_major else '' for _ in row]

            # 7. Display
            styled_df = df.style.apply(style_major_rows, axis=1)
            st.dataframe(styled_df, use_container_width=True, hide_index=True)
        
        with tab3:
                st.subheader("Total Project Cost Breakdown")

                # 1. Define the dictionary first
                detailed_items = {
                    "Item": [
                        "Preliminary", "Earthwork", "Foundation", "Structural", 
                        "Architecture Work", "FF&E", "MEP Works", "Utility",
                        "External/Landscape", "Misc/Facility", "Contingency",
                        "Soft Cost/Consultancy & Insurance"
                    ],
                    "Amount": [
                        group_prelim, group_earth, group_found, group_struc, 
                        group_arch, group_ffe, group_mep, group_utility, 
                        group_ext, group_misc,
                        group_conting, group_soft_cost
                    ],
                    # Adding 'Type' helps with color coding the chart
                    "Type": ["Hard Cost"]*11 + ["Soft Cost"]*1 
                }

                # 2. Convert to DataFrame and FILTER OUT zeros
                df_detailed = pd.DataFrame(detailed_items)
                df_detailed = df_detailed[df_detailed["Amount"] > 0] 

                # 3. Dynamic height calculation
                chart_height = max(400, len(df_detailed) * 25)

                # 4. Create the Chart
                hover = alt.selection_point(on='mouseover', nearest=True, fields=['Item'], empty=False)

                detailed_chart = alt.Chart(df_detailed).mark_bar().encode(
                    x=alt.X("Amount:Q", title="Cost (Rp)"),
                    y=alt.Y("Item:N", sort="-x", title=""),
                    opacity=alt.condition(hover, alt.value(1), alt.value(0.7)),
                    color=alt.Color("Type:N", scale=alt.Scale(domain=['Hard Cost', 'Soft Cost'], range=["#1f77b4", "#c2a136"])),
                    tooltip=["Item", "Type", alt.Tooltip("Amount:Q", format=",.2f")]
                ).properties(height=chart_height).add_params(hover)

                st.altair_chart(detailed_chart, use_container_width=True)

# --- SAVE ALL METRICS TO SESSION STATE ---
    # We use .update() so we NEVER delete the Area Analysis's data!
    st.session_state.projects[curr_id]["data"].update({
        "ht_spec_type": get_val("ht_spec_type", "Type1"),
        "vin_spec_type": get_val("vin_spec_type", "Type1"),
        "mar_spec_type": get_val("mar_spec_type", "Type1"),
        "m_land": land_area, "m_gba": gba, "m_gfa": gfa, "m_sgfa": sgfa,
        "m_facade": facade, "m_rooms": rooms, "m_lobby": lobby_interior,
        "m_gondola": gondola_unit, "m_carpet": carpet_m2, "m_glass": glass_m2,
        "m_skylight": skylight_area, "m_door_g": glass_door, "m_door_w": wooden_door,
        "m_door_s": steel_door, "m_toil_m": toilet_male, "m_toil_f": toilet_female,
        "m_toil_d": disabled_toil, "m_mushola": mushola_unit, "m_fac_res": res_fac_m2,
        "m_fac_pub": pub_fac_m2, "m_fac_proj": proj_fac_u, "m_land_m2": land_m2,
        "r_fac_pre": facade_precast_pct, "r_fac_win": facade_window_pct, "r_fac_doub": facade_double_pct,
        "r_fl_ht": fl_ht_pct, "r_fl_vin": fl_vinyl_pct, "r_fl_mar": fl_marmer_pct,
        "r_san_qty": san_qty_room, "r_rail_qty": railing_qty,
        "u_earth": struc_earth, "u_found": struc_found, "u_struc": struc_work, "u_arch": arch_base,
        "u_lobby": lobby_rate, "u_f_pre": fac_precast_rate, "u_f_win": fac_window_rate, "u_f_doub": fac_double_rate,
        "u_d_wood": door_wood, "u_d_glass": door_glass, "u_d_steel": door_steel,
        "u_hw_wood": hw_wood, "u_hw_steel": hw_steel,
        "u_s_room": san_room_rate, "u_s_pub_m": san_pub_m, "u_s_pub_f": san_pub_f,
        "u_s_dis": san_dis, "u_s_mushola": san_mushola,
        "w_floor": fl_waste, "s_floor" :fl_skirt,
        "u_fl_ht": fl_ht_rate, "u_fl_vin": fl_vinyl_rate, "u_fl_mar": fl_marmer_rate,
        "u_carpet": carpet_rate, "u_glass": glass_rate, "u_sky": skylight_rate,
        "u_gondola": gondola_rate, "u_rail": railing_rate,
        "u_mep": mep_rate, "u_util": utility_rate,
        "u_ffe": ffe_rate, "u_kit": kitchen_rate, "u_misc": misc_rate,
        "u_ext": ext_land_rate, "u_fac_p": fac_pub_rate,
        "u_fac_r": fac_res_rate, "u_fac_pr": fac_proj_rate,
        "sc_cons": consultancy_rate, "sc_qs_m": qs_months, "sc_qs_r": qs_rate,
        "sc_pm_m": pm_months, "sc_pm_r": pm_rate, "sc_ins": insurance_pct
    })
    st.session_state.projects[curr_id]["data"]["grand_total_project"] = grand_total_project
    save_data()

    with tab8:
            st.header("Detail Pembuktian & Logika Perhitungan")

            # Group 1: Structural Audit
            with st.expander("1. Struktur & Pondasi", expanded=True):
                audit_struc = pd.DataFrame([
                    {"Item": "Earthwork", "Formula": f"GBA ({gba:,.0f} m2) × Rate (Rp {struc_earth:,.0f})", "Total": t_earth},
                    {"Item": "Foundation", "Formula": f"GBA ({gba:,.0f} m2) × Rate (Rp {struc_found:,.0f})", "Total": t_found},
                    {"Item": "Structural Work", "Formula": f"GBA ({gba:,.0f} m2) × Rate (Rp {struc_work:,.0f})", "Total": t_struc},
                ])
                st.table(audit_struc.style.format({"Total": "Rp {:,.2f}"}))

            # Group 2: Architectural Base & Facade Audit
            with st.expander("2. Arsitektur Dasar & Fasad"):
                audit_arch = pd.DataFrame([
                    {"Item": "Architecture Base", "Formula": f"GFA ({gfa:,.0f} m2) × Rate (Rp {arch_base:,.0f})", "Total": t_arch_base},
                    {"Item": "Facade Precast", "Formula": f"Facade ({facade:,.0f} m2) × {facade_precast_pct}% × Rate (Rp {fac_precast_rate:,.0f})", "Total": t_precast},
                    {"Item": "Facade Window Wall", "Formula": f"Facade ({facade:,.0f} m2) × {facade_window_pct}% × Rate (Rp {fac_window_rate:,.0f})", "Total": t_window},
                    {"Item": "Facade Double Skin", "Formula": f"Facade ({facade:,.0f} m2) × {facade_double_pct}% × Rate (Rp {fac_double_rate:,.0f})", "Total": t_double},
                ])
                st.table(audit_arch.style.format({"Total": "Rp {:,.2f}"}))

            # Group 3: Doors & Hardware
            with st.expander("3. Pintu & Hardware"):
                audit_door = pd.DataFrame([
                    {"Item": "Wooden Door", "Formula": f"{wooden_door:,.0f} Unit × Rate (Rp {door_wood:,.0f})", "Total": t_w_door},
                    {"Item": "Glass Door", "Formula": f"{glass_door:,.0f} Unit × Rate (Rp {door_glass:,.0f})", "Total": t_g_door},
                    {"Item": "Steel Door", "Formula": f"{steel_door:,.0f} Unit × Rate (Rp {door_steel:,.0f})", "Total": t_s_door},
                    {"Item": "Hardware Wooden Door", "Formula": f"{wooden_door:,.0f} Unit × Rate (Rp {hw_wood:,.0f})", "Total": t_hw_w},
                    {"Item": "Hardware Steel Door", "Formula": f"{steel_door:,.0f} Unit × Rate (Rp {hw_steel:,.0f})", "Total": t_hw_s},
                ])
                st.table(audit_door.style.format({"Total": "Rp {:,.2f}"}))

            # Group 4: Sanitary Audit
            with st.expander("4. Sanitari"):
                audit_san = pd.DataFrame([
                    {"Item": "Typical Unit Sanitary", "Formula": f"{rooms:,.0f} Rooms × {san_qty_room} Unit/Room × Rate (Rp {san_room_rate:,.0f})", "Total": t_unit_san},
                    {"Item": "Public Toilet Male", "Formula": f"{toilet_male:,.0f} Unit × Rate (Rp {san_pub_m:,.0f})", "Total": t_t_male},
                    {"Item": "Public Toilet Female", "Formula": f"{toilet_female:,.0f} Unit × Rate (Rp {san_pub_f:,.0f})", "Total": t_t_female},
                    {"Item": "Disabled Toilet", "Formula": f"{disabled_toil:,.0f} Unit × Rate (Rp {san_dis:,.0f})", "Total": t_t_dis},
                    {"Item": "Mushola", "Formula": f"{mushola_unit:,.0f} Unit × Rate (Rp {san_mushola:,.0f})", "Total": t_mushola},
                ])
                st.table(audit_san.style.format({"Total": "Rp {:,.2f}"}))

            # Group 5: Flooring Logic (Waste & Skirting Proof)
            with st.expander("5. Lantai (Termasuk Waste & Skirting)"):
                st.markdown(fr"""
                **Rumus Pengali Lantai (f_mult):**  
                $(1 + \text{{Waste}} \%) \times (1 + \text{{Skirting}} \%) = (1 + {fl_waste/100}) \times (1 + {fl_skirt/100}) = **{f_mult:.4f}**$
                """)
                audit_floor = pd.DataFrame([
                    {"Item": "HT / Tile", "Formula": f"GFA ({gfa:,.0f} m2) × {fl_ht_pct}% × {f_mult:.4f} × Rate (Rp {fl_ht_rate:,.0f})", "Total": t_ht},
                    {"Item": "Vinyl", "Formula": f"GFA ({gfa:,.0f} m2) × {fl_vinyl_pct}% × {f_mult:.4f} × Rate (Rp {fl_vinyl_rate:,.0f})", "Total": t_vinyl},
                    {"Item": "Marmer", "Formula": f"GFA ({gfa:,.0f} m2) × {fl_marmer_pct}% × {f_mult:.4f} × Rate (Rp {fl_marmer_rate:,.0f})", "Total": t_marmer},
                ])
                st.table(audit_floor.style.format({"Total": "Rp {:,.2f}"}))

            # Group 6: Finishing & Interior Misc
            with st.expander("6. Finishing, Interior & Lainnya"):
                audit_fin = pd.DataFrame([
                    {"Item": "Lobby Interior", "Formula": f"{lobby_interior:,.0f} m2 × Rate (Rp {lobby_rate:,.0f})", "Total": t_lobby},
                    {"Item": "Carpet Work", "Formula": f"{carpet_m2:,.0f} m2 × Rate (Rp {carpet_rate:,.0f})", "Total": t_carpet},
                    {"Item": "Glass Work", "Formula": f"{glass_m2:,.0f} m2 × Rate (Rp {glass_rate:,.0f})", "Total": t_glass_work},
                    {"Item": "Skylight", "Formula": f"{skylight_area:,.0f} m2 × Rate (Rp {skylight_rate:,.0f})", "Total": t_skylight},
                    {"Item": "Railing", "Formula": f"({rooms:,.0f} Rooms × {railing_qty} m'/room) × Rate (Rp {railing_rate:,.0f})", "Total": t_railing},
                    {"Item": "Gondola", "Formula": f"{gondola_unit:,.0f} Unit × Rate (Rp {gondola_rate:,.0f})", "Total": t_gondola},
                ])
                st.table(audit_fin.style.format({"Total": "Rp {:,.2f}"}))

            # Group 7: MEP, Dapur & FF&E
            with st.expander("7. MEP, Dapur & FF&E"):
                audit_mep = pd.DataFrame([
                    {"Item": "MEP Works", "Formula": f"GBA ({gba:,.0f} m2) × Rate (Rp {mep_rate:,.0f})", "Total": t_mep},
                    {"Item": "Utility Connection", "Formula": f"GBA ({gba:,.0f} m2) × Rate (Rp {utility_rate:,.0f})", "Total": t_utility},
                    {"Item": "FF&E", "Formula": f"{rooms:,.0f} Rooms × Rate (Rp {ffe_rate:,.0f})", "Total": t_ffe},
                    {"Item": "Kitchen Equipment", "Formula": f"{rooms:,.0f} Rooms × Rate (Rp {kitchen_rate:,.0f})", "Total": t_kitchen},
                    {"Item": "Misc (Linen/Gym)", "Formula": f"Switch ({misc_switch}) × Rate (Rp {misc_rate:,.0f})", "Total": t_misc},
                ])
                st.table(audit_mep.style.format({"Total": "Rp {:,.2f}"}))

            # Group 8: External & Facilities
            with st.expander("8. Fasilitas & Area Eksternal"):
                audit_ext = pd.DataFrame([
                    {"Item": "External Works (Landscape)", "Formula": f"{land_m2:,.0f} m2 × Rate (Rp {ext_land_rate:,.0f})", "Total": t_external},
                    {"Item": "Public Facilities", "Formula": f"{pub_fac_m2:,.0f} m2 × Rate (Rp {fac_pub_rate:,.0f})", "Total": t_pub_fac},
                    {"Item": "Resident Facilities", "Formula": f"GFA ({gfa:,.0f} m2) × Rate (Rp {fac_res_rate:,.0f})", "Total": t_res_fac},
                    {"Item": "Project Facilities", "Formula": f"{proj_fac_u:,.0f} Unit × Rate (Rp {fac_proj_rate:,.0f})", "Total": t_proj_fac},
                ])
                st.table(audit_ext.style.format({"Total": "Rp {:,.2f}"}))

            # Group 9: Markups (Prelim & Contingency)
            with st.expander("9. Preliminaries & Contingency"):
                audit_markup = pd.DataFrame([
                    {"Item": "Preliminary Works", "Formula": f"Construction Subtotal (Rp {construction_subtotal:,.0f}) × 5%", "Total": t_preliminary},
                    {"Item": "Contingencies", "Formula": f"(Subtotal + Prelim) (Rp {construction_subtotal + t_preliminary:,.0f}) × 3%", "Total": t_contingency},
                ])
                st.table(audit_markup.style.format({"Total": "Rp {:,.2f}"}))

            # Group 10: Soft Costs (Consultants & Insurance)
            with st.expander("10. Soft Costs (Consultants, QS, PM, Insurance)"):
                audit_soft = pd.DataFrame([
                    {"Item": "Consultancy Fee", "Formula": f"GFA ({gfa:,.0f} m2) × Rate (Rp {consultancy_rate:,.0f})", "Total": t_consultancy},
                    {"Item": "QS Services", "Formula": f"{qs_months} Months × Rate (Rp {qs_rate:,.0f})", "Total": t_qs},
                    {"Item": "PM Services", "Formula": f"{pm_months} Months × Rate (Rp {pm_rate:,.0f})", "Total": t_pm},
                    {"Item": "Insurance Coverage", "Formula": f"Total Hard Cost (Rp {grand_total_hc:,.0f}) × {insurance_pct}%", "Total": t_insurance},
                ])
                st.table(audit_soft.style.format({"Total": "Rp {:,.2f}"}))
                
            # Group 11: Custom Items Tracker
            with st.expander("11. Item Tambahan (Custom)"):
                if smart_custom_costs > 0:
                    st.markdown(f"**Total Item Tambahan:** Rp {smart_custom_costs:,.2f}")
                    for detail in breakdown_details:
                        st.markdown(f"- {detail}")
                else:
                    st.info("Tidak ada item tambahan (custom) yang dimasukkan.")

#region --- DO NOT CHANGE#3 (OR GOD HELP ME) ---
def generate_exact_portfolio_excel(port_meta, port_data, port_assumptions):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"

    # --- 1. Styling Definitions ---
    blue_fill = PatternFill(start_color="005A9C", end_color="005A9C", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    white_font = Font(color="FFFFFF", bold=True, name='Calibri', size=11)
    bold_font = Font(bold=True, name='Calibri', size=10)
    reg_font = Font(bold=False, name='Calibri', size=10)
    small_font = Font(name='Calibri', size=9)
    
    black_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # --- 2. Blue Header Section ---
    headers = [
        ["ASG GROUP PROPERTY DEVELOPMENT", f"VERSION : {port_meta.get('version', '')}"],
        ["QS & PROCUREMENT DIVISION", f"UPDATED : {port_meta.get('updated', '')}"],
        [port_meta.get('title', ''), f"CREATED : {port_meta.get('created', '')}"],
        [port_meta.get('ref', ''), ""]
    ]

    for r_idx, (text_left, text_right) in enumerate(headers, 1):
        for c_idx in range(1, 12):
            c = ws.cell(row=r_idx, column=c_idx)
            c.fill = blue_fill
            c.font = white_font
            if c_idx == 1: c.value = text_left
            if c_idx == 11: 
                c.value = text_right
                c.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=10)

    # --- 3. Table Header (Rows 6 & 7) ---
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1) # SN
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=2) # AREA
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=5) # BLDG AREA
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7) # UNIT
    ws.merge_cells(start_row=6, start_column=8, end_row=7, end_column=8) # BUDGET
    ws.merge_cells(start_row=6, start_column=9, end_row=6, end_column=11) # COST RATIO

    header_labels = {
        (6, 1): "SN", (6, 2): "AREA", (6, 3): "BUILDING AREA (M2)", 
        (6, 6): "UNIT", (6, 8): "BUDGET ESTIMATE\nRP", (6, 9): "COST RATIO RP/M2",
        (7, 3): "GBA", (7, 4): "GFA", (7, 5): "SGFA", (7, 9): "GBA", (7, 10): "GFA", (7, 11): "SGFA"
    }

    for (r, c), text in header_labels.items():
        cell = ws.cell(row=r, column=c, value=text)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = gray_fill

    for r in range(6, 8):
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = black_border

    # --- 4. Data Rows ---
    current_row = 8
    for p_row in port_data:
        is_total = p_row.get("AREA") == "TOTAL"
        is_parking = "PARKING" in str(p_row.get("AREA", "")).upper()
        
        # Set SN and Area
        ws.cell(row=current_row, column=1, value=p_row.get("SN", "")).alignment = center_align
        
        area_cell = ws.cell(row=current_row, column=2, value=p_row.get("AREA", ""))
        area_cell.alignment = Alignment(horizontal='left', vertical='top' if is_parking else 'center', wrap_text=True)
        
        # Set Values
        cols = ["GBA", "GFA", "SGFA", "QTY", "UNIT", "BUDGET", "R_GBA", "R_GFA", "R_SGFA"]
        for i, key in enumerate(cols, 3):
            val = p_row.get(key, "")
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(vertical='top' if is_parking else 'center', horizontal='right' if i != 7 else 'center')
            
            if key in ["GBA", "GFA", "SGFA"]: cell.number_format = "#,##0.00"
            if key in ["BUDGET", "R_GBA", "R_GFA", "R_SGFA"]: cell.number_format = "#,##0"

        # Apply Row Styles
        for c in range(1, 12):
            ws.cell(row=current_row, column=c).border = black_border
            ws.cell(row=current_row, column=c).font = bold_font if (p_row.get("SN") or is_total) else reg_font
            if is_total: ws.cell(row=current_row, column=c).fill = gray_fill

        if is_total:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=6, value="TOTAL").alignment = center_align

        # Adjust height for multi-line rows (Parking)
        if is_parking:
            ws.row_dimensions[current_row].height = 100 

        current_row += 1

    # --- 5. Assumptions Section ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    cell_assump = ws.cell(row=current_row, column=1, value="I.  ASSUMPTIONS")
    cell_assump.font = bold_font
    for c in range(1, 12):
        ws.cell(row=current_row, column=c).fill = yellow_fill
        ws.cell(row=current_row, column=c).border = black_border

    current_row += 1
    for _, a_row in port_assumptions.iterrows():
        ws.cell(row=current_row, column=1, value=a_row.get("No.", "")).alignment = center_align
        
        desc_cell = ws.cell(row=current_row, column=2, value=a_row.get("Assumption Description", ""))
        desc_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
        
        for c in range(1, 12):
            ws.cell(row=current_row, column=c).border = black_border
        current_row += 1

    # Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for c in ['C', 'D', 'E', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[c].width = 16

    wb.save(output)
    return output.getvalue()

def generate_recap_excel(port_meta, projects):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Recap Cost"

    # --- EXACT MATH ENGINE ---
    def get_recap_values(pdata):
        d = pdata.get("data", {})
        curr_type = pdata.get("type", "Hotel")
        pt_data = PROJECT_DATABASE.get(curr_type, {})
        
        def get_val(key, default_db_key, default_val=0.0):
            val = d.get(key)
            if val is not None and val != "":
                try: return _safe_float(val)
                except: pass
            if default_db_key and default_db_key in pt_data:
                try: return _safe_float(pt_data[default_db_key])
                except: pass
            return _safe_float(default_val)
            
        gba = get_val("m_gba", None, 0); gfa = get_val("m_gfa", None, 0)
        struc_earth = get_val("u_earth", "struc_earth", 0)
        struc_found = get_val("u_found", "struc_found", 0)
        struc_work = get_val("u_struc", "struc_work", 0)
        arch_base = get_val("u_arch", "arch_base", 0)
        facade = get_val("m_facade", None, 0)
        facade_precast_pct = get_val("r_fac_pre", "facade_precast_pct", 0)
        facade_precast_rate = get_val("u_f_pre", "facade_precast_rate", 0)
        facade_window_pct = get_val("r_fac_win", "facade_window_pct", 0)
        facade_window_rate = get_val("u_f_win", "facade_window_rate", 0)
        facade_double_pct = get_val("r_fac_doub", "facade_double_pct", 0)
        facade_double_rate = get_val("u_f_doub", "facade_double_rate", 0)
        wooden_door = get_val("m_door_w", None, 0); door_wood = get_val("u_d_wood", "door_wood", 0)
        glass_door = get_val("m_door_g", None, 0); door_glass = get_val("u_d_glass", "door_glass", 0)
        steel_door = get_val("m_door_s", None, 0); door_steel = get_val("u_d_steel", "door_steel", 0)
        lobby_interior = get_val("m_lobby", None, 0); lobby_rate = get_val("u_lobby", "lobby", 0)
        gondola_unit = get_val("m_gondola", None, 0); gondola_rate = get_val("u_gondola", "gondola", 0)
        rooms = get_val("m_rooms", None, 0)
        san_qty_room = get_val("r_san_qty", "san_room_qty", 0); san_room_rate = get_val("u_s_room", "san_room_rate", 0)
        toilet_male = get_val("m_toil_m", None, 0); san_pub_m = get_val("u_s_pub_m", "san_pub_m", 0)
        toilet_female = get_val("m_toil_f", None, 0); san_pub_f = get_val("u_s_pub_f", "san_pub_f", 0)
        disabled_toil = get_val("m_toil_d", None, 0); san_dis = get_val("u_s_dis", "san_dis", 0)
        mushola_unit = get_val("m_mushola", None, 0); san_mushola = get_val("u_s_mushola", "san_mushola", 0)
        kitchen_rate = get_val("u_kit", "kitchen", 0)
        hw_wood = get_val("u_hw_wood", "hw_wood", 0); hw_steel = get_val("u_hw_steel", "hw_steel", 0)
        fl_waste = get_val("w_floor", "fl_waste", 10); fl_skirt = get_val("s_floor", "fl_skirt", 20)
        fl_ht_pct = get_val("r_fl_ht", "fl_ht_pct", 0); fl_ht_rate = get_val("u_fl_ht", None, 0) 
        fl_vinyl_pct = get_val("r_fl_vin", "fl_vinyl_pct", 0); fl_vinyl_rate = get_val("u_fl_vin", None, 0)
        fl_marmer_pct = get_val("r_fl_mar", "fl_marmer_pct", 0); fl_marmer_rate = get_val("u_fl_mar", None, 0)
        carpet_m2 = get_val("m_carpet", None, 0); carpet_rate = get_val("u_carpet", "carpet", 0)
        glass_m2 = get_val("m_glass", None, 0); glass_rate = get_val("u_glass", "glass", 0)
        ffe_rate = get_val("u_ffe", "ffe", 0); misc_rate = get_val("u_misc", "misc", 0); misc_switch = get_val("misc_switch", None, 0)
        mep_rate = get_val("u_mep", "mep", 0); utility_rate = get_val("u_util", "utility", 0)
        railing_qty = get_val("r_rail_qty", "railing_qty", 0); railing_rate = get_val("u_rail", "railing_rate", 0)
        skylight_area = get_val("m_skylight", None, 0); skylight_rate = get_val("u_sky", "skylight_rate", 0)
        land_m2 = get_val("m_land_m2", None, 0); ext_land_rate = get_val("u_ext", "ext_land", 0)
        pub_fac_m2 = get_val("m_fac_pub", None, 0); fac_pub_rate = get_val("u_fac_p", "fac_pub", 0)
        res_fac_m2 = get_val("m_fac_res", None, 0); fac_res_rate = get_val("u_fac_r", "fac_res", 0)
        proj_fac_u = get_val("m_fac_proj", None, 0); fac_proj_rate = get_val("u_fac_pr", "fac_proj", 0)
        consultancy_rate = get_val("sc_cons", "cons", 0); qs_months = get_val("sc_qs_m", None, 0); qs_rate = get_val("sc_qs_r", None, 0)
        pm_months = get_val("sc_pm_m", None, 0); pm_rate = get_val("sc_pm_r", None, 0); insurance_pct = get_val("sc_ins", None, 0.12)
        smart_custom_costs = sum(_safe_float(i.get("Rate (Rp)", 0)) * _safe_float(i.get("Quantity", 1)) for i in d.get("smart_custom_costs", []) if isinstance(i, dict))

        t_earth = gba * struc_earth; t_found = gba * struc_found; t_struc = gba * struc_work; t_arch_base = gfa * arch_base
        t_precast = facade * (facade_precast_pct / 100) * facade_precast_rate; t_window = facade * (facade_window_pct / 100) * facade_window_rate
        t_double = facade * (facade_double_pct / 100) * facade_double_rate; t_w_door = wooden_door * door_wood
        t_g_door = glass_door * door_glass; t_s_door = steel_door * door_steel; t_lobby = lobby_interior * lobby_rate
        t_gondola = gondola_unit * gondola_rate; t_unit_san = rooms * san_qty_room * san_room_rate
        t_t_male = toilet_male * san_pub_m; t_t_female = toilet_female * san_pub_f; t_t_dis = disabled_toil * san_dis
        t_mushola = mushola_unit * san_mushola; t_kitchen = rooms * kitchen_rate; t_hw_w = wooden_door * hw_wood
        t_hw_s = steel_door * hw_steel; f_mult = (1 + (fl_waste/100)) * (1 + (fl_skirt/100))
        t_ht = gfa * (fl_ht_pct / 100) * fl_ht_rate * f_mult; t_vinyl = gfa * (fl_vinyl_pct / 100) * fl_vinyl_rate * f_mult
        t_marmer = gfa * (fl_marmer_pct / 100) * fl_marmer_rate * f_mult; t_carpet = carpet_m2 * carpet_rate
        t_glass_work = glass_m2 * glass_rate; t_ffe = rooms * ffe_rate; t_misc = misc_rate * misc_switch
        t_mep = gba * mep_rate; t_utility = gba * utility_rate; t_railing = (rooms * railing_qty) * railing_rate
        t_skylight = skylight_area * skylight_rate; t_external = land_m2 * ext_land_rate
        t_pub_fac = pub_fac_m2 * fac_pub_rate; t_res_fac = res_fac_m2 * fac_res_rate; t_proj_fac = proj_fac_u * fac_proj_rate

        construction_subtotal = sum([
            t_earth, t_found, t_struc, t_arch_base, t_precast, t_window, t_double, t_w_door, t_g_door, t_s_door, 
            t_lobby, t_gondola, t_unit_san, t_t_male, t_t_female, t_t_dis, t_mushola, t_kitchen, t_hw_w, t_hw_s, 
            t_ht, t_vinyl, t_marmer, t_carpet, t_glass_work, t_ffe, t_misc, t_mep, t_utility, t_railing, t_skylight, 
            t_external, t_pub_fac, t_res_fac, t_proj_fac, smart_custom_costs
        ])

        t_preliminary = construction_subtotal * 0.05
        t_contingency = (construction_subtotal + t_preliminary) * 0.03
        grand_total_hc = construction_subtotal + t_preliminary + t_contingency

        t_consultancy = gfa * consultancy_rate
        t_qs = qs_months * qs_rate
        t_pm = pm_months * pm_rate
        t_insurance = grand_total_hc * (insurance_pct / 100.0)

        total_soft_cost = t_consultancy + t_qs + t_pm + t_insurance
        grand_total_project = grand_total_hc + total_soft_cost

        group_arch = (t_arch_base + t_lobby + t_carpet + t_gondola + t_glass_work + t_kitchen + t_railing + t_skylight + 
                      (t_precast + t_window + t_double) + (t_unit_san + t_t_male + t_t_female + t_t_dis + t_mushola) + 
                      (t_ht + t_vinyl + t_marmer) + (t_w_door + t_g_door + t_s_door + t_hw_w + t_hw_s) + smart_custom_costs)
        
        return {
            "EARTHWORKS": t_earth, "FOUNDATIONS": t_found, "STRUCTURAL WORKS": t_struc,
            "ARCHITECTURAL WORKS": group_arch, "FF & E": t_ffe + t_misc, "M.E.P WORKS": t_mep,
            "UTILITY CONNECTION": t_utility, "EXTERNAL WORKS": t_external, "FACILITY": t_pub_fac + t_res_fac + t_proj_fac,
            "PRELIMINARIES WORKS": t_preliminary, "CONTINGENCIES": t_contingency, "HARDCOST": grand_total_hc,
            "CONSULTANCY SERVICES FEE": t_consultancy, "QS SERVICES": t_qs, 
            "PROJECT MANAGEMENT SERVICES": t_pm, "INSURANCE COVERAGE": t_insurance,
            "SOFTCOST": total_soft_cost, "TOTAL, EXCLD PPN": grand_total_project
        }
    
    if "recap_math_engine" not in st.session_state:
        st.session_state.recap_math_engine = get_recap_values

    # --- Generate Global Totals ---
    tot_cache = {}; global_cost = {}; tot_gba = tot_gfa = tot_sgfa = 0
    for pid, pdata in projects.items():
        vals = get_recap_values(pdata)
        tot_cache[pid] = vals
        for k, v in vals.items(): global_cost[k] = global_cost.get(k, 0) + v
        d = pdata.get("data", {})
        tot_gba += _safe_float(d.get("m_gba", 0)); tot_gfa += _safe_float(d.get("m_gfa", 0)); tot_sgfa += _safe_float(d.get("m_sgfa", 0))

    # Calculate global % divisors
    global_hc = global_cost.get("HARDCOST", 0)
    global_sc = global_cost.get("SOFTCOST", 0)
    safe_hc = global_hc if global_hc > 0 else 1
    safe_sc = global_sc if global_sc > 0 else 1

    # --- 1. Styling Definitions ---
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    white_font, bold_font, reg_font = Font(color="FFFFFF", bold=True, size=10), Font(bold=True, size=9), Font(size=9)
    black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align, left_align = Alignment(horizontal='center', vertical='center', wrap_text=True), Alignment(horizontal='left', vertical='center')

    # --- 2. Blue Metadata Header ---
    headers = [["ASG GROUP PROPERTY DEVELOPMENT", f"VERSION : {port_meta.get('version', '')}"], ["QS & PROCUREMENT DIVISION", f"UPDATED : {port_meta.get('updated', '')}"], [port_meta.get('title', ''), f"CREATED : {port_meta.get('created', '')}"], [port_meta.get('ref', ''), ""]]
    for r_idx, (text_left, text_right) in enumerate(headers, 1):
        for c in range(1, 100): ws.cell(row=r_idx, column=c).fill = blue_fill
        ws.cell(row=r_idx, column=1, value=text_left).font = white_font
        ws.cell(row=r_idx, column=15, value=text_right).font = white_font
        ws.cell(row=r_idx, column=15).alignment = Alignment(horizontal='right', vertical='center')

    # --- 3. Static Table Headers ---
    static_cols = [("SN", 1), ("DESCRIPTION", 2), ("COA", 3), ("%", 4)]
    for name, col_idx in static_cols:
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=9, end_column=col_idx)
        c = ws.cell(row=6, column=col_idx, value=name)
        c.alignment, c.font, c.fill = center_align, bold_font, PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for r in range(6, 10): ws.cell(row=r, column=col_idx).border = black_border

    # --- 4. Dynamic Project Headers ---
    bg_colors = ["EAEAEA", "FCE4D6", "F2DCDB", "E1D5E7", "DDEBF7", "E2EFDA", "D9E1F2", "F4B084", "FFF2CC"]
    project_list = [("TOTAL", {"name": "TOTAL"})] + list(projects.items())
    current_col = 5
    
    for i, (pid, pdata) in enumerate(project_list):
        group_fill = PatternFill(start_color=bg_colors[i % len(bg_colors)], end_color=bg_colors[i % len(bg_colors)], fill_type="solid")
        ws.merge_cells(start_row=6, start_column=current_col, end_row=6, end_column=current_col+4)
        c = ws.cell(row=6, column=current_col, value=pdata.get('name', 'PROJECT').upper())
        c.alignment, c.font, c.fill = center_align, bold_font, group_fill
        
        ws.cell(row=7, column=current_col, value="ESTIMATE").alignment = center_align
        ws.merge_cells(start_row=7, start_column=current_col+1, end_row=7, end_column=current_col+4)
        ws.cell(row=7, column=current_col+1, value="Cost Ratio (Rp/m2)").alignment = center_align
        
        for j, lbl in enumerate(["TOTAL", "GBA", "GFA", "SGFA", "NFA"]):
            ws.cell(row=8, column=current_col+j, value=lbl).alignment = center_align
            
        if pid == "TOTAL":
            gba, gfa, sgfa, nfa = tot_gba, tot_gfa, tot_sgfa, tot_gfa * 0.82
        else:
            d = pdata.get("data", {})
            gba, gfa, sgfa = _safe_float(d.get("m_gba", 0)), _safe_float(d.get("m_gfa", 0)), _safe_float(d.get("m_sgfa", 0))
            nfa = gfa * 0.82
        
        gba_f = gba if gba > 0 else 1; gfa_f = gfa if gfa > 0 else 1; sgfa_f = sgfa if sgfa > 0 else 1; nfa_f = nfa if nfa > 0 else 1
            
        for j, val in enumerate(["Rp", gba, gfa, sgfa, nfa]):
            c = ws.cell(row=9, column=current_col+j, value=val)
            c.alignment = center_align
            if j > 0: c.number_format = '#,##0'

        for r in range(6, 10):
            for c_idx in range(current_col, current_col+5):
                ws.cell(row=r, column=c_idx).fill = group_fill
                ws.cell(row=r, column=c_idx).border = black_border
        current_col += 5

    # --- 5. Data Rows Map (With Categories for %) ---
    row_mapping = [
        ("I", "HARDCOST", "118-14-000", True, "HC"), ("1", "PRELIMINARIES WORKS", "118-14-100", False, "HC"), 
        ("2", "EARTHWORKS", "118-14-200", False, "HC"), ("3", "FOUNDATIONS", "118-14-300", False, "HC"), 
        ("4", "STRUCTURAL WORKS", "118-14-500", False, "HC"), ("5", "ARCHITECTURAL WORKS", "118-14-600", False, "HC"),
        ("6", "FF & E", "118-14-700", False, "HC"), ("7", "M.E.P WORKS", "118-14-800", False, "HC"), 
        ("8", "UTILITY CONNECTION", "118-13-900", False, "HC"), ("9", "EXTERNAL WORKS", "118-14-930", False, "HC"), 
        ("10", "FACILITY", "118-14-960", False, "HC"), ("11", "CONTINGENCIES", "", False, "HC"),
        ("II", "SOFTCOST", "118-13-000", True, "SC_TOTAL"), ("1", "CONSULTANCY SERVICES FEE", "118-13-202", False, "SC"), 
        ("2", "QS SERVICES", "118-13-201", False, "SC"), ("3", "PROJECT MANAGEMENT SERVICES", "118-13-203", False, "SC"), 
        ("4", "INSURANCE COVERAGE", "118-13-300", False, "SC"), ("IV", "TOTAL, EXCLD PPN", "", True, "TOTAL")
    ]

    r_idx = 10
    for sn, desc, coa, is_bold, cat in row_mapping:
        # Calculate % based on Global Totals and Category
        global_val = global_cost.get(desc, 0)
        if cat == "HC": pct = global_val / safe_hc
        elif cat == "SC": pct = global_val / safe_sc
        elif cat == "SC_TOTAL": pct = global_val / safe_hc # Softcost total vs Hardcost
        elif cat == "TOTAL": pct = global_val / safe_hc # Grand total vs Hardcost
        else: pct = 0

        ws.cell(row=r_idx, column=1, value=sn).alignment = center_align
        ws.cell(row=r_idx, column=2, value=desc).alignment = left_align
        ws.cell(row=r_idx, column=3, value=coa).alignment = center_align
        
        # Write Percentage
        pct_cell = ws.cell(row=r_idx, column=4, value=pct)
        pct_cell.alignment = center_align
        pct_cell.number_format = '0.00%'

        for c in range(1, 5):
            ws.cell(row=r_idx, column=c).border = black_border
            ws.cell(row=r_idx, column=c).font = bold_font if is_bold else reg_font

        col_offset = 5
        for pid, pdata in project_list:
            val = global_cost.get(desc, 0) if pid == "TOTAL" else tot_cache[pid].get(desc, 0)
            
            if pid == "TOTAL":
                gba_f = tot_gba if tot_gba > 0 else 1; gfa_f = tot_gfa if tot_gfa > 0 else 1
                sgfa_f = tot_sgfa if tot_sgfa > 0 else 1; nfa_f = (tot_gfa*0.82) if tot_gfa > 0 else 1
            else:
                d = pdata.get("data", {})
                gba_f = _safe_float(d.get("m_gba", 1) if d.get("m_gba", 0) > 0 else 1)
                gfa_f = _safe_float(d.get("m_gfa", 1) if d.get("m_gfa", 0) > 0 else 1)
                sgfa_f = _safe_float(d.get("m_sgfa", 1) if d.get("m_sgfa", 0) > 0 else 1)
                nfa_f = gfa_f * 0.82
            
            for j, v in enumerate([val, val/gba_f, val/gfa_f, val/sgfa_f, val/nfa_f]):
                c = ws.cell(row=r_idx, column=col_offset+j, value=v)
                c.number_format = '#,##0'
                c.border = black_border
                c.font = bold_font if is_bold else reg_font
                if is_bold: c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            col_offset += 5
        r_idx += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    for c in range(5, col_offset): ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = 'E10'
    wb.save(output)
    return output.getvalue()
#endregion

def show_portfolio_summary():
    st.title("Summary")
    
    # ==========================================
    # 1. INITIALIZE EDITABLE SESSION STATE
    # ==========================================
    init_report_config()

    port_meta = get_port_meta()
    port_assumptions_df = get_port_assumptions_df()

    # ==========================================
    # 1B. SHARED SUMMARY UI + HEADER RENDERER
    # ==========================================
    from html import escape

    def safe_text(value):
        return escape(str(value if value is not None else ""))

    def render_portfolio_header(meta):
        title = safe_text(meta.get("title", ""))
        ref = safe_text(meta.get("ref", ""))
        version = safe_text(meta.get("version", ""))
        updated = safe_text(meta.get("updated", ""))
        created = safe_text(meta.get("created", ""))

        return f"""
<div class="asg-header">
<div class="asg-header-left">
ASG GROUP PROPERTY DEVELOPMENT<br>
QS & PROCUREMENT DIVISION<br>
{title}<br>
{ref}
</div>
<div class="asg-header-right">
VERSION &nbsp;&nbsp;: {version}<br>
UPDATED &nbsp;: {updated}<br>
CREATED &nbsp;: {created}
</div>
</div>
        """

    SUMMARY_CSS = """
    <style>
        /* ===============================
           CONFIG PAGE
        =============================== */
        .summary-config-panel {
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            padding: 16px 18px;
            margin-bottom: 14px;
            box-shadow: 0 1px 3px rgba(16,24,40,0.04);
        }

        .summary-config-title {
            font-size: 12px;
            font-weight: 750;
            color: #111827;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            margin-bottom: 8px;
        }

        .summary-config-desc {
            font-size: 13px;
            color: #6B7280;
            line-height: 1.55;
            margin-bottom: 12px;
        }

        .summary-kpi-grid {
            display: grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 12px;
            margin-bottom: 14px;
        }

        .summary-kpi-card {
            background: #F9FAFB;
            border: 1px solid #E5E7EB;
            border-radius: 12px;
            padding: 12px 14px;
        }

        .summary-kpi-label {
            font-size: 11px;
            color: #6B7280;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            font-weight: 700;
            margin-bottom: 5px;
        }

        .summary-kpi-value {
            font-size: 16px;
            color: #111827;
            font-weight: 750;
            line-height: 1.35;
        }

        .summary-note-box {
            background: #F8F9FF;
            border: 1px solid #DDE1FF;
            border-left: 4px solid #3E4095;
            border-radius: 12px;
            padding: 12px 14px;
            font-size: 13px;
            color: #4B5563;
            line-height: 1.55;
            margin-bottom: 14px;
        }

        /* ===============================
           SHARED ASG REPORT PREVIEW
        =============================== */
        .asg-container {
            font-family: Calibri, Arial, sans-serif;
            font-size: 13px;
            color: #000000 !important;
            background-color: #FFFFFF !important;
            padding: 15px;
            border: 1px solid #D1D5DB;
            border-radius: 8px;
            box-shadow: 0 1px 3px rgba(16,24,40,0.05);
            overflow-x: auto;
        }

        .asg-header {
            background-color: #0070C0 !important;
            color: #FFFFFF !important;
            padding: 7px 12px;
            font-weight: bold;
            font-size: 13px;
            display: flex;
            justify-content: space-between;
            gap: 16px;
            line-height: 1.45;
            margin-bottom: 15px;
            border: 1px solid #005A9C;
        }

        .asg-header-left {
            text-align: left;
        }

        .asg-header-right {
            text-align: right;
            min-width: 190px;
        }

        .asg-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background-color: #FFFFFF !important;
        }

        .asg-table th,
        .asg-table td {
            border: 2px solid #000000 !important;
            padding: 5px 8px;
            text-align: right;
            vertical-align: middle;
            color: #000000 !important;
        }

        .asg-table th {
            background-color: #F2F2F2 !important;
            text-align: center;
            font-weight: bold;
            color: #000000 !important;
        }

        .asg-table td.left {
            text-align: left;
            font-weight: bold;
        }

        .asg-table td.center {
            text-align: center;
            font-weight: bold;
        }

        .asg-table .bold-row td {
            font-weight: bold;
        }

        .asg-assumptions {
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
            background-color: #FFFFFF !important;
        }

        .asg-assumptions td {
            border: 1px solid #D9D9D9 !important;
            padding: 4px 8px;
            color: #000000 !important;
            vertical-align: top;
        }

        .asg-assumptions .yellow-header {
            background-color: #FFD966 !important;
            font-weight: bold;
            text-align: left;
            color: #000000 !important;
        }

        /* ===============================
           REKAP TABLE CONTAINER
        =============================== */
        .recap-wrapper {
            width: 100%;
            overflow-x: auto;
            font-family: Calibri, Arial, sans-serif;
            font-size: 11px;
        }

        .recap-table {
            border-collapse: separate;
            border-spacing: 0;
            white-space: nowrap;
        }

        .recap-table th {
            text-align: center !important;
            font-weight: bold;
            vertical-align: middle;
            border: 1px solid #000;
            padding: 4px 6px;
            background-color: #FFFFFF;
        }

        .recap-table td {
            border-right: 1px solid #000;
            border-bottom: 1px solid #000;
            border-left: 1px solid #000;
            padding: 4px 6px;
            background-color: #FFFFFF;
        }

        .sticky-col,
        .sticky-col2 {
            position: sticky;
            background-color: #F2F2F2 !important;
            z-index: 5;
        }

        .sticky-col3,
        .sticky-col4 {
            background-color: #F2F2F2 !important;
            z-index: 5;
        }

        .sticky-col {
            left: 0;
        }

        .sticky-col2 {
            left: 20px;
            text-align: left !important;
            min-width: 200px;
        }

        .bold-row {
            font-weight: bold;
            background-color: #F9F9F9;
        }

        @media (max-width: 900px) {
            .summary-kpi-grid {
                grid-template-columns: 1fr;
            }

            .asg-header {
                display: block;
            }

            .asg-header-right {
                text-align: left;
                margin-top: 8px;
            }
        }
    </style>
    """

    st.markdown(SUMMARY_CSS, unsafe_allow_html=True)

    # ==========================================
    # 2. TABS SETUP
    # ==========================================
    summary_list = ["Config", "FAD", "Rekap"]
    summary_tabs = st.tabs(summary_list)
    
    # --- TAB 1: EDITABLE NATIVE COMPONENTS ---
    # --- TAB 1: CONFIG ---
    with summary_tabs[0]:
        st.subheader("Config")
        st.caption("Configure the report header and assumptions used in both FAD and Rekap previews.")

        cfg = get_report_config()
        export_settings = cfg.setdefault("export_settings", {
            "prepared_by": "",
            "checked_by": ""
        })

        st.markdown("##### Approval / Export Settings")

        c_prepared, c_checked = st.columns(2)

        export_settings["prepared_by"] = c_prepared.text_input(
            "Prepared By",
            value=export_settings.get("prepared_by", "")
        )

        export_settings["checked_by"] = c_checked.text_input(
            "Checked By",
            value=export_settings.get("checked_by", "")
        )

        project_count = len(st.session_state.projects)
        assumption_count = len(port_assumptions_df)

        st.markdown(
            f"""
            <div class="summary-kpi-grid">
                <div class="summary-kpi-card">
                    <div class="summary-kpi-label">Active Projects</div>
                    <div class="summary-kpi-value">{project_count}</div>
                </div>
                <div class="summary-kpi-card">
                    <div class="summary-kpi-label">Assumptions</div>
                    <div class="summary-kpi-value">{assumption_count} items</div>
                </div>
                <div class="summary-kpi-card">
                    <div class="summary-kpi-label">Output Format</div>
                    <div class="summary-kpi-value">FAD + Rekap</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        col_cfg, col_preview = st.columns([1.1, 1], gap="large")

        with col_cfg:
            with st.container(border=True):
                st.markdown("##### Header Configuration")

                st.markdown(
                    """
                    <div class="summary-config-desc">
                        These fields control the blue report header. The same header will appear in both FAD and Rekap.
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                port_meta["title"] = st.text_area(
                    "Project Title",
                    value=port_meta["title"],
                    height=80
                )

                port_meta["ref"] = st.text_area(
                    "Reference Data",
                    value=port_meta["ref"],
                    height=70,
                    help="Drawing reference, data source, or revision reference."
                )

                c1, c2, c3 = st.columns(3)

                port_meta["version"] = c1.text_input(
                    "Version",
                    value=port_meta["version"]
                )

                port_meta["updated"] = c2.text_input(
                    "Updated Date",
                    value=port_meta["updated"]
                )

                port_meta["created"] = c3.text_input(
                    "Created Date",
                    value=port_meta["created"]
                )

        with col_preview:
            st.markdown("##### Header Preview")

            st.markdown(
                f"""
<div class="summary-note-box">
This preview uses the same header renderer used in FAD and Rekap.
Adjust the fields on the left and the preview will update automatically.
</div>
<div class="asg-container">
{render_portfolio_header(port_meta)}
</div>
                """,
                unsafe_allow_html=True
            )

        st.divider()

        st.markdown("##### Assumptions Configuration")
        st.caption("These assumptions are displayed below the FAD table. Keep the wording aligned with the required report format.")

        edited_assumptions = st.data_editor(
            get_port_assumptions_df(),
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "No.": st.column_config.TextColumn("No.", width="small"),
                "Assumption Description": st.column_config.TextColumn("Assumption Description", width="large"),
            }
        )

        set_port_assumptions_df(edited_assumptions)

        if st.button("Save Config", type="primary"):
            save_data()
            st.success("Config saved.")

    # --- TAB 2: EXACT FORMAT MIRROR (HTML/CSS) ---
    with summary_tabs[1]:
        st.subheader("Feasibility Analysis Data (FAD)")

        # 1. DATA PREPARATION (Define raw_data BEFORE anything else)
        raw_data = []
        tot_gba = tot_gfa = tot_sgfa = tot_budget = 0
        
        # Loop through projects to build the data list
        for sn, (pid, pdata) in enumerate(st.session_state.projects.items(), 1):
            # Recalculate fresh numbers
            gba, gfa, sgfa, budget, qty = calculate_project_totals(pdata, pdata.get("type", "Hotel"))
            
            raw_data.append({
                "SN": sn,
                "AREA": pdata.get("name", f"Project {sn}"),
                "GBA": gba, "GFA": gfa, "SGFA": sgfa,
                "QTY": qty, 
                "UNIT": "Units" if "Hotel" not in pdata.get("type", "") else "RoomKey",
                "BUDGET": budget,
                "R_GBA": budget / gba if gba > 0 else 0,
                "R_GFA": budget / gfa if gfa > 0 else 0,
                "R_SGFA": budget / sgfa if sgfa > 0 else 0
            })
            tot_gba += gba; tot_gfa += gfa; tot_sgfa += sgfa; tot_budget += budget

        # Add the TOTAL row
        raw_data.append({
            "SN": "", "AREA": "TOTAL",
            "GBA": tot_gba, "GFA": tot_gfa, "SGFA": tot_sgfa,
            "QTY": None, "UNIT": "TOTAL",
            "BUDGET": tot_budget,
            "R_GBA": tot_budget / tot_gba if tot_gba > 0 else 0,
            "R_GFA": tot_budget / tot_gfa if tot_gfa > 0 else 0,
            "R_SGFA": tot_budget / tot_sgfa if tot_sgfa > 0 else 0
        })

        # 2. UI CONTROLS (Now they can safely see raw_data)
        col_btn1, col_btn2, _ = st.columns([1.5, 1.5, 3])
        
        with col_btn1:
            # This will now work because raw_data was defined in Step 1
            excel_output = generate_exact_portfolio_excel(
                port_meta,
                raw_data,
                get_port_assumptions_df()
            )
            
            st.download_button(
                label="Download Excel",
                data=excel_output,
                file_name="ASG_Portfolio_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
                
        header_html = f"""
        <div class="asg-container">
            {render_portfolio_header(port_meta)}
        """

        # 3. Dynamic Data Table Core
        table_start = """
<table class="asg-table">
<thead>
<tr>
<th rowspan="2" style="width: 3%;">SN</th>
<th rowspan="2" style="width: 18%;">AREA</th>
<th colspan="3">BUILDING AREA (M2)</th>
<th colspan="2" style="width: 10%;">UNIT</th>
<th rowspan="2" style="width: 14%;">BUDGET ESTIMATE<br>RP</th>
<th colspan="3">COST RATIO RP/M2</th>
</tr>
<tr>
<th>GBA</th><th>GFA</th><th>SGFA</th>
<th></th><th></th>
<th>GBA</th><th>GFA</th><th>SGFA</th>
</tr>
</thead>
<tbody>
        """
        
        # 4. Generate Dynamic Rows from Active Projects
        table_rows = ""
        tot_gba = tot_gfa = tot_sgfa = tot_budget = 0
        
        for sn, (pid, pdata) in enumerate(st.session_state.projects.items(), 1):
            d = pdata.get("data", {})
            name = pdata.get("name", f"Project {sn}")
            ptype = pdata.get("type", "")
            
            gba = _safe_float(d.get("m_gba", 0.0))
            gfa = _safe_float(d.get("m_gfa", 0.0))
            sgfa = _safe_float(d.get("m_sgfa", 0.0))
            qty = _safe_float(d.get("m_rooms", 0.0))
            budget = _safe_float(d.get("grand_total_project", 0.0))
            
            if "Hotel" in ptype: unit_lbl = "RoomKey"
            elif "Parking" in ptype: unit_lbl = "lots"
            else: unit_lbl = "Units"
            
            r_gba = budget / gba if gba > 0 else 0
            r_gfa = budget / gfa if gfa > 0 else 0
            r_sgfa = budget / sgfa if sgfa > 0 else 0
            
            table_rows += f"""<tr class="bold-row">
<td class="center">{sn}</td>
<td class="left">{name}</td>
<td>{gba:,.2f}</td><td>{gfa:,.2f}</td><td>{sgfa:,.2f}</td>
<td class="center">{qty:,.0f}</td><td class="center">{unit_lbl}</td>
<td>{budget:,.0f}</td>
<td>{r_gba:,.0f}</td><td>{r_gfa:,.0f}</td><td>{r_sgfa:,.0f}</td>
</tr>"""
            
            tot_gba += gba
            tot_gfa += gfa
            tot_sgfa += sgfa
            tot_budget += budget

        tot_r_gba = tot_budget / tot_gba if tot_gba > 0 else 0
        tot_r_gfa = tot_budget / tot_gfa if tot_gfa > 0 else 0
        tot_r_sgfa = tot_budget / tot_sgfa if tot_sgfa > 0 else 0

        table_end = f"""<tr class="bold-row" style="background-color: #F2F2F2;">
<td class="center" colspan="2">TOTAL</td>
<td>{tot_gba:,.2f}</td><td>{tot_gfa:,.2f}</td><td>{tot_sgfa:,.2f}</td>
<td class="center" colspan="2">TOTAL</td>
<td>{tot_budget:,.0f}</td>
<td>{tot_r_gba:,.0f}</td><td>{tot_r_gfa:,.0f}</td><td>{tot_r_sgfa:,.0f}</td>
</tr>
</tbody>
</table>"""

        assumptions_html = """<table class="asg-assumptions">
<tr>
    <td class="yellow-header" style="width: 3%;">I.</td>
    <td class="yellow-header">ASSUMPTIONS</td>
</tr>"""
        for _, row in get_port_assumptions_df().iterrows():
            num = row.get("No.", "")
            desc = row.get("Assumption Description", "")
            if pd.notna(desc) and str(desc).strip() != "":
                assumptions_html += f"""<tr>
<td style="text-align: center;">{num}</td>
<td>{desc}</td>
</tr>"""
        assumptions_html += """</table>"""

        full_html = header_html + table_start + table_rows + table_end + assumptions_html + "</div>"

        st.markdown(full_html, unsafe_allow_html=True)

# --- TAB 3: WIDE RECAP COST ---
    with summary_tabs[2]:
        st.subheader("Comprehensive Recap Matrix (Cost & Ratios)")
        
        if "recap_math_engine" not in st.session_state:
            _ = generate_recap_excel(port_meta, st.session_state.projects)
            
        math_engine = st.session_state.recap_math_engine
        
        col_btn, col_info = st.columns([1.5, 4.5])
        with col_btn:
            recap_excel_data = generate_recap_excel(port_meta, st.session_state.projects)
            st.download_button(
                label="Download Excel",
                data=recap_excel_data,
                file_name="ASG_Recap_Cost_Wide.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )


        # --- GENERATE HTML PREVIEW ---
        bg_colors = ["#EAEAEA", "#FCE4D6", "#F2DCDB", "#E1D5E7", "#DDEBF7", "#E2EFDA", "#D9E1F2", "#F4B084", "#FFF2CC"]
        project_list = [("TOTAL", {"name": "TOTAL"})] + list(st.session_state.projects.items())

        tot_cache = {}; global_cost = {}; tot_gba = tot_gfa = tot_sgfa = 0
        for pid, pdata in st.session_state.projects.items():
            vals = math_engine(pdata)
            tot_cache[pid] = vals
            for k, v in vals.items(): global_cost[k] = global_cost.get(k, 0) + v
            d = pdata.get("data", {})
            tot_gba += _safe_float(d.get("m_gba", 0)); tot_gfa += _safe_float(d.get("m_gfa", 0)); tot_sgfa += _safe_float(d.get("m_sgfa", 0))

        global_hc = global_cost.get("HARDCOST", 0); safe_hc = global_hc if global_hc > 0 else 1
        global_sc = global_cost.get("SOFTCOST", 0); safe_sc = global_sc if global_sc > 0 else 1

        html_str = f"""
        <div class="asg-container">
        {render_portfolio_header(port_meta)}

<div class="recap-wrapper">
<table class="recap-table">
        """

        html_str += """
<tr><th rowspan='4' class='sticky-col'>SN</th>
<th rowspan='4' class='sticky-col2' style='min-width:200px;'>DESCRIPTION</th>
<th rowspan='4' class='sticky-col3'>COA</th><th rowspan='4' class='sticky-col4'>%</th>
        """
        for i, (pid, pdata) in enumerate(project_list):
            color = bg_colors[i % len(bg_colors)]
            html_str += f"<th colspan='5' style='background-color:{color}; color:#000;'>{pdata.get('name', 'PROJECT').upper()}</th>"
        html_str += "</tr><tr>"
        for i in range(len(project_list)):
            color = bg_colors[i % len(bg_colors)]
            html_str += f"<th style='background-color:{color};'>ESTIMATE</th><th colspan='4' style='background-color:{color};'>Cost Ratio (Rp/m2)</th>"
        html_str += "</tr><tr>"
        for i in range(len(project_list)):
            color = bg_colors[i % len(bg_colors)]
            html_str += f"<th style='background-color:{color};'>TOTAL</th><th style='background-color:{color};'>GBA</th><th style='background-color:{color};'>GFA</th><th style='background-color:{color};'>SGFA</th><th style='background-color:{color};'>NFA</th>"
        html_str += "</tr><tr>"
        for i, (pid, pdata) in enumerate(project_list):
            color = bg_colors[i % len(bg_colors)]
            if pid == "TOTAL":
                gba, gfa, sgfa, nfa = tot_gba, tot_gfa, tot_sgfa, tot_gfa * 0.82
            else:
                d = pdata.get("data", {})
                gba, gfa, sgfa = _safe_float(d.get("m_gba", 0)), _safe_float(d.get("m_gfa", 0)), _safe_float(d.get("m_sgfa", 0))
                nfa = gfa * 0.82
            html_str += f"<th style='background-color:{color};'>Rp</th><th style='background-color:{color};'>{gba:,.0f}</th><th style='background-color:{color};'>{gfa:,.0f}</th><th style='background-color:{color};'>{sgfa:,.0f}</th><th style='background-color:{color};'>{nfa:,.0f}</th>"
        html_str += "</tr>"

        row_mapping = [
            ("I", "HARDCOST", "118-14-000", True, "HC"), ("1", "PRELIMINARIES WORKS", "118-14-100", False, "HC"), 
            ("2", "EARTHWORKS", "118-14-200", False, "HC"), ("3", "FOUNDATIONS", "118-14-300", False, "HC"), 
            ("4", "STRUCTURAL WORKS", "118-14-500", False, "HC"), ("5", "ARCHITECTURAL WORKS", "118-14-600", False, "HC"),
            ("6", "FF & E", "118-14-700", False, "HC"), ("7", "M.E.P WORKS", "118-14-800", False, "HC"), 
            ("8", "UTILITY CONNECTION", "118-13-900", False, "HC"), ("9", "EXTERNAL WORKS", "118-14-930", False, "HC"), 
            ("10", "FACILITY", "118-14-960", False, "HC"), ("11", "CONTINGENCIES", "", False, "HC"),
            ("II", "SOFTCOST", "118-13-000", True, "SC_TOTAL"), ("1", "CONSULTANCY SERVICES FEE", "118-13-202", False, "SC"), 
            ("2", "QS SERVICES", "118-13-201", False, "SC"), ("3", "PROJECT MANAGEMENT SERVICES", "118-13-203", False, "SC"), 
            ("4", "INSURANCE COVERAGE", "118-13-300", False, "SC"), ("IV", "TOTAL, EXCLD PPN", "", True, "TOTAL")
        ]

        for sn, desc, coa, is_bold, cat in row_mapping:
            global_val = global_cost.get(desc, 0)
            if cat == "HC": pct = global_val / safe_hc
            elif cat == "SC": pct = global_val / safe_sc
            elif cat in ["SC_TOTAL", "TOTAL"]: pct = global_val / safe_hc
            else: pct = 0

            tr_class = " class='bold-row'" if is_bold else ""
            html_str += f"<tr{tr_class}>"
            # Sticky Columns (Anchor columns remain neutral grey)
            html_str += f"<td class='sticky-col' style='text-align:center;'>{sn}</td>"
            html_str += f"<td class='sticky-col2'>{desc}</td>"
            html_str += f"<td class='sticky-col3'>{coa}</td>"
            html_str += f"<td class='sticky-col4'>{pct*100:.2f}%</td>"
            
            # Project Data Columns (Colored to match headers)
            for i, (pid, pdata) in enumerate(project_list):
                val = global_cost.get(desc, 0) if pid == "TOTAL" else tot_cache[pid].get(desc, 0)
                color = bg_colors[i % len(bg_colors)] # Get the header's color
                
                # Calculate divisors
                if pid == "TOTAL":
                    gba_f, gfa_f, sgfa_f = tot_gba or 1, tot_gfa or 1, tot_sgfa or 1
                    nfa_f = (tot_gfa * 0.82) or 1
                else:
                    d = pdata.get("data", {})
                    gba_f = _safe_float(d.get("m_gba") or 1)
                    gfa_f = _safe_float(d.get("m_gfa") or 1)
                    sgfa_f = _safe_float(d.get("m_sgfa") or 1)
                    nfa_f = gfa_f * 0.82 or 1
                
                # Apply the background color style to every <td> in this column
                c_style = f"style='background-color:{color};'"
                
                html_str += f"<td {c_style}>{val:,.0f}</td>"
                html_str += f"<td {c_style}>{val/gba_f:,.0f}</td>"
                html_str += f"<td {c_style}>{val/gfa_f:,.0f}</td>"
                html_str += f"<td {c_style}>{val/sgfa_f:,.0f}</td>"
                html_str += f"<td {c_style}>{val/nfa_f:,.0f}</td>"
            
            # Spacer for desktop 'over-scroll' comparison
            html_str += "<td style='border:none; background:transparent; min-width:600px;'></td>"
            html_str += "</tr>"

        html_str += "</table></div></div>"
        st.markdown(html_str, unsafe_allow_html=True)

def show_snapshots():
    st.title("Archive")
    curr_id, curr_proj = get_current_project()

    atab1, atab2= st.tabs([
        "Online Backup", "Offline Backup"
    ])
    with atab1:
        # --- SAVE NEW SNAPSHOT ---
        st.header("Online Backup")
        st.subheader("Save File")
        col1, _ = st.columns([4, 3])
        snapshot_name = col1.text_input(
            "Name", 
            placeholder="e.g. Project X - Opt 1 - Rev 0"
        )
        col1, _ = st.columns([1, 6])
        if col1.button("Save", use_container_width=True, icon=icon_safe("save_as")):
            if snapshot_name.strip() == "":
                col1.warning("Please enter Project name.")
            else:
                if save_snapshot(snapshot_name):
                    st.success(f"Project **{snapshot_name}** saved!")
                    st.rerun()

        st.divider()

        # --- LIST EXISTING SNAPSHOTS ---
        st.subheader("Load File")
        snapshots = load_snapshots()

        if not snapshots:
            st.info("No saved projects yet.")
        else:
            # ==================================================
            # PAGINATION SETUP
            # ==================================================
            PAGE_SIZE = 10

            if "archive_page" not in st.session_state:
                st.session_state.archive_page = 0

            total_items = len(snapshots)
            total_pages = max(1, (total_items + PAGE_SIZE - 1) // PAGE_SIZE)

            # Safety clamp if files were deleted / changed
            if st.session_state.archive_page >= total_pages:
                st.session_state.archive_page = total_pages - 1

            if st.session_state.archive_page < 0:
                st.session_state.archive_page = 0

            start_idx = st.session_state.archive_page * PAGE_SIZE
            end_idx = start_idx + PAGE_SIZE
            page_snapshots = snapshots[start_idx:end_idx]

            st.caption(
                f"Showing {start_idx + 1}–{min(end_idx, total_items)} of {total_items} saved files"
            )

            st.divider()

            # ==================================================
            # PAGINATED SNAPSHOT LIST
            # ==================================================
            for snap in page_snapshots:
                snap_id = snap["id"]

                rename_key = f"renaming_{snap_id}"
                delete_key = f"deleting_{snap_id}"

                if rename_key not in st.session_state:
                    st.session_state[rename_key] = False

                if delete_key not in st.session_state:
                    st.session_state[delete_key] = False

                col1, col2, col3, col4 = st.columns([4, 1, 1, 1])

                from datetime import datetime, timedelta

                created_utc = datetime.fromisoformat(snap["created_at"].replace("Z", "+00:00"))
                created_local = created_utc + timedelta(hours=7)
                formatted_date = created_local.strftime("%d %b %Y, %H:%M")

                is_active = st.session_state.get("loaded_snapshot_id") == snap_id
                active_label = " — ACTIVE" if is_active else ""

                # ==================================================
                # NORMAL DISPLAY MODE
                # ==================================================
                if not st.session_state[rename_key] and not st.session_state[delete_key]:
                    with col1:
                        st.markdown(f"**{snap['snapshot_name']}**")
                        st.caption(f"Saved: {formatted_date} WIB{active_label}")

                    if col2.button("Rename", key=f"rename_start_{snap_id}", use_container_width=True):
                        st.session_state[rename_key] = True
                        st.session_state[delete_key] = False
                        st.rerun()

                    if col3.button("Load", key=f"load_{snap_id}", type="primary", use_container_width=True):
                        data = load_snapshot_data(snap_id)

                        if data:
                            restore_app_payload(data)

                            st.session_state.loaded_snapshot_id = snap_id
                            st.session_state.loaded_snapshot_name = snap["snapshot_name"]

                            save_data()
                            st.success(f"Loaded **{snap['snapshot_name']}**.")
                            st.rerun()

                    if col4.button("Delete", key=f"delete_start_{snap_id}", use_container_width=True):
                        st.session_state[delete_key] = True
                        st.session_state[rename_key] = False
                        st.rerun()

                # ==================================================
                # RENAME MODE
                # ==================================================
                elif st.session_state[rename_key]:
                    with col1:
                        new_archive_name = st.text_input(
                            "New file name",
                            value=snap["snapshot_name"],
                            key=f"rename_input_{snap_id}",
                            label_visibility="collapsed"
                        )
                        st.caption(f"Saved: {formatted_date} WIB{active_label}")

                    if col2.button("Save Name", key=f"rename_save_{snap_id}", type="primary", use_container_width=True):
                        if rename_snapshot(snap_id, new_archive_name):
                            st.session_state[rename_key] = False
                            st.success("File renamed.")
                            st.rerun()

                    if col3.button("Cancel", key=f"rename_cancel_{snap_id}", use_container_width=True):
                        st.session_state[rename_key] = False
                        st.rerun()

                    with col4:
                        st.empty()

                # ==================================================
                # DELETE CONFIRMATION MODE
                # ==================================================
                elif st.session_state[delete_key]:
                    with col1:
                        st.warning(f"Delete **{snap['snapshot_name']}**?")
                        st.caption("This archived file will be removed from the library.")

                    if col2.button("Confirm", key=f"delete_confirm_{snap_id}", type="primary", use_container_width=True):
                        if delete_snapshot(snap_id):
                            if st.session_state.get("loaded_snapshot_id") == snap_id:
                                st.session_state.loaded_snapshot_id = None
                                st.session_state.loaded_snapshot_name = None
                                save_data()

                            st.session_state[delete_key] = False
                            st.success("File deleted.")
                            st.rerun()

                    if col3.button("Cancel", key=f"delete_cancel_{snap_id}", use_container_width=True):
                        st.session_state[delete_key] = False
                        st.rerun()

                    with col4:
                        st.empty()

                st.divider()

            # ==================================================
            # PAGINATION CONTROLS
            # ==================================================
            c1, col_prev, col_page, col_next, c2 = st.columns([5, 1, 2, 1, 5])

            with col_prev:
                if st.button(
                    "Previous",
                    key="archive_prev_page",
                    use_container_width=True,
                    disabled=st.session_state.archive_page <= 0
                ):
                    st.session_state.archive_page -= 1
                    st.rerun()

            with col_page:
                st.markdown(
                    f"<div style='text-align:center; padding-top: 0.45rem;'>"
                    f"Page {st.session_state.archive_page + 1} of {total_pages}"
                    f"</div>",
                    unsafe_allow_html=True
                )

            with col_next:
                if st.button(
                    "Next",
                    key="archive_next_page",
                    use_container_width=True,
                    disabled=st.session_state.archive_page >= total_pages - 1
                ):
                    st.session_state.archive_page += 1
                    st.rerun()  

    with atab2:
        st.header("Offline Backup")
        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Import")
            uploaded_file = st.file_uploader("Upload CSV Database:", type=["csv"])

            if uploaded_file is not None:
                file_key = getattr(uploaded_file, 'file_id', uploaded_file.name)
                
                if "last_loaded_file" not in st.session_state or st.session_state.last_loaded_file != file_key:
                    try:
                        import ast  # Needed to turn strings back into tables
                        df_import = pd.read_csv(uploaded_file)
                        df_import = df_import.replace({np.nan: None})
                        
                        if df_import is not None and not df_import.empty:
                            global_temp_custom = {}
                            import_pid_map = {} # Track mapped IDs to prevent row-by-row renaming

                            for index, row in df_import.iterrows():
                                # Get raw ID
                                raw_pid = str(row.get("Project_ID", curr_id)).strip()
                                key = str(row.get("Metric_Key", "")).strip()
                                val = row.get("Value", "")

                                if not key or pd.isna(val): continue 

                                # 1. Handle Project ID Mapping (Renaming if duplicate exists)
                                if raw_pid not in import_pid_map:
                                    target_pid = raw_pid
                                    if target_pid in st.session_state.projects:
                                        # Unique suffix for this file import
                                        target_pid = f"{target_pid}_imported_{file_key[:4]}"
                                    import_pid_map[raw_pid] = target_pid

                                pid = import_pid_map[raw_pid]

                                if pid not in st.session_state.projects:
                                    st.session_state.projects[pid] = {
                                        "name": f"Imported Project {pid}",
                                        "type": "Hotel", 
                                        "data": {}
                                    }
                                
                                # 2. Handle Metadata
                                if key == "proj_name":
                                    st.session_state.projects[pid]["name"] = str(val)
                                elif key == "proj_type":
                                    st.session_state.projects[pid]["type"] = str(val)
                                
                                # 3. Handle Custom Item Reconstruction
                                elif key.startswith(("input_name", "input_rate", "input_qty")):
                                    if pid not in global_temp_custom:
                                        global_temp_custom[pid] = {}
                                    try:
                                        idx = int(''.join(filter(str.isdigit, key)))
                                        if idx not in global_temp_custom[pid]:
                                            global_temp_custom[pid][idx] = {"Item Description": "", "Rate (Rp)": 0.0, "Quantity": 1.0}
                                        
                                        if "name" in key: global_temp_custom[pid][idx]["Item Description"] = str(val)
                                        elif "rate" in key: global_temp_custom[pid][idx]["Rate (Rp)"] = _safe_float(val)
                                        elif "qty" in key: global_temp_custom[pid][idx]["Quantity"] = _safe_float(val)
                                    except: continue

                                # 4. Handle Standard Metrics & Nested Tables (Area Analysis)
                                else:
                                    try:
                                        str_val = str(val).strip()

                                        # Strip surrounding quotes that CSV may add
                                        if str_val.startswith('"') and str_val.endswith('"'):
                                            str_val = str_val[1:-1]

                                        if str_val.startswith("[{") or str_val.startswith("['"): 
                                            # Try JSON first (clean), then fall back to Python literal
                                            try:
                                                st.session_state.projects[pid]["data"][key] = _json.loads(str_val)
                                            except Exception:
                                                try:
                                                    st.session_state.projects[pid]["data"][key] = ast.literal_eval(str_val)
                                                except Exception:
                                                    st.session_state.projects[pid]["data"][key] = str_val
                                        else:
                                            try:
                                                st.session_state.projects[pid]["data"][key] = _safe_float(val)
                                            except (ValueError, TypeError):
                                                st.session_state.projects[pid]["data"][key] = str(val)
                                    except (ValueError, TypeError, SyntaxError):
                                        st.session_state.projects[pid]["data"][key] = str(val)

                            # 5. Finalize Custom Item Lists
                            for pid_key, items_dict in global_temp_custom.items():
                                sorted_custom = [items_dict[i] for i in sorted(items_dict.keys())]
                                st.session_state.projects[pid_key]["data"]["smart_custom_costs"] = sorted_custom

                            # 6. CLEAR UI CACHE ANCHORS (Crucial for Area Analysis reload)
                            keys_to_clear = [k for k in st.session_state.keys() if "base_table_" in k or "area_editor_" in k]
                            for k in keys_to_clear:
                                del st.session_state[k]

                            st.session_state.last_loaded_file = file_key
                            save_data()
                            st.success(f"✅ Import Successful! New projects created to avoid overwrites.")
                            st.rerun()
                        else:
                            st.warning("⚠️ The uploaded CSV is empty.")
                    except Exception as e:
                        st.error(f"❌ Error during import: {e}")
                        
        with c2:
            st.subheader("Export")
            # --- 1. CURRENT PROJECT ONLY ---
            current_project_csv = []
            current_project_csv.append({"Project_ID": curr_id, "Metric_Key": "proj_name", "Value": curr_proj["name"]})
            current_project_csv.append({"Project_ID": curr_id, "Metric_Key": "proj_type", "Value": curr_proj["type"]})
            
            for k, v in st.session_state.projects[curr_id]["data"].items():
                if k not in ("header_info", "assumptions"):
                    serialized_v = _json.dumps(v) if isinstance(v, list) else v
                    current_project_csv.append({"Project_ID": curr_id, "Metric_Key": k, "Value": serialized_v})

            df_curr = pd.DataFrame(current_project_csv)
            csv_buffer = df_curr.to_csv(index=False).encode("utf-8")

            # --- 2. GLOBAL DATABASE ---
            all_projects_csv = []
            for pid, pdata in st.session_state.projects.items():
                all_projects_csv.append({"Project_ID": pid, "Metric_Key": "proj_name", "Value": pdata["name"]})
                all_projects_csv.append({"Project_ID": pid, "Metric_Key": "proj_type", "Value": pdata["type"]})
                for k, v in pdata["data"].items():
                    if k not in ("header_info", "assumptions"):
                        serialized_v = _json.dumps(v) if isinstance(v, list) else v
                        all_projects_csv.append({
                            "Project_ID": pid,
                            "Metric_Key": k,
                            "Value": serialized_v
                        })

            df_all = pd.DataFrame(all_projects_csv)
            csv_all_buffer = df_all.to_csv(index=False).encode("utf-8")

            # --- 3. DOWNLOAD BUTTONS ---
            st.download_button(
                label=f"Download {curr_proj['name']} only",
                data=csv_buffer,
                file_name=f"Project_{curr_id}.csv",
                mime="text/csv",
                use_container_width=True
            )

            active_file_name = st.session_state.get("loaded_snapshot_name")

            st.download_button(
                label=f"Download {active_file_name} File",
                data=csv_all_buffer,
                file_name="ProCalc_Global_Database.csv",
                mime="text/csv",
                use_container_width=True,
                type="primary",
                icon=icon_safe("download")
            )

#region --- LOGIN SCREEN AND SIDE BAR(INSIDE MAIN APP) ---
from supabase import create_client, Client

# 1. SETUP & SESSION CHECK
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

import time

def login_screen():
    st.markdown("""
        <style>
            /* Hide the default Streamlit sidebar and top header */
            [data-testid="stSidebar"] {display: none;}
            [data-testid="stHeader"] {display: none;}
            
            /* --- RESPONSIVE SPACING FIX --- */
            /* 1. Default spacing for Desktop (PC) */
            .block-container {
                padding-top: 4rem !important; 
            }
            
            /* 2. Overrides for Mobile (Screens smaller than 768px) */
            @media (max-width: 768px) {
                .block-container {
                    padding-top: 1rem !important; /* Pulls everything up on mobile */
                }
                /* Hides the empty left column on mobile so it doesn't push the form down */
                [data-testid="column"]:nth-of-type(1) {
                    display: none !important;
                }
            }
            
            /* Typography & Alignment */
            .login-header {
                text-align: center;
                margin-top: 1rem;
                margin-bottom: 1.5rem;
            }
            .login-title {
                color: #1E3A8A;
                font-size: 26px;
                font-weight: 700;
                margin-bottom: 5px;
            }
            .login-subtitle {
                color: #6B7280;
                font-size: 15px;
            }
            
            /* Form Card Styling */
            [data-testid="stForm"] {
                border: 1px solid rgba(49, 51, 63, 0.1);
                border-radius: 12px;
                padding: 2rem;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03);
                background-color: var(--background-color);
            }
            
            /* HTML Logo Styling */
            .logo-container {
                display: flex;
                justify-content: center;
                margin-bottom: 0.5rem;
            }
            .logo-container img {
                width: 45%; /* Keeps the logo scaled similarly to the old column layout */
                max-width: 200px;
            }
        </style>
    """, unsafe_allow_html=True)
        
    # 2. Adjust Layout Proportions (creates a clean, focused center column)
    col1, center_col, col3 = st.columns([1.5, 2, 1.5])
    
    with center_col:
        # 3. Nesting the logo right above the text for perfect alignment
        logo_col_1, logo_col_2, logo_col_3 = st.columns([1, 1.5, 1])
        with logo_col_2:
            st.image("https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTQWTNsgu_c6WzJAehb4zQ3qdTKNauleAXe4w&s", use_container_width=True)
            
        st.markdown("""
            <div class="login-header">
                <div class="login-title">Project Feasibility Study</div>
                <div class="login-subtitle">Please sign in to your account</div>
            </div>
        """, unsafe_allow_html=True)
        
        # 4. The Login Form
        with st.form("login_gate", clear_on_submit=False):
            # Added placeholders for better UX
            email = st.text_input("Email", placeholder="name@agungsedayu.com")
            password = st.text_input("Password", type="password", placeholder="••••••••")
            
            st.write("") # Small gap before button
            submit = st.form_submit_button("Sign In", use_container_width=True, type="primary")
            
            if submit:
                # Basic input validation
                if not email or not password:
                    st.warning("Please enter both email and password.", icon="⚠️")
                else:
                    # Added a spinner so the UI doesn't freeze during API calls
                    with st.spinner("Authenticating securely..."):
                        try:
                            res = supabase.auth.sign_in_with_password({
                                "email": email, 
                                "password": password
                            })
                            
                            supabase.postgrest.auth(res.session.access_token)
                            st.session_state.logged_in = True
                            st.session_state.user = res.user
                            st.session_state.access_token = res.session.access_token

                            # Clear projects so main_app() re-loads from Supabase
                            for key in ["projects", "storage_loaded"]:
                                if key in st.session_state:
                                    del st.session_state[key]
                            
                            # Format username to look cleaner (e.g., john.doe -> John Doe)
                            raw_username = res.user.email.split("@")[0]
                            clean_username = raw_username.replace('.', ' ').title()
                            
                            st.success(f"Identity Verified. Welcome back, **{clean_username}**! 👋")
                            time.sleep(0.8)  # Slightly faster transition
                            st.rerun()
                            
                        except Exception as e:
                            # Catch generic invalid credential messages and make them user-friendly
                            error_msg = str(e)
                            if "Invalid login credentials" in error_msg or "400" in error_msg:
                                st.error("Invalid email or password. Please try again.", icon="❌")
                            else:
                                st.error(f"Authentication error: {error_msg}", icon="🚨")

    # 5. Professional Footer
    st.markdown(f"""
        <hr style="border: none; border-top: 1px solid #E5E7EB; margin-top: 50px; margin-bottom: 20px;">
        <div style='text-align: center; color: #9CA3AF; font-size: 12px; font-family: sans-serif; line-height: 1.6;'>
            v{APP_VERSION} | &copy; 2026 QS & Procurement - ASG. All rights reserved.<br>
            <span style="letter-spacing: 1px; font-weight: 500;">INTERNAL AGUNG SEDAYU GROUP USE ONLY</span>
        </div>
    """, unsafe_allow_html=True)

# 3. THE ACTUAL APPLICATION
def main_app():
    # The 'Assembler'
    ensure_app_state_loaded()

    curr_id, curr_proj = get_current_project()

    #region --- SIDEBAR ----
    st.sidebar.title("Main Navigation")

    user = st.session_state.get("user")
    user_email = getattr(user, "email", "user@example.com")
    username = user_email.split("@")[0]
    st.sidebar.markdown(f"Hello, **{username}**!")

    page_choice = st.sidebar.radio(
        "Pilih Pekerjaan:",
        ["Start", "Area Analysis", "Cost Analysis", "Database", "Summary", "Archive"]
    )

    st.sidebar.markdown("---")

    # Always build sidebar list AFTER project repair
    curr_id, curr_proj = get_current_project()

    proj_ids = list(st.session_state.projects.keys())
    proj_labels = [
        f"{st.session_state.projects[pid]['name']} ({st.session_state.projects[pid]['type']})"
        for pid in proj_ids
    ]

    current_index = proj_ids.index(curr_id) if curr_id in proj_ids else 0
    
    # ==================================================
    # SIDEBAR CURRENT FILE / QUICK SAVE
    # ==================================================
    active_archive_id = st.session_state.get("loaded_snapshot_id")
    active_archive_name = st.session_state.get("loaded_snapshot_name")

    if active_archive_id:

        if st.sidebar.button(
            "Quick Save",
            key="sidebar_save_current_archive",
            type="primary",
            use_container_width=True,
            icon=mi("save") if "mi" in globals() else None,
            help="Overwrite the currently loaded file."
        ):
            if overwrite_current_snapshot():
                save_data()
                st.sidebar.success("File saved.")
                st.rerun()
    else:
        st.sidebar.info("No file linked yet.")
        st.sidebar.button(
            "Quick Save",
            key="sidebar_save_disabled_no_archive",
            use_container_width=True,
            disabled=True,
            icon=mi("save") if "mi" in globals() else None,
        )

    st.sidebar.markdown("---")

    # ==================================================
    # ACTIVE COMPONENT SELECTOR
    # ==================================================
    st.sidebar.subheader("Project List")

    st.sidebar.radio(
        "Active Component:",
        options=proj_labels,
        index=current_index,
        key="project_selector",
        on_change=cb_switch_project,
        label_visibility="collapsed"
    )

    curr_id, curr_proj = get_current_project()

    # ==================================================
    # SIDEBAR COMPONENT ACTION MODE
    # ==================================================
    if "sidebar_component_mode" not in st.session_state:
        st.session_state.sidebar_component_mode = None

    project_type_options = list(PROJECT_DATABASE.keys())


    # ==================================================
    # DEFAULT ACTION BUTTONS
    # ==================================================
    if st.session_state.sidebar_component_mode is None:
        c1, c2 = st.sidebar.columns(2)

        with c1:
            if st.button(
                "Add",
                key="sidebar_component_add_start",
                type="primary",
                use_container_width=True,
                icon=icon_safe("add")
            ):
                st.session_state.sidebar_component_mode = "add"
                st.rerun()

        with c2:
            if st.button(
                "Edit",
                key="sidebar_component_edit_start",
                use_container_width=True,
                icon=icon_safe("edit")
            ):
                st.session_state.sidebar_component_mode = "edit"
                st.rerun()

        if st.sidebar.button(
            "Delete",
            key="sidebar_component_delete_start",
            use_container_width=True,
            icon=icon_safe("delete")
        ):
            st.session_state.sidebar_component_mode = "delete"
            st.rerun()


    # ==================================================
    # ADD COMPONENT MODE
    # ==================================================
    elif st.session_state.sidebar_component_mode == "add":
        with st.sidebar.form("sidebar_add_component_form", clear_on_submit=False):
            st.markdown("**Add Component**")

            new_component_name = st.text_input(
                "Component Name",
                placeholder="e.g. Apartment Tower A"
            )

            new_component_type = st.selectbox(
                "Component Type",
                options=project_type_options,
                index=project_type_options.index("Hotel") if "Hotel" in project_type_options else 0
            )

            c_create, c_cancel = st.columns(2)

            with c_create:
                create_clicked = st.form_submit_button(
                    "Create",
                    type="primary",
                    use_container_width=True
                )

            with c_cancel:
                cancel_clicked = st.form_submit_button(
                    "Cancel",
                    use_container_width=True
                )

            if create_clicked:
                clean_name = new_component_name.strip()

                if clean_name == "":
                    st.warning("Please enter component name.")
                else:
                    st.session_state.proj_counter += 1
                    new_id = f"proj_{st.session_state.proj_counter}"

                    st.session_state.projects[new_id] = {
                        "name": clean_name,
                        "type": new_component_type,
                        "data": {}
                    }

                    st.session_state.current_proj_id = new_id
                    st.session_state.sidebar_component_mode = None

                    save_data()
                    st.rerun()

            if cancel_clicked:
                st.session_state.sidebar_component_mode = None
                st.rerun()


    # ==================================================
    # EDIT COMPONENT MODE
    # ==================================================
    elif st.session_state.sidebar_component_mode == "edit":
        curr_id, curr_proj = get_current_project()

        current_name = curr_proj.get("name", "")
        current_type = curr_proj.get("type", "Hotel")

        if current_type not in project_type_options:
            current_type = "Hotel"

        with st.sidebar.form("sidebar_edit_component_form", clear_on_submit=False):
            st.markdown("**Edit Component**")

            edited_component_name = st.text_input(
                "Component Name",
                value=current_name
            )

            edited_component_type = st.selectbox(
                "Component Type",
                options=project_type_options,
                index=project_type_options.index(current_type)
            )

            c_save, c_cancel = st.columns(2)

            with c_save:
                save_clicked = st.form_submit_button(
                    "Save",
                    type="primary",
                    use_container_width=True
                )

            with c_cancel:
                cancel_clicked = st.form_submit_button(
                    "Cancel",
                    use_container_width=True
                )

            if save_clicked:
                clean_name = edited_component_name.strip()

                if clean_name == "":
                    st.warning("Component name cannot be empty.")
                else:
                    old_type = st.session_state.projects[curr_id].get("type", "Hotel")

                    st.session_state.projects[curr_id]["name"] = clean_name

                    if edited_component_type != old_type:
                        st.session_state.projects[curr_id]["type"] = edited_component_type

                        # Clear old cost/input values so new database defaults can take effect
                        st.session_state.projects[curr_id]["data"] = {}

                        # Clear possible widget/session cache tied to old type
                        keys_to_clear = [
                            k for k in st.session_state.keys()
                            if str(curr_id) in str(k)
                            or "temp_spec_" in str(k)
                            or "u_fl_" in str(k)
                        ]

                        for k in keys_to_clear:
                            if k not in ["projects", "current_proj_id", "proj_counter"]:
                                del st.session_state[k]
                    else:
                        st.session_state.projects[curr_id]["type"] = edited_component_type

                    st.session_state.sidebar_component_mode = None

                    save_data()
                    st.rerun()


    # ==================================================
    # DELETE COMPONENT CONFIRMATION MODE
    # ==================================================
    elif st.session_state.sidebar_component_mode == "delete":
        curr_id, curr_proj = get_current_project()

        st.sidebar.warning(
            f"Delete **{curr_proj.get('name', 'Current Component')}**?"
        )
        st.sidebar.caption("This will remove the active project from the current file.")

        c_confirm, c_cancel = st.sidebar.columns(2)

        with c_confirm:
            if st.button(
                "Confirm",
                key="sidebar_component_delete_confirm",
                type="primary",
                use_container_width=True
            ):
                projects = st.session_state.get("projects", {})

                if not isinstance(projects, dict) or len(projects) <= 1:
                    st.sidebar.warning("At least one component must remain.")
                    st.session_state.sidebar_component_mode = None
                    repair_projects_state(save=True)
                    st.rerun()

                if curr_id in projects:
                    del projects[curr_id]

                repair_projects_state(save=True)
                st.session_state.sidebar_component_mode = None

                save_data()
                st.rerun()

        with c_cancel:
            if st.button(
                "Cancel",
                key="sidebar_component_delete_cancel",
                use_container_width=True
            ):
                st.session_state.sidebar_component_mode = None
                st.rerun()  

    st.sidebar.divider()

    if st.sidebar.button("Logout", type="primary", icon=icon_safe("logout")):
        st.session_state.logged_in = False
        st.session_state.access_token = None
        st.session_state.user = None

        for key_to_clear in [
            "projects",
            "storage_loaded",
            "report_config",
            "port_meta",
            "port_assumptions"
        ]:
            if key_to_clear in st.session_state:
                del st.session_state[key_to_clear]

        st.rerun()
        
    st.sidebar.caption(f"v{APP_VERSION} | © 2026 QS & Procurement - ASG")
    #endregion
    
    if page_choice == "Area Analysis":
        show_area_calculator()
    elif page_choice == "Database":
        show_project_database()
    elif page_choice == "Summary":
        show_portfolio_summary()
    elif page_choice == "Archive":
        show_snapshots()
    elif page_choice == "Cost Analysis":
        show_cost_estimator()
    else:
        render_feasibility_study_landing()

# 4. THE GATEKEEPER LOGIC
# Replace your entire gatekeeper section at the bottom with this:

if not st.session_state.logged_in:
    login_screen()
else:
    # Re-apply token on EVERY script run, not just after login
    token = st.session_state.get("access_token")
    if token:
        supabase.postgrest.auth(token)
    main_app()
#endregion

#region --- THANK YOU ---
# Version: 1.1.0
# Environment: Streamlit 1.56.0, Python 3.13
# for the future me or any IT person that might look at this code,
# this code is made in 2026, by a totally newbie programmer wannabe, 
# but with over 8 years of work experience and Architecture Bachelor 
# (architure as in construction, not that architecture)
# if someday this might not work, know that I (Boris Prilyan Sidabutar, B. Arch) 
# make this alone (many thanks especially to Jesus and for Gemini and Claude too)
#endregion
